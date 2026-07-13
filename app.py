import os, io, csv, zipfile, requests, threading, json, re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
from datetime import datetime, timedelta
from flask import Flask, render_template, jsonify, request, Response, stream_with_context

app = Flask(__name__)

API_KEY  = os.environ.get("CC_API_KEY", "80c5abfc86864f8d9330288c3521bb78")
BULK_BASE= "https://ccewuksprdoneregsadata1.blob.core.windows.net/data/txt"
PARTA_URL  = f"{BULK_BASE}/publicextract.charity_annual_return_parta.zip"
CHARITY_URL= f"{BULK_BASE}/publicextract.charity.zip"
CLASSIF_URL= f"{BULK_BASE}/publicextract.charity_classification.zip"
CC_API_BASE = "https://api.charitycommission.gov.uk/register/api"
AREA_URL   = f"{BULK_BASE}/publicextract.charity_area_of_operation.zip"
POLICY_URL = f"{BULK_BASE}/publicextract.charity_policy.zip"

# ─── Turso DB ─────────────────────────────────────────────────────────────────
TURSO_URL   = os.environ.get("TURSO_DATABASE_URL", "")
TURSO_TOKEN = os.environ.get("TURSO_AUTH_TOKEN", "")

# ─── OpenAI (AI pre-call insights) ────────────────────────────────────────────
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
OPENAI_MODEL   = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
# What 9 Mountains offers — editable here or overridden via the NINEM_PITCH env var.
NINEM_PITCH = os.environ.get("NINEM_PITCH",
    "9 Mountains helps UK charities strengthen and diversify their income — "
    "particularly around statutory and government funding, grants, and financial "
    "sustainability. The caller is reaching out to start a conversation about how "
    "9 Mountains could support this charity's funding and growth.")
_db_lock = threading.Lock()
_db = None

def _new_conn():
    """Create a fresh Turso client per call — pure HTTP, no Rust runtime."""
    if not TURSO_URL or not TURSO_TOKEN: return None
    try:
        import libsql_client
        # Force HTTPS - libsql-client otherwise tries wss:// which Turso rejects with 505
        url = TURSO_URL.replace("libsql://", "https://", 1)
        return libsql_client.create_client_sync(url=url, auth_token=TURSO_TOKEN)
    except Exception as e:
        print(f"[Turso] connect failed: {e}")
        return None

def db_init():
    """Create tables if they don't exist."""
    client = _new_conn()
    if not client: return
    try:
        client.execute("""CREATE TABLE IF NOT EXISTS users (
            name TEXT PRIMARY KEY,
            added_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        client.execute("""CREATE TABLE IF NOT EXISTS called_log (
            key TEXT PRIMARY KEY,
            reg_number TEXT, page TEXT, name TEXT,
            called_by TEXT, timestamp TEXT, data TEXT
        )""")
        client.execute("""CREATE TABLE IF NOT EXISTS comments (
            key TEXT PRIMARY KEY, text TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        # Comment history — every comment ever added (for viewing past notes)
        client.execute("""CREATE TABLE IF NOT EXISTS comment_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_number TEXT, page TEXT, text TEXT,
            caller TEXT, timestamp TEXT,
            synced_to_maximizer INTEGER DEFAULT 0
        )""")
        client.execute("CREATE INDEX IF NOT EXISTS idx_ch_reg ON comment_history(reg_number)")
        client.execute("""CREATE TABLE IF NOT EXISTS statuses (
            key TEXT PRIMARY KEY, status TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        client.execute("""CREATE TABLE IF NOT EXISTS saved_searches (
            name TEXT PRIMARY KEY, criteria TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        # Phone overrides — user-edited phone numbers per charity
        # (Table also holds email + website overrides — kept this name for migration compat)
        client.execute("""CREATE TABLE IF NOT EXISTS phone_overrides (
            reg_number TEXT PRIMARY KEY,
            phone TEXT,
            updated_by TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )""")
        # Idempotent column adds for email + website + contact person (no-op if already present)
        for _col in ("email","website","contact_name","contact_role","contact_phone","contact_email"):
            try: client.execute(f"ALTER TABLE phone_overrides ADD COLUMN {_col} TEXT")
            except Exception: pass
        try: client.execute("ALTER TABLE phone_overrides ADD COLUMN contacts_json TEXT")
        except Exception: pass
        # Call log — every call made through dashboard
        client.execute("""CREATE TABLE IF NOT EXISTS call_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_number TEXT, page TEXT,
            phone_used TEXT, outcome TEXT, notes TEXT,
            duration_sec INTEGER DEFAULT 0,
            caller TEXT, timestamp TEXT,
            synced_to_maximizer INTEGER DEFAULT 0
        )""")
        client.execute("CREATE INDEX IF NOT EXISTS idx_call_reg ON call_log(reg_number)")
        # Retry cooldown: when a call outcome is one that means "didn't get through"
        # (No Answer / Voicemail / Busy / Line Dropped), retry_at holds when this
        # charity becomes eligible to be called again.
        try: client.execute("ALTER TABLE call_log ADD COLUMN retry_at TEXT")
        except Exception: pass
        # Follow-up calls scheduled during a call
        client.execute("""CREATE TABLE IF NOT EXISTS follow_ups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_number TEXT, name TEXT, page TEXT,
            follow_up_at TEXT,
            caller TEXT, notes TEXT,
            created_at TEXT,
            done INTEGER DEFAULT 0,
            completed_at TEXT
        )""")
        try: client.execute("ALTER TABLE follow_ups ADD COLUMN completed_at TEXT")
        except Exception: pass
        try: client.execute("ALTER TABLE follow_ups ADD COLUMN days_late INTEGER")
        except Exception: pass
        client.execute("CREATE INDEX IF NOT EXISTS idx_fu_at ON follow_ups(follow_up_at)")
        # Persisted Daily Calling batch (a fixed set of 100, stable until regenerated)
        client.execute("""CREATE TABLE IF NOT EXISTS calling_batch (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_no INTEGER,
            reg_number TEXT,
            position INTEGER,
            completed INTEGER DEFAULT 0,
            created_at TEXT,
            category TEXT,
            active INTEGER DEFAULT 1,
            data TEXT
        )""")
        client.execute("CREATE INDEX IF NOT EXISTS idx_cb_active ON calling_batch(active)")
        client.execute("""CREATE TABLE IF NOT EXISTS newsletter_flags (
            reg_number TEXT PRIMARY KEY,
            flag INTEGER DEFAULT 0,
            updated_by TEXT,
            updated_at TEXT
        )""")
        client.execute("""CREATE TABLE IF NOT EXISTS calling_exclusions (
            reg_number TEXT PRIMARY KEY,
            excluded_by TEXT,
            excluded_at TEXT
        )""")
        # Saved, shareable column layout presets (Daily Calling / Call History)
        client.execute("""CREATE TABLE IF NOT EXISTS column_presets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            config_json TEXT,
            created_by TEXT,
            created_at TEXT,
            visibility TEXT DEFAULT 'private',
            allowed_users_json TEXT
        )""")
        # Seed default user if empty
        try:
            rs = client.execute("SELECT COUNT(*) FROM users")
            if rs.rows and rs.rows[0][0] == 0:
                client.execute("INSERT INTO users(name) VALUES(?)", ["Muhanna"])
        except: pass
        print("[Turso] Tables initialized")
    except Exception as e:
        print(f"[Turso] db_init error: {e}")
    finally:
        try: client.close()
        except: pass

def db_exec(sql, params=()):
    """Execute write with a fresh short-lived client."""
    client = _new_conn()
    if not client: return None
    try:
        client.execute(sql, list(params) if params else [])
        return True
    except Exception as e:
        print(f"[Turso] db_exec error on '{sql[:50]}': {e}")
        return None
    finally:
        try: client.close()
        except: pass

def db_query(sql, params=()):
    """Read query with a fresh short-lived client. Returns list of tuples."""
    client = _new_conn()
    if not client: return []
    try:
        rs = client.execute(sql, list(params) if params else [])
        return [tuple(r) for r in rs.rows]
    except Exception as e:
        print(f"[Turso] db_query error on '{sql[:50]}': {e}")
        return []
    finally:
        try: client.close()
        except: pass

# ─── Cache and load in-memory state from Turso ────────────────────────────────
_cache={}; _called_log={}; _comments={}; _statuses={}; _saved_searches={}
_contact_overrides={}   # {reg_number: {phone,email,website}} user-edited contact details
_newsletter={}          # {reg_number: 1} newsletter opt-in flags
_calling_excl=set()     # reg_numbers permanently excluded from calling lists
_bg_status={}; _bg_lock=threading.Lock()
_users=["Muhanna"]  # admin-managed caller list

def db_load_state():
    """Load all DB state into in-memory dicts. Clears existing state first."""
    global _users, _called_log, _comments, _statuses, _saved_searches, _contact_overrides, _newsletter, _calling_excl
    if not TURSO_URL or not TURSO_TOKEN: return

    try:
        users = db_query("SELECT name FROM users ORDER BY added_at")
        if users:
            _users = [r[0] for r in users]
        print(f"[Turso] Loaded {len(_users)} users")

        # Clear existing state so removed rows are also removed from memory
        _called_log.clear()
        _comments.clear()
        _statuses.clear()
        _saved_searches.clear()

        rows = db_query("SELECT key, reg_number, page, name, called_by, timestamp, data FROM called_log")
        for r in rows:
            extra = json.loads(r[6]) if r[6] else {}
            _called_log[r[0]] = {"reg_number": r[1], "page": r[2], "name": r[3],
                                  "called_by": r[4], "timestamp": r[5], **extra}
        print(f"[Turso] Loaded {len(_called_log)} called_log entries")

        rows = db_query("SELECT key, text FROM comments")
        for r in rows: _comments[r[0]] = r[1]
        print(f"[Turso] Loaded {len(_comments)} comments")

        rows = db_query("SELECT key, status FROM statuses")
        for r in rows: _statuses[r[0]] = r[1]
        print(f"[Turso] Loaded {len(_statuses)} statuses")

        rows = db_query("SELECT name, criteria FROM saved_searches")
        for r in rows:
            try: _saved_searches[r[0]] = json.loads(r[1])
            except: _saved_searches[r[0]] = {}
        print(f"[Turso] Loaded {len(_saved_searches)} saved searches")

        _newsletter.clear()
        try:
            rows = db_query("SELECT reg_number, flag FROM newsletter_flags WHERE flag=1")
            for r in rows: _newsletter[r[0]] = 1
            print(f"[Turso] Loaded {len(_newsletter)} newsletter flags")
        except Exception as e:
            print(f"[Turso] newsletter load skipped: {e}")

        _calling_excl.clear()
        try:
            rows = db_query("SELECT reg_number FROM calling_exclusions")
            for r in rows: _calling_excl.add(r[0])
            print(f"[Turso] Loaded {len(_calling_excl)} calling exclusions")
        except Exception as e:
            print(f"[Turso] calling_excl load skipped: {e}")

        _contact_overrides.clear()
        try:
            rows = db_query("SELECT reg_number, phone, email, website, contact_name, contact_role, contact_phone, contact_email, contacts_json FROM phone_overrides")
            for r in rows:
                d = {}
                if r[1]: d["phone"]   = r[1]
                if r[2]: d["email"]   = r[2]
                if r[3]: d["website"] = r[3]
                if len(r) > 4 and r[4]: d["name"]   = r[4]
                if len(r) > 5 and r[5]: d["role"]   = r[5]
                if len(r) > 6 and r[6]: d["cphone"] = r[6]
                if len(r) > 7 and r[7]: d["cemail"] = r[7]
                if len(r) > 8 and r[8]:
                    try: d["contacts"] = json.loads(r[8])
                    except Exception: pass
                if d: _contact_overrides[r[0]] = d
            print(f"[Turso] Loaded {len(_contact_overrides)} contact overrides")
        except Exception as e:
            # Fallback if newer columns aren't there yet (very first deploy)
            try:
                rows = db_query("SELECT reg_number, phone, email, website FROM phone_overrides")
                for r in rows:
                    d = {}
                    if r[1]: d["phone"]=r[1]
                    if r[2]: d["email"]=r[2]
                    if r[3]: d["website"]=r[3]
                    if d: _contact_overrides[r[0]] = d
            except Exception:
                rows = db_query("SELECT reg_number, phone FROM phone_overrides")
                for r in rows:
                    if r[1]: _contact_overrides[r[0]] = {"phone": r[1]}
            print(f"[Turso] Loaded {len(_contact_overrides)} contact overrides (legacy schema)")
    except Exception as e:
        print(f"[Turso] db_load_state error: {e}")

# Initialize on import - never let DB issues crash the app
try:
    db_init()
    db_load_state()
except Exception as _e:
    print(f"[Turso] FATAL during init (continuing without DB): {_e}")
    import traceback; traceback.print_exc()

def cache_get(key,max_age=180):
    if key in _cache:
        ts,val=_cache[key]
        if (datetime.utcnow()-ts).total_seconds()<max_age*60: return val
    return None
def cache_set(key,val): _cache[key]=(datetime.utcnow(),val)
def flt(v):
    try: return float(v) if v else 0.0
    except: return 0.0

def stream_zip_csv(url, needed_cols=None):
    r=requests.get(url,timeout=180,stream=True)
    r.raise_for_status()
    buf=io.BytesIO()
    for chunk in r.iter_content(chunk_size=1024*1024): buf.write(chunk)
    buf.seek(0)
    print(f"  Downloaded {buf.getbuffer().nbytes//1024}KB from {url.split('/')[-1]}")
    with zipfile.ZipFile(buf) as z:
        inner=z.namelist()[0]
        with z.open(inner) as f:
            header=f.readline().decode('utf-8',errors='replace').rstrip('\r\n')
            all_cols=[c.strip().lower() for c in header.split('\t')]
            print(f"  [stream_zip_csv] {url.split('/')[-1]} actual columns: {all_cols}")
            keep={i:c for i,c in enumerate(all_cols) if not needed_cols or c in needed_cols}
            if needed_cols and not keep:
                print(f"  [stream_zip_csv] WARNING: none of {sorted(needed_cols)} matched the actual columns above")
            count=0
            for raw in f:
                try:
                    fields=raw.decode('utf-8',errors='replace').rstrip('\r\n').split('\t')
                    row={c:(fields[i].strip() if i<len(fields) else '') for i,c in keep.items()}
                    row={k:('' if v.lower() in ('nan','none') else v) for k,v in row.items()}
                    yield row; count+=1
                except: continue
    print(f"  Streamed {count} rows")

def classify(ti,stat,pct,cont):
    l=[]
    if 250_000<=ti<=3_000_000 and 50_000<=stat<=300_000 and 10<=pct<=40: l.append("Best Immediate")
    if 150_000<=ti<=1_500_000 and 10_000<=stat<=100_000: l.append("Readiness Package")
    if 500_000<=ti<=5_000_000 and 100_000<=cont<=500_000 and (cont/ti*100<35 if ti>0 else False): l.append("Contract Growth")
    if 500_000<=ti<=5_000_000 and stat>=200_000 and 40<=pct<=70: l.append("Retention/Replacement")
    return l

def build_flags(ti,te,stat,pstat,pgc,cont):
    f=[]
    if ti>0 and te>=ti*0.95: f.append({"label":"Spend≥Income","cls":"red"})
    if pstat>0 and stat<pstat*0.9: f.append({"label":"Stat↓10%+","cls":"orange"})
    if pgc>0 and cont==0: f.append({"label":"Lost contracts","cls":"red"})
    return f

def bg_run(key,fn):
    with _bg_lock:
        if _bg_status.get(key)=="loading": return
        _bg_status[key]="loading"
    def _w():
        try:
            result=fn(); cache_set(key,result); print(f"BG {key}: done")
        except Exception as e:
            import traceback; print(f"BG {key} ERROR: {traceback.format_exc()}")
        finally:
            with _bg_lock: _bg_status[key]="idle"
    threading.Thread(target=_w,daemon=True).start()

def bg_check(key,empty):
    cached=cache_get(key,max_age=60)
    if cached: return jsonify(cached)
    with _bg_lock: s=_bg_status.get(key,"idle")
    if s=="loading": return jsonify({"status":"loading","message":"Loading…",**empty}),202
    return None

def compute_prospects():
    PC={"registered_charity_number","income_from_government_grants",
        "income_from_government_contracts","total_gross_income",
        "total_gross_expenditure","fin_period_end_date"}
    CC={"registered_charity_number","charity_name","charity_contact_web",
        "charity_registration_status","charity_contact_phone","charity_contact_email",
        "charity_contact_address3","charity_contact_address4"}
    latest={}; prev={}
    for row in stream_zip_csv(PARTA_URL,PC):
        reg=row.get("registered_charity_number","").strip()
        if not reg: continue
        yr=row.get("fin_period_end_date","")[:10]
        if reg not in latest or yr>latest[reg].get("fin_period_end_date",""):
            if reg in latest: prev[reg]=latest[reg]
            latest[reg]=row
        elif reg not in prev or yr>prev[reg].get("fin_period_end_date",""):
            prev[reg]=row

    # PRE-FILTER: figure out which charities pass financial criteria BEFORE loading contacts
    # This saves ~200MB by only storing contact info for ~5K prospects, not all 200K charities
    prospect_regs = set()
    for reg, row in latest.items():
        ti=flt(row.get("total_gross_income"))
        if ti<150_000 or ti>5_000_000: continue
        gg=flt(row.get("income_from_government_grants"))
        gc=flt(row.get("income_from_government_contracts"))
        stat=gg+gc; pct=round(stat/ti*100,1) if ti>0 else 0.0
        if classify(ti,stat,pct,gc):
            prospect_regs.add(reg)
            prospect_regs.add(reg.lstrip("0"))

    # Only load contact info for prospects (much smaller memory footprint)
    contacts={}
    for row in stream_zip_csv(CHARITY_URL,CC):
        reg=row.get("registered_charity_number","").strip()
        if reg in prospect_regs and row.get("charity_name",""):
            contacts[reg]={"name":row.get("charity_name",""),
                           "website":row.get("charity_contact_web",""),
                           "phone":row.get("charity_contact_phone",""),
                           "email":row.get("charity_contact_email",""),
                           "town":row.get("charity_contact_address3",""),
                           "county":row.get("charity_contact_address4",""),
                           "status":row.get("charity_registration_status","")}

    results={"best":[],"readiness":[],"contract":[],"retention":[]}
    for reg,row in latest.items():
        ti=flt(row.get("total_gross_income"))
        if ti<150_000 or ti>5_000_000: continue
        te=flt(row.get("total_gross_expenditure"))
        gg=flt(row.get("income_from_government_grants"))
        gc=flt(row.get("income_from_government_contracts"))
        stat=gg+gc; pct=round(stat/ti*100,1) if ti>0 else 0.0
        yr=row.get("fin_period_end_date","")[:4]
        pr=prev.get(reg,{}); pgg=flt(pr.get("income_from_government_grants")); pgc=flt(pr.get("income_from_government_contracts")); pstat=pgg+pgc
        lists=classify(ti,stat,pct,gc)
        if not lists: continue
        ct=contacts.get(reg,{}) or contacts.get(reg.lstrip("0"),{}) or {}
        fin={"total_income":ti,"total_expenditure":te,"gov_grants":gg,"gov_contracts":gc,
             "statutory":stat,"stat_pct":pct,"prev_grants":pgg,"prev_contracts":pgc,
             "prev_statutory":pstat,"fin_year":yr,"flags":build_flags(ti,te,stat,pstat,pgc,gc)}
        obj={"reg_number":reg,"name":ct.get("name",""),"website":ct.get("website",""),
             "phone":ct.get("phone",""),"email":ct.get("email",""),
             "town":ct.get("town",""),"county":ct.get("county",""),
             "removed":str(ct.get("status","")).upper().strip() in ("REMOVED","D"),
             "financials":fin,"lists":lists}
        if "Best Immediate" in lists: results["best"].append(obj)
        if "Readiness Package" in lists: results["readiness"].append(obj)
        if "Contract Growth" in lists: results["contract"].append(obj)
        if "Retention/Replacement" in lists: results["retention"].append(obj)
    return {"lists":results,"counts":{k:len(v) for k,v in results.items()},
            "total":sum(len(v) for v in results.values()),
            "generated":datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}

def compute_new(days):
    cutoff=(datetime.utcnow()-timedelta(days=int(days))).strftime("%Y-%m-%d")
    COLS={"registered_charity_number","charity_name","date_of_registration",
          "charity_registration_status","charity_contact_web",
          "charity_contact_phone","charity_contact_email",
          "charity_contact_address3","charity_contact_address4"}
    results=[]
    for row in stream_zip_csv(CHARITY_URL,COLS):
        if row.get("charity_registration_status","").upper().strip() not in ("REGISTERED","R"): continue
        dt=row.get("date_of_registration","")[:10]
        if dt>=cutoff:
            results.append({"reg_number":row.get("registered_charity_number",""),
                            "name":row.get("charity_name",""),"date_registered":dt,
                            "website":row.get("charity_contact_web",""),
                            "phone":row.get("charity_contact_phone",""),
                            "email":row.get("charity_contact_email",""),
                            "town":row.get("charity_contact_address3",""),
                            "county":row.get("charity_contact_address4",""),
                            "status":"Active"})
    results.sort(key=lambda x:x.get("date_registered",""),reverse=True)
    return {"charities":results,"count":len(results),"days":days,"cutoff":cutoff}

def compute_late():
    today=datetime.utcnow().date()
    PC={"registered_charity_number","ar_due_date","latest_fin_period_submitted_ind",
        "total_gross_income","fin_period_end_date"}
    NC={"registered_charity_number","charity_name","charity_contact_phone",
        "charity_contact_email","charity_contact_address3","charity_contact_address4",
        "charity_contact_web","charity_registration_status"}
    latest={}
    for row in stream_zip_csv(PARTA_URL,PC):
        reg=row.get("registered_charity_number","").strip()
        if not reg: continue
        yr=row.get("fin_period_end_date","")[:10]
        if reg not in latest or yr>latest[reg].get("fin_period_end_date",""): latest[reg]=row

    # PRE-FILTER: determine which charities are late, then only load contacts for those
    late_regs = set()
    late_data = {}
    for reg,row in latest.items():
        if row.get("latest_fin_period_submitted_ind","").upper().strip() not in ("FALSE","0","NO","N"): continue
        due_str=row.get("ar_due_date","")[:10]
        if not due_str: continue
        try: due_d=datetime.strptime(due_str,"%Y-%m-%d").date()
        except: continue
        if due_d>=today: continue
        ml=round((today-due_d).days/30.4,1)
        late_regs.add(reg)
        late_data[reg]={"due_date":str(due_d),"months_late":ml,
                        "latest_income":flt(row.get("total_gross_income",""))}

    # Only load contact info for late charities (much smaller dict)
    contacts={}
    for row in stream_zip_csv(CHARITY_URL,NC):
        reg=row.get("registered_charity_number","").strip()
        if reg in late_regs:
            contacts[reg]={"name":row.get("charity_name",""),
                           "phone":row.get("charity_contact_phone",""),
                           "email":row.get("charity_contact_email",""),
                           "town":row.get("charity_contact_address3",""),
                           "county":row.get("charity_contact_address4",""),
                           "website":row.get("charity_contact_web",""),
                           "status":row.get("charity_registration_status","")}

    results=[]
    for reg, d in late_data.items():
        ct=contacts.get(reg,{})
        results.append({"reg_number":reg,"name":ct.get("name",reg) or reg,
                        "due_date":d["due_date"],"months_late":d["months_late"],
                        "latest_income":d["latest_income"],
                        "phone":ct.get("phone",""),"email":ct.get("email",""),
                        "town":ct.get("town",""),"county":ct.get("county",""),
                        "website":ct.get("website",""),
                        "removed":str(ct.get("status","")).upper().strip() in ("REMOVED","D")})
    results.sort(key=lambda x:x.get("months_late",0),reverse=True)
    o12=len([r for r in results if r["months_late"]>12])
    b612=len([r for r in results if 6<=r["months_late"]<=12])
    u6=len([r for r in results if r["months_late"]<6])
    return {"charities":results,"count":len(results),"over_12":o12,"bet_6_12":b612,"under_6":u6}

def compute_old():
    today=datetime.utcnow().date(); cutoff_yr=today.year-40
    PC={"registered_charity_number","income_from_government_grants",
        "income_from_government_contracts","total_gross_income","fin_period_end_date"}
    CC={"registered_charity_number","charity_name","date_of_registration",
        "charity_registration_status","charity_contact_web",
        "charity_contact_phone","charity_contact_email",
        "charity_contact_address3","charity_contact_address4"}
    pl={}
    for row in stream_zip_csv(PARTA_URL,PC):
        reg=row.get("registered_charity_number","").strip()
        if not reg: continue
        yr=row.get("fin_period_end_date","")[:10]
        if reg not in pl or yr>pl[reg].get("fin_period_end_date",""): pl[reg]=row
    results=[]
    for row in stream_zip_csv(CHARITY_URL,CC):
        if row.get("charity_registration_status","").upper().strip() not in ("REGISTERED","R"): continue
        dt=row.get("date_of_registration","")[:10]
        if not dt or len(dt)<4: continue
        try: ry=int(dt[:4])
        except: continue
        if ry>cutoff_yr: continue
        reg=row.get("registered_charity_number","").strip()
        if not reg: continue
        fin=pl.get(reg,{})
        if flt(fin.get("income_from_government_grants"))+flt(fin.get("income_from_government_contracts"))>0: continue
        web=row.get("charity_contact_web","")
        results.append({"reg_number":reg,"name":row.get("charity_name","") or reg,
                        "website":web if web not in ("nan","None") else "",
                        "phone":row.get("charity_contact_phone",""),
                        "email":row.get("charity_contact_email",""),
                        "town":row.get("charity_contact_address3",""),
                        "county":row.get("charity_contact_address4",""),
                        "date_registered":dt,"age_years":today.year-ry,
                        "total_income":flt(fin.get("total_gross_income","")),
                        "fin_year":fin.get("fin_period_end_date","")[:4]})
    results.sort(key=lambda x:x.get("date_registered",""))
    return {"charities":results,"count":len(results)}

@app.route("/")
def index(): return render_template("index.html")

@app.route("/api/rc/config")
def rc_config():
    """Expose non-secret RC config to the frontend for the Embeddable widget."""
    return jsonify({
        "client_id":  RC_EMBEDDABLE_CLIENT_ID,
        "server_url": RC_SERVER_URL,
        "configured": bool(RC_EMBEDDABLE_CLIENT_ID)
    })

@app.route("/healthz")
def healthz():
    """Lightweight keep-alive endpoint for an external uptime pinger (no heavy work)."""
    return "ok", 200

@app.route("/api/prospects")
def api_prospects():
    r=bg_check("prospects",{"lists":{"best":[],"readiness":[],"contract":[],"retention":[]},"counts":{},"total":0})
    if r: return r
    bg_run("prospects",compute_prospects)
    return jsonify({"status":"loading","message":"Loading charity data from Charity Commission…",
                    "lists":{"best":[],"readiness":[],"contract":[],"retention":[]},"counts":{},"total":0}),202

@app.route("/api/new_charities")
def api_new_charities():
    days=request.args.get("days","30"); key=f"new_{days}"
    r=bg_check(key,{"charities":[],"count":0,"days":days,"cutoff":""})
    if r: return r
    bg_run(key,lambda:compute_new(days))
    return jsonify({"status":"loading","message":"Loading newly registered charities…",
                    "charities":[],"count":0,"days":days,"cutoff":""}),202

@app.route("/api/late_accounts")
def api_late_accounts():
    r=bg_check("late",{"charities":[],"count":0,"over_12":0,"bet_6_12":0,"under_6":0})
    if r: return r
    bg_run("late",compute_late)
    return jsonify({"status":"loading","message":"Loading late accounts…",
                    "charities":[],"count":0,"over_12":0,"bet_6_12":0,"under_6":0}),202

@app.route("/api/old_charities")
def api_old_charities():
    r=bg_check("old_charities",{"charities":[],"count":0})
    if r: return r
    bg_run("old_charities",compute_old)
    return jsonify({"status":"loading","message":"Loading established charities…",
                    "charities":[],"count":0}),202

@app.route("/api/search")
def api_search():
    term=request.args.get("q","").strip()
    if not term: return jsonify([])
    COLS={"registered_charity_number","charity_name","charity_registration_status",
          "charity_contact_web","charity_contact_phone","charity_contact_email",
          "charity_contact_address3","charity_contact_address4","date_of_removal",
          "charity_activities","organisation_number","charity_company_registration_number",
          "linked_charity_number","date_of_registration","charity_type",
          "latest_acc_fin_period_start_date","latest_acc_fin_period_end_date",
          "latest_expenditure","charity_gift_aid","charity_has_land"}
    results=[]; is_num=term.isdigit()
    for row in stream_zip_csv(CHARITY_URL,COLS):
        reg=row.get("registered_charity_number","").strip()
        name=row.get("charity_name","")
        if is_num:
            if reg!=term: continue
        else:
            if term.upper() not in name.upper(): continue
        rem=row.get("date_of_removal","")
        results.append({"reg_number":reg,"name":name,"status":"Removed" if rem else "Active",
                        "phone":row.get("charity_contact_phone",""),"email":row.get("charity_contact_email",""),
                        "town":row.get("charity_contact_address3",""),"county":row.get("charity_contact_address4",""),
                        "website":row.get("charity_contact_web",""),"financials":{},"lists":[],
                        "activities":row.get("charity_activities",""),
                        "organisation_number":row.get("organisation_number",""),
                        "company_number":row.get("charity_company_registration_number",""),
                        "charity_suffix":row.get("linked_charity_number","") or "0",
                        "date_registered":row.get("date_of_registration","")[:10],
                        "charity_type":row.get("charity_type",""),
                        "fin_year_start":row.get("latest_acc_fin_period_start_date","")[:10],
                        "fin_year_end":row.get("latest_acc_fin_period_end_date","")[:10],
                        "latest_expenditure":row.get("latest_expenditure",""),
                        "gift_aid":row.get("charity_gift_aid",""),
                        "has_land":row.get("charity_has_land","")})
        if len(results)>=25: break
    return jsonify(results)

def _run_advanced_search(criteria, limit=50000):
    """Core advanced-search logic, usable both by the /api/advanced_search endpoint
    and by anything that needs to run a SAVED search live (e.g. the Daily Calling
    'Charity Search (saved)' source), since saved searches only store the filter
    criteria, not stale results."""
    nq=str(criteria.get("name","")).upper().strip()
    rq=str(criteria.get("reg_number","")).strip()
    pq=str(criteria.get("postcode","")).upper().strip()
    cq=str(criteria.get("county","")).upper().strip()
    wq=str(criteria.get("what","")).upper().strip()
    sq=str(criteria.get("reg_status","registered")).strip()
    imin=flt(criteria.get("inc_min","")); imax=flt(criteria.get("inc_max","")) or float("inf")
    df=str(criteria.get("reg_date_from","")).strip(); dt=str(criteria.get("reg_date_to","")).strip()
    tq=str(criteria.get("charity_type","")).upper().strip()
    limit = max(1, min(int(limit or 50000), 200000))
    COLS={"registered_charity_number","charity_name","charity_registration_status",
          "date_of_registration","date_of_removal","charity_contact_web",
          "charity_contact_phone","charity_contact_email",
          "charity_contact_address3","charity_contact_address4",
          "charity_contact_postcode","charity_activities","charity_type","latest_income",
          "organisation_number","charity_company_registration_number","linked_charity_number",
          "latest_acc_fin_period_start_date","latest_acc_fin_period_end_date",
          "latest_expenditure","charity_gift_aid","charity_has_land"}
    results=[]
    truncated=False
    for row in stream_zip_csv(CHARITY_URL,COLS):
        st=row.get("charity_registration_status","").upper().strip()
        if sq=="registered" and st not in ("REGISTERED","R"): continue
        if sq=="removed" and st not in ("REMOVED","D"): continue
        name=row.get("charity_name","")
        if nq and nq not in name.upper(): continue
        reg=row.get("registered_charity_number","").strip()
        if rq and rq!=reg: continue
        pc=row.get("charity_contact_postcode","").upper()
        if pq and not pc.startswith(pq): continue
        county=row.get("charity_contact_address4","").upper()
        if cq and cq not in county: continue
        what=row.get("charity_activities","").upper()
        if wq and wq not in what: continue
        ctype=row.get("charity_type","").upper()
        if tq and tq not in ctype: continue
        inc=flt(row.get("latest_income",""))
        if imin and inc<imin: continue
        if imax<float("inf") and inc>imax: continue
        dtreg=row.get("date_of_registration","")[:10]
        if df and dtreg<df: continue
        if dt and dtreg>dt: continue
        results.append({"reg_number":reg,"name":name,"status":st,"latest_income":inc,
                        "date_registered":dtreg,"website":row.get("charity_contact_web",""),
                        "phone":row.get("charity_contact_phone",""),"email":row.get("charity_contact_email",""),
                        "town":row.get("charity_contact_address3",""),"county":row.get("charity_contact_address4",""),
                        "postcode":pc,"activities":row.get("charity_activities",""),
                        "charity_type":row.get("charity_type",""),"financials":{},"lists":[],
                        "organisation_number":row.get("organisation_number",""),
                        "company_number":row.get("charity_company_registration_number",""),
                        "charity_suffix":row.get("linked_charity_number","") or "0",
                        "fin_year_start":row.get("latest_acc_fin_period_start_date","")[:10],
                        "fin_year_end":row.get("latest_acc_fin_period_end_date","")[:10],
                        "latest_expenditure":row.get("latest_expenditure",""),
                        "gift_aid":row.get("charity_gift_aid",""),
                        "has_land":row.get("charity_has_land","")})
        if len(results)>=limit:
            truncated=True
            break
    return results, truncated

@app.route("/api/advanced_search")
def advanced_search():
    criteria = {
        "name": request.args.get("name",""), "reg_number": request.args.get("reg_number",""),
        "postcode": request.args.get("postcode",""), "county": request.args.get("county",""),
        "what": request.args.get("what",""), "reg_status": request.args.get("reg_status","registered"),
        "inc_min": request.args.get("inc_min",""), "inc_max": request.args.get("inc_max",""),
        "reg_date_from": request.args.get("reg_date_from",""), "reg_date_to": request.args.get("reg_date_to",""),
        "charity_type": request.args.get("charity_type",""),
    }
    try: limit = max(1, min(int(request.args.get("limit","50000")), 200000))
    except: limit = 50000
    results, truncated = _run_advanced_search(criteria, limit)
    return jsonify({"charities":results,"count":len(results),"truncated":truncated,"limit":limit})

@app.route("/api/search/enrich_export", methods=["POST"])
def api_search_enrich_export():
    """Look up What/Who/How, Region/Local Authority/Country, and Policy for a
    bounded set of charities, for the Charity Search CSV export ONLY. This is
    deliberately separate from the live search endpoints — it only runs when the
    user clicks Download CSV, and only for the specific charities being exported,
    capped at EXPORT_ENRICH_MAX to keep memory and response time bounded."""
    data = request.get_json(force=True) or {}
    regs = [str(r).strip() for r in (data.get("regs") or []) if str(r).strip()]
    capped = len(regs) > EXPORT_ENRICH_MAX
    regs = regs[:EXPORT_ENRICH_MAX]
    if not regs:
        return jsonify({"enriched": {}, "capped": False, "count": 0})
    enriched = enrich_regs_for_export(regs)
    return jsonify({"enriched": enriched, "capped": capped, "count": len(regs),
                    "max": EXPORT_ENRICH_MAX})

@app.route("/api/mark_called",methods=["POST"])
def mark_called():
    data=request.json or {}; reg=str(data.get("reg_number","")).strip(); page=str(data.get("page","")).strip()
    name=str(data.get("name","")).strip()
    caller=str(data.get("caller","")).strip() or "Unknown"
    if not reg or not page: return jsonify({"ok":False}),400
    key=f"{page}|{reg}"
    if key in _called_log:
        del _called_log[key]
        db_exec("DELETE FROM called_log WHERE key=?", (key,))
        return jsonify({"ok":True,"marked":False})
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    _called_log[key]={"reg_number":reg,"name":name,"page":page,"called_by":caller,"timestamp":ts}
    db_exec("INSERT OR REPLACE INTO called_log(key,reg_number,page,name,called_by,timestamp,data) VALUES(?,?,?,?,?,?,?)",
            (key, reg, page, name, caller, ts, "{}"))
    return jsonify({"ok":True,"marked":True})

@app.route("/api/comment",methods=["POST"])
def save_comment():
    data = request.json or {}
    reg  = str(data.get("reg_number","")).strip()
    page = str(data.get("page","")).strip()
    text = str(data.get("comment","")).strip()[:400]
    caller = str(data.get("caller","")).strip() or "Unknown"
    if not reg or not page: return jsonify({"ok":False}),400
    key = f"{page}|{reg}"

    # Save latest in comments (for dashboard display)
    _comments[key] = text
    db_exec("INSERT OR REPLACE INTO comments(key,text) VALUES(?,?)", (key, text))

    # Also save to history if non-empty
    synced = 0
    if text:
        ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        # Push to Maximizer Note if entry exists
        try:
            found = mx_find_by_org_number(reg)
            if found and found.get("key"):
                resp = mx_create_note(found["key"], text, caller)
                if resp and resp.get("Code") == 0:
                    synced = 1
        except Exception as e:
            print(f"  Note sync error: {e}")
        db_exec("INSERT INTO comment_history(reg_number,page,text,caller,timestamp,synced_to_maximizer) VALUES(?,?,?,?,?,?)",
                (reg, page, text, caller, ts, synced))
    return jsonify({"ok":True, "synced_to_maximizer": synced})

# ── Phone overrides ───────────────────────────────────────────────────────────
@app.route("/api/phone", methods=["GET","POST"])
def phone_override():
    """GET ?reg=X → returns override; POST → saves override. Preserves email/website."""
    if request.method == "GET":
        reg = str(request.args.get("reg","")).strip()
        if not reg: return jsonify({"phone": ""})
        rows = db_query("SELECT phone FROM phone_overrides WHERE reg_number=?", (reg,))
        return jsonify({"reg_number": reg, "phone": rows[0][0] if rows else ""})
    # POST → route through the unified contact endpoint logic
    data = request.json or {}
    return contact_override.__wrapped__() if False else _save_contact(
        reg=str(data.get("reg_number","")).strip(),
        field="phone",
        val=str(data.get("phone","")).strip(),
        caller=str(data.get("caller","")).strip() or "Unknown",
    )

def _save_contact(reg, field, val, caller):
    """Internal helper used by /api/phone and /api/contact.
    Handles phone/email/website (Maximizer field updates) and name/role
    (contact person — stored locally + pushed to Maximizer as a Note)."""
    if not reg:
        return jsonify({"ok":False, "error":"Missing reg_number"}), 400
    if field not in CONTACT_LIMITS:
        return jsonify({"ok":False, "error":"Bad field"}), 400
    val = (val or "")[:CONTACT_LIMITS[field]]
    existing = db_query("SELECT phone,email,website,contact_name,contact_role,contact_phone,contact_email FROM phone_overrides WHERE reg_number=?", (reg,))
    e0 = existing[0] if existing else ("","","","","","","")
    def _g(i): return (e0[i] if len(e0)>i else "") or ""
    cur = {"phone":_g(0),"email":_g(1),"website":_g(2),"name":_g(3),"role":_g(4),"cphone":_g(5),"cemail":_g(6)}
    cur[field] = val
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    if not any(cur.values()):
        db_exec("DELETE FROM phone_overrides WHERE reg_number=?", (reg,))
        _contact_overrides.pop(reg, None)
    else:
        db_exec("""INSERT OR REPLACE INTO phone_overrides
                   (reg_number,phone,email,website,contact_name,contact_role,contact_phone,contact_email,updated_by,updated_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (reg, cur["phone"], cur["email"], cur["website"], cur["name"], cur["role"], cur["cphone"], cur["cemail"], caller, ts))
        _contact_overrides[reg] = {k:v for k,v in cur.items() if v}
    synced = False
    try:
        found = mx_find_by_org_number(reg)
        if found and found.get("key"):
            if field in MX_CONTACT_FIELD:
                payload = {"Key": found["key"], MX_CONTACT_FIELD[field]: val}
                r = mx_write_update(payload)
                if r and r.get("Code") == 0: synced = True
            elif field in ("name", "role", "cphone", "cemail"):
                # No safe contact-person fields on a Company entry — record as a Note
                nm = cur["name"]; rl = cur["role"]; cp = cur["cphone"]; ce = cur["cemail"]
                parts = ["👤 Contact: " + (nm or "(name not set)") + (f" — {rl}" if rl else "")]
                if cp: parts.append("☎ " + cp)
                if ce: parts.append("✉ " + ce)
                note = "  |  ".join(parts)
                resp = mx_create_note(found["key"], note, caller, timestamp=ts)
                if resp and resp.get("Code") == 0: synced = True
    except Exception as e:
        print(f"  contact→mx update error: {e}")
    return jsonify({"ok": True, "phone": val if field=="phone" else None,
                    "field": field, "value": val, "synced_to_maximizer": synced})

@app.route("/api/phones/bulk")
def phones_bulk():
    """Return all phone overrides as a {reg_number: phone} dict for frontend caching."""
    rows = db_query("SELECT reg_number, phone FROM phone_overrides")
    return jsonify({r[0]: r[1] for r in rows if r[1]})

# ── Contact overrides (phone / email / website) — manual edits per charity ────
MX_CONTACT_FIELD = {"phone": "Phone1", "email": "Email1", "website": "WebSite"}
CONTACT_LIMITS   = {"phone": 30, "email": 100, "website": 200, "name": 80, "role": 80, "cphone": 40, "cemail": 120}

@app.route("/api/contacts/bulk")
def contacts_bulk():
    """All contact overrides as {reg: {phone,email,website}} for frontend bootstrap."""
    return jsonify(_contact_overrides)

@app.route("/api/contacts/save", methods=["POST"])
def contacts_save():
    """Save up to 3 manually-entered contacts for a charity."""
    data = request.json or {}
    reg = str(data.get("reg_number","")).strip()
    caller = str(data.get("caller","")).strip() or "Manual"
    raw = data.get("contacts") or []
    if not reg:
        return jsonify({"ok": False, "error": "Missing reg_number"}), 400
    contacts = []
    for c in raw[:3]:
        cc = {"name":  str(c.get("name") or "").strip()[:CONTACT_LIMITS["name"]],
              "role":  str(c.get("role") or "").strip()[:CONTACT_LIMITS["role"]],
              "phone": str(c.get("phone") or "").strip()[:CONTACT_LIMITS["cphone"]],
              "email": str(c.get("email") or "").strip()[:CONTACT_LIMITS["cemail"]]}
        if any(cc.values()): contacts.append(cc)
    top = contacts[0] if contacts else {"name":"","role":"","phone":"","email":""}
    existing = db_query("SELECT phone,email,website FROM phone_overrides WHERE reg_number=?", (reg,))
    e0 = existing[0] if existing else ("","","")
    def _g(i): return (e0[i] if len(e0)>i else "") or ""
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db_exec("""INSERT OR REPLACE INTO phone_overrides
               (reg_number,phone,email,website,contact_name,contact_role,contact_phone,contact_email,contacts_json,updated_by,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (reg, _g(0), _g(1), _g(2), top["name"], top["role"], top["phone"], top["email"],
             json.dumps(contacts), caller, ts))
    ov = _contact_overrides.get(reg, {}) or {}
    if top["name"]:  ov["name"]   = top["name"]
    if top["role"]:  ov["role"]   = top["role"]
    if top["phone"]: ov["cphone"] = top["phone"]
    if top["email"]: ov["cemail"] = top["email"]
    ov["contacts"] = contacts
    _contact_overrides[reg] = ov
    # Sync a note to Maximizer listing all contacts
    try:
        found = mx_find_by_org_number(reg)
        if found and found.get("key") and contacts:
            lines = ["👥 Contacts (manual entry):"]
            for i,c in enumerate(contacts,1):
                seg = f"{i}. {c['name'] or '(no name)'}" + (f" — {c['role']}" if c['role'] else "")
                if c["phone"]: seg += f"  ☎ {c['phone']}"
                if c["email"]: seg += f"  ✉ {c['email']}"
                lines.append(seg)
            mx_create_note(found["key"], "  |  ".join(lines), caller, timestamp=ts)
    except Exception as e:
        print(f"  contacts_save→mx note error: {e}")
    return jsonify({"ok": True, "contacts": contacts})

@app.route("/api/state/bulk")
def state_bulk():
    """Saved statuses + called state for frontend bootstrap (persist across reloads)."""
    called_out = {}
    for k, log in _called_log.items():
        called_out[k] = {"caller": (log.get("called_by") or ""),
                         "ts": log.get("timestamp", "")}
    return jsonify({"statuses": _statuses, "called": called_out})

def _fmt_money(v):
    try:
        n = float(str(v).replace(",", "").replace("£", "").strip())
        if n >= 1_000_000: return f"£{n/1_000_000:.1f}m"
        if n >= 1_000:     return f"£{n/1_000:.0f}k"
        return f"£{n:.0f}"
    except: return str(v or "")

def _build_charity_profile_text(c, comments, calls, transcripts=None):
    """Assemble everything we know about a charity into a readable profile for the AI."""
    L = []
    L.append(f"Name: {c.get('name','')}")
    L.append(f"Charity Commission reg number: {c.get('reg_number','')}")
    if c.get("what"):  L.append(f"What they do (purposes): {c.get('what')}")
    if c.get("who"):   L.append(f"Who they help: {c.get('who')}")
    if c.get("how"):   L.append(f"How they work: {c.get('how')}")
    if c.get("region"): L.append(f"Region: {c.get('region')}")
    if c.get("local_authority"): L.append(f"Local authority: {c.get('local_authority')}")
    city = c.get("city") or c.get("county") or ""
    if city: L.append(f"Location: {city} {c.get('postcode','')}".strip())
    # Financials
    fin = []
    if c.get("latest_income"):      fin.append(f"total income {_fmt_money(c.get('latest_income'))}")
    elif c.get("income"):           fin.append(f"total income {_fmt_money(c.get('income'))}")
    if c.get("latest_expenditure"): fin.append(f"expenditure {_fmt_money(c.get('latest_expenditure'))}")
    if c.get("statutory_income") or c.get("statutory"):
        fin.append(f"statutory income {_fmt_money(c.get('statutory_income') or c.get('statutory'))}")
    if c.get("govt_grants") or c.get("government_grants"):
        fin.append(f"government grants {_fmt_money(c.get('govt_grants') or c.get('government_grants'))}")
    if c.get("govt_contracts") or c.get("government_contracts"):
        fin.append(f"government contracts {_fmt_money(c.get('govt_contracts') or c.get('government_contracts'))}")
    if c.get("statutory_pct") or c.get("stat_pct"):
        fin.append(f"statutory income is {c.get('statutory_pct') or c.get('stat_pct')}% of total")
    if fin: L.append("Financials: " + ", ".join(fin))
    if c.get("fin_year_end") or c.get("fin_year"):
        L.append(f"Latest financial year: {c.get('fin_year_end') or c.get('fin_year')}")
    if c.get("income_trend") or c.get("trend"):
        L.append(f"Income trend vs prior year: {c.get('income_trend') or c.get('trend')}")
    if c.get("website"): L.append(f"Website: {c.get('website')}")
    if c.get("email"):   L.append(f"Email on file: {c.get('email')}")
    if c.get("phone"):   L.append(f"Phone on file: {c.get('phone')}")
    if comments:
        L.append("\nPrevious dashboard notes/comments:")
        for cm in comments[-8:]: L.append(f"  - {cm}")
    if calls:
        L.append("\nPrevious call history:")
        for cl in calls[-8:]: L.append(f"  - {cl}")
    if transcripts:
        L.append("\nPREVIOUS CALL TRANSCRIPTS (use these for continuity — pick up where the last conversation left off):")
        for ts, cname, body in transcripts:
            L.append(f"\n--- Call on {ts} (caller: {cname}) ---")
            L.append(body)
    return "\n".join(L)

def _trim_transcript(text, limit=2200):
    """Trim a long transcript while keeping the opening context and the close
    (where outcomes / next steps usually are)."""
    body = str(text).replace("📝 Call transcript:", "").strip()
    if len(body) <= limit:
        return body
    head = body[:800].rsplit(" ", 1)[0]
    tail = body[-1200:].split(" ", 1)[-1]
    return head + " …[middle of call omitted]… " + tail

@app.route("/api/ai/coach", methods=["POST"])
def ai_coach():
    """Generate AI performance coaching for a caller based on their call history in a date range."""
    if not OPENAI_API_KEY:
        return jsonify({"ok": False,
            "error": "AI is not configured. Add the OPENAI_API_KEY environment variable in Render."}), 400
    data = request.json or {}
    callers = [c.strip() for c in (data.get("callers") or []) if c.strip()]
    date_from = str(data.get("from","")).strip()
    date_to   = str(data.get("to","")).strip()
    if not date_from or not date_to:
        return jsonify({"ok": False, "error": "Missing date range"}), 400
    ts_from = date_from + " 00:00:00"; ts_to = date_to + " 23:59:59"

    # Build caller filter
    if len(callers)==1:
        cf = "AND caller=?"; cp = (callers[0],)
    elif len(callers)>1:
        ph=','.join(['?']*len(callers)); cf=f"AND caller IN ({ph})"; cp=tuple(callers)
    else:
        cf=""; cp=()
    who = ", ".join(callers) if callers else "the team"

    # Gather aggregate stats
    base = f"FROM call_log WHERE timestamp>=? AND timestamp<=? {cf}"
    bp = (ts_from, ts_to) + cp
    def one(sql, p=bp):
        r = db_query(sql, p); return (r or [[0]])[0][0]
    total       = one(f"SELECT COUNT(*) {base}")
    if not total:
        return jsonify({"ok": False, "error": "No calls found for this caller and date range."}), 400
    total_dur   = one(f"SELECT COALESCE(SUM(duration_sec),0) {base}")
    conv2       = one(f"SELECT COUNT(*) {base} AND duration_sec>=120")
    connected   = one(f"SELECT COUNT(*) {base} AND outcome IN ('Connected','Engaged','Interested','Meeting secured','Not interested')")
    meetings    = one(f"SELECT COUNT(*) {base} AND outcome='Meeting secured'")
    interested  = one(f"SELECT COUNT(*) {base} AND outcome='Interested'")
    notint      = one(f"SELECT COUNT(*) {base} AND outcome='Not interested'")
    voicemail   = one(f"SELECT COUNT(*) {base} AND outcome='Voicemail'")
    noans       = one(f"SELECT COUNT(*) {base} AND outcome='No Answer'")
    # Outcome breakdown
    outcome_rows = db_query(f"SELECT outcome, COUNT(*), COALESCE(AVG(duration_sec),0) {base} AND outcome!='' GROUP BY outcome", bp) or []
    outcome_lines = [f"  {r[0]}: {r[1]} calls, avg {int(r[2]//60)}m{int(r[2]%60)}s" for r in outcome_rows]
    # Duration distribution
    very_short = one(f"SELECT COUNT(*) {base} AND duration_sec>0 AND duration_sec<30")
    avg_dur = round(total_dur/total) if total else 0
    connect_rate = round(connected/total*100) if total else 0
    conv_rate = round(conv2/total*100) if total else 0
    meeting_rate = round(meetings/total*100) if total else 0

    # Sample recent call notes (for qualitative insight, trimmed)
    note_rows = db_query(f"SELECT outcome, duration_sec, notes {base} AND notes!='' ORDER BY id DESC LIMIT 12", bp) or []
    note_lines = []
    for r in note_rows:
        dur = f"{int((r[1] or 0)//60)}m{int((r[1] or 0)%60)}s"
        note_lines.append(f"  [{r[0]}, {dur}] {str(r[2])[:160]}")

    # ── Charity-TYPE success analysis (what/who/how → outcome) ─────────────────
    # For every call, look up the charity's classification, then tally positive outcomes
    # by charity "what" category to see which types convert best.
    POSITIVE = ('Connected','Engaged','Interested','Meeting secured')
    type_stats = {}  # what-category → {calls, positive, meetings}
    try:
        call_regs = db_query(f"SELECT reg_number, outcome {base}", bp) or []
        for reg, oc in call_regs:
            cls = get_charity_classification(str(reg))
            whats = [w.strip() for w in (cls.get("what","") or "").split(",") if w.strip()]
            if not whats: whats = ["(unclassified)"]
            for w in whats:
                if w not in type_stats: type_stats[w] = {"calls":0,"positive":0,"meetings":0}
                type_stats[w]["calls"] += 1
                if oc in POSITIVE: type_stats[w]["positive"] += 1
                if oc == "Meeting secured": type_stats[w]["meetings"] += 1
    except Exception as e:
        print(f"  charity-type analysis error: {e}")
    # Build a ranked summary (only types with >=2 calls, sorted by positive rate)
    type_ranked = []
    for w, s in type_stats.items():
        if s["calls"] >= 2:
            rate = round(s["positive"]/s["calls"]*100)
            type_ranked.append({"type":w, "calls":s["calls"], "positive":s["positive"],
                                "meetings":s["meetings"], "rate":rate})
    type_ranked.sort(key=lambda x: (-x["rate"], -x["calls"]))
    type_lines = [f"  {t['type']}: {t['calls']} calls, {t['rate']}% positive, {t['meetings']} meetings"
                  for t in type_ranked[:12]]

    stats_block = (
        f"CALLER(S): {who}\n"
        f"PERIOD: {date_from} to {date_to}\n\n"
        f"VOLUME & EFFICIENCY:\n"
        f"  Total calls: {total}\n"
        f"  Total talk time: {int(total_dur//3600)}h {int((total_dur%3600)//60)}m\n"
        f"  Average call length: {avg_dur//60}m {avg_dur%60}s\n"
        f"  Connect rate (someone answered): {connect_rate}% ({connected}/{total})\n"
        f"  Conversations >2min: {conv_rate}% ({conv2}/{total})\n"
        f"  Very short calls (<30s): {very_short}\n\n"
        f"OUTCOMES:\n"
        f"  Meetings secured: {meetings} ({meeting_rate}% of all calls)\n"
        f"  Interested: {interested}\n"
        f"  Not interested: {notint}\n"
        f"  Voicemail: {voicemail}\n"
        f"  No answer: {noans}\n"
        + "\n".join(outcome_lines) + "\n\n"
        + ("CHARITY-TYPE SUCCESS (which charity categories respond best — ranked by positive rate):\n"
           + "\n".join(type_lines) + "\n\n" if type_lines else "")
        + ("RECENT CALL NOTES (sample):\n" + "\n".join(note_lines) if note_lines else "")
    )

    system_prompt = (
        "You are an experienced sales coach for charity-sector outbound prospecting. " + NINEM_PITCH + " "
        "You are reviewing a caller's performance to help them improve. "
        "Be specific, constructive and practical — like a supportive manager. "
        "Base every observation on the actual numbers provided; do not invent data. "
        "Be CONCISE — short punchy sentences, no padding. Use UK English."
    )
    user_prompt = (
        "Review this performance data and write a SHORT, scannable coaching report. "
        "Keep the whole thing tight — aim for brevity over completeness.\n\n"
        f"{stats_block}\n\n"
        "Use these exact headings, each with 1-3 short bullet points (not paragraphs):\n"
        "**Summary** — 1-2 sentences: overall verdict + biggest opportunity.\n"
        "**Strengths** — up to 3 bullets, grounded in numbers.\n"
        "**Improve** — up to 3 bullets: the key issues + a one-line fix each.\n"
        "**Best charity types** — 1-2 bullets: which charity categories convert best for this caller (from the charity-type data) and what to call more of.\n"
        "**Do next week** — up to 3 concrete bullet actions.\n"
        "Keep each bullet under 20 words.\n"
    )

    try:
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
            json={"model": OPENAI_MODEL,
                  "messages": [{"role": "system", "content": system_prompt},
                               {"role": "user", "content": user_prompt}],
                  "temperature": 0.6, "max_tokens": 650},
            timeout=60)
        if resp.status_code != 200:
            return jsonify({"ok": False, "error": f"OpenAI error {resp.status_code}: {resp.text[:200]}"}), 502
        out = resp.json()
        coaching = (out.get("choices") or [{}])[0].get("message", {}).get("content", "").strip()
        if not coaching:
            return jsonify({"ok": False, "error": "Empty response from OpenAI"}), 502
        return jsonify({"ok": True, "coaching": coaching, "who": who,
                        "charity_types": type_ranked[:8],
                        "stats": {"total": total, "connect_rate": connect_rate,
                                  "conv_rate": conv_rate, "meetings": meetings, "avg_dur": avg_dur}})
    except Exception as e:
        return jsonify({"ok": False, "error": f"AI request failed: {str(e)[:200]}"}), 500

@app.route("/api/ai/analyze", methods=["POST"])
def ai_analyze():
    """Generate a pre-call sales briefing for a charity using OpenAI."""
    if not OPENAI_API_KEY:
        return jsonify({"ok": False,
            "error": "AI is not configured. Add the OPENAI_API_KEY environment variable in Render."}), 400
    data = request.json or {}
    c = dict(data.get("charity") or {})
    reg = str(c.get("reg_number", "")).strip()
    if not reg:
        return jsonify({"ok": False, "error": "Missing charity reg_number"}), 400

    # Enrich with full Charity Commission data (purposes, address, financials, etc.)
    try:
        c = enrich_charity_from_cc(c)
    except Exception as e:
        print(f"  ai enrich error: {e}")

    # Gather dashboard comments + call history (separating out call transcripts)
    comments, calls, transcripts = [], [], []
    try:
        for text, cname, ts in db_query(
            "SELECT text, caller, timestamp FROM comment_history WHERE reg_number=? ORDER BY id ASC", (reg,)):
            if str(text).startswith("📝 Call transcript:"):
                transcripts.append((ts, cname or "Unknown", str(text)))
            else:
                comments.append(f"[{ts}] {cname or 'Unknown'}: {text}")
    except Exception: pass
    try:
        for outcome, dur, cname, ts in db_query(
            "SELECT outcome, duration_sec, caller, timestamp FROM call_log WHERE reg_number=? ORDER BY id ASC", (reg,)):
            d = f" ({int(dur or 0)//60}m {int(dur or 0)%60}s)" if dur else ""
            calls.append(f"[{ts}] {cname or 'Unknown'}: {outcome}{d}")
    except Exception: pass

    # Include the 2 most recent transcripts (newest first), trimmed for cost
    trans_for_prompt = [(ts, cname, _trim_transcript(text))
                        for ts, cname, text in transcripts[-2:][::-1]]
    has_transcripts = len(trans_for_prompt) > 0

    profile = _build_charity_profile_text(c, comments, calls, trans_for_prompt)

    # Category / list context — why this charity is a prospect
    category      = str(data.get("category", "")).strip()
    category_note = str(data.get("category_note", "")).strip()
    all_lists     = str(data.get("all_lists", "")).strip()
    source_detail = str(c.get("source_detail", "")).strip()
    cat_block = ""
    if category:
        cat_block = (f"\n\nPROSPECT LIST CONTEXT — this charity surfaced in 9 Mountains' "
                     f"'{category}' list. What that list means: {category_note}")
        if source_detail:
            cat_block += (f"\nSPECIFIC LIST DETAIL for this charity: {source_detail}. "
                          "Reference this concretely (e.g. a 1-year anniversary is a natural, timely reason to reach out; "
                          "an overdue filing suggests they may need support).")
        if all_lists and all_lists != category:
            cat_block += f"\n(It also qualifies for: {all_lists}.)"
        cat_block += ("\nThis is the SPECIFIC reason 9 Mountains is calling — make the "
                      "'Why they're a fit' and 'Talking points' sections speak directly to it.")
    elif source_detail:
        cat_block = (f"\n\nWHY THIS CHARITY IS ON THE LIST: {source_detail}. "
                     "Use this as a timely, concrete hook in the talking points.")

    # Known contact person + their job role (so the briefing is tuned to who we're calling)
    contact_block = ""
    try:
        cov = _contact_overrides.get(reg, {}) or {}
        cname2 = cov.get("name") or ""
        crole = cov.get("role") or ""
        if cname2 or crole:
            contact_block = (f"\n\nPERSON BEING CALLED: {cname2 or '(name unknown)'}"
                             + (f" — {crole}" if crole else "")
                             + ".\nTune the talking points and tone to this person's seniority and role: "
                             "a CEO/Founder wants strategic, big-picture, outcome-focused conversation; "
                             "a Fundraising/Development lead wants funding, donor and campaign specifics; "
                             "an Operations/Finance lead wants efficiency, cost and process angles; "
                             "a Chair/Trustee wants governance and mission alignment. "
                             "Address what THIS role typically cares about most.")
    except Exception: pass

    system_prompt = (
        "You are a sharp, concise sales-intelligence analyst preparing a caller from 9 Mountains "
        "for a phone call to a UK charity. " + NINEM_PITCH + " "
        "Give the caller a focused pre-call briefing so they sound informed and can quickly establish "
        "relevance and rapport. Be specific and practical. Use ONLY the charity data provided — do not "
        "invent facts, figures, names, or programmes. Where useful data is missing, briefly note it rather "
        "than guessing. Keep the whole briefing tight and scannable (UK English)."
    )
    user_prompt = (
        "Prepare a pre-call briefing from the data below.\n\n"
        f"{profile}"
        f"{cat_block}"
        f"{contact_block}\n\n"
        "Structure it with these short sections (use the exact headings):\n"
        + ("**Last call recap** — what was discussed on the previous call(s) from the transcript: "
           "key points, any interest/objections/commitments, and what to follow up on now. "
           "Reference specifics from what was actually said.\n" if has_transcripts else "")
        + "**Snapshot** — 1-2 sentences on what they do, who they serve, and their size.\n"
        "**Financial picture** — income, funding mix, and any notable trend or risk.\n"
        "**Why they're a fit for 9 Mountains** — likely needs or pain points, grounded in the data "
        "and the prospect-list reason above.\n"
        "**Talking points** — 3-4 specific, concrete things to mention or ask on the call, tuned to this list"
        + (" and to where the last conversation left off" if has_transcripts else "") + ".\n"
        "**Watch-outs** — anything to be tactful about, or gaps in what we know.\n"
    )

    try:
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
            json={"model": OPENAI_MODEL,
                  "messages": [{"role": "system", "content": system_prompt},
                               {"role": "user", "content": user_prompt}],
                  "temperature": 0.6, "max_tokens": 900},
            timeout=60)
        if resp.status_code != 200:
            return jsonify({"ok": False,
                "error": f"OpenAI error {resp.status_code}: {resp.text[:200]}"}), 502
        out = resp.json()
        analysis = (out.get("choices") or [{}])[0].get("message", {}).get("content", "").strip()
        if not analysis:
            return jsonify({"ok": False, "error": "Empty response from OpenAI"}), 502
        return jsonify({"ok": True, "analysis": analysis,
                        "model": OPENAI_MODEL, "name": c.get("name", "")})
    except Exception as e:
        return jsonify({"ok": False, "error": f"AI request failed: {str(e)[:200]}"}), 500

@app.route("/api/calling/exclude", methods=["POST"])
def calling_exclude():
    data = request.get_json(force=True) or {}
    reg = str(data.get("reg_number","")).strip()
    excl = bool(data.get("exclude", True))
    caller = str(data.get("caller","")).strip() or "Unknown"
    if not reg: return jsonify({"ok":False,"error":"Missing reg_number"}),400
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    if excl:
        _calling_excl.add(reg)
        db_exec("INSERT OR REPLACE INTO calling_exclusions(reg_number,excluded_by,excluded_at) VALUES(?,?,?)",
                (reg, caller, ts))
    else:
        _calling_excl.discard(reg)
        db_exec("DELETE FROM calling_exclusions WHERE reg_number=?", (reg,))
    return jsonify({"ok":True,"excluded":excl})

@app.route("/api/calling/exclusions", methods=["GET"])
def calling_exclusions_bulk():
    return jsonify(list(_calling_excl))

@app.route("/api/report")
def api_report():
    """Weekly calling performance report for a given caller + date range."""
    callers = [c.strip() for c in request.args.getlist("caller") if c.strip()]
    date_from = request.args.get("from","").strip()   # YYYY-MM-DD
    date_to   = request.args.get("to","").strip()     # YYYY-MM-DD
    if not date_from or not date_to:
        return jsonify({"ok":False,"error":"Missing from/to dates"}),400
    # Expand to full day range
    ts_from = date_from + " 00:00:00"
    ts_to   = date_to   + " 23:59:59"
    def q(sql, params=()):
        return db_query(sql, params)
    # Caller filter (empty = all callers)
    if len(callers)==1:
        cal_filter = "AND caller=?"
        cal_params = (callers[0],)
    elif len(callers)>1:
        ph=','.join(['?']*len(callers))
        cal_filter = f"AND caller IN ({ph})"
        cal_params = tuple(callers)
    else:
        cal_filter = ""
        cal_params = ()
    caller = ", ".join(callers) if callers else "All callers"
    # ── Call metrics ────────────────────────────────────────────────────────
    base = f"FROM call_log WHERE timestamp>=? AND timestamp<=? {cal_filter}"
    bp = (ts_from, ts_to) + cal_params
    calls_made      = (q(f"SELECT COUNT(*) {base}", bp) or [[0]])[0][0]
    total_dur_sec   = (q(f"SELECT COALESCE(SUM(duration_sec),0) {base}", bp) or [[0]])[0][0]
    conv_2min       = (q(f"SELECT COUNT(*) {base} AND duration_sec>=120", bp) or [[0]])[0][0]
    meaningful      = (q(f"SELECT COUNT(*) {base} AND duration_sec>=120 AND outcome IN ('Connected','Engaged','Interested','Meeting secured')", bp) or [[0]])[0][0]
    meetings        = (q(f"SELECT COUNT(*) {base} AND outcome='Meeting secured'", bp) or [[0]])[0][0]
    a_grade         = (q(f"SELECT COUNT(*) {base} AND duration_sec>=120 AND outcome IN ('Interested','Meeting secured')", bp) or [[0]])[0][0]
    voicemail       = (q(f"SELECT COUNT(*) {base} AND outcome='Voicemail'", bp) or [[0]])[0][0]
    no_answer       = (q(f"SELECT COUNT(*) {base} AND outcome='No Answer'", bp) or [[0]])[0][0]
    engaged_ct      = (q(f"SELECT COUNT(*) {base} AND outcome='Engaged'", bp) or [[0]])[0][0]
    # Actual calls = Calls Made minus No Answer minus Voicemail minus Engaged (a
    # busy/engaged line — not an actual conversation — every other outcome means
    # she actually got through to someone, even if it was a quick decline).
    actual_calls    = max(0, calls_made - voicemail - no_answer - engaged_ct)
    # Follow-up / SLT emails sent (replaces the manual promo-email figure)
    email_followups = (q(f"""SELECT COUNT(*) {base} AND outcome IN
        ('Email Sent Follow Up','SLT Unavailable - Send Email')""", bp) or [[0]])[0][0]
    # Outcome breakdown
    outcome_rows = q(f"SELECT outcome, COUNT(*) {base} AND outcome IS NOT NULL AND outcome!='' GROUP BY outcome", bp)
    outcomes = {r[0]:r[1] for r in (outcome_rows or [])}
    # Lists used
    # Lists used — join with calling_batch to get the real source category
    # For calls made from Daily Calling (page='calling'), the source is in calling_batch.category
    # For other pages (prospects, new, late etc.) page IS the source
    list_rows = q(f"""SELECT
        CASE WHEN cl.page='calling' THEN COALESCE(cb.category, 'calling') ELSE cl.page END as src,
        COUNT(*)
        FROM call_log cl
        LEFT JOIN calling_batch cb ON cb.reg_number=cl.reg_number AND cb.active=1
        WHERE cl.timestamp>=? AND cl.timestamp<=? {cal_filter}
        GROUP BY src ORDER BY COUNT(*) DESC""", (ts_from, ts_to)+cal_params)
    lists_used = [{"list": CALLING_CATEGORIES.get(r[0], r[0] or "Unknown"), "calls": r[1]} for r in (list_rows or [])]

    # Per-caller per-list breakdown for stacked bar chart
    callers_q = q(f"SELECT DISTINCT caller {base} AND caller IS NOT NULL AND caller!=''", bp) or []
    all_callers = [r[0] for r in callers_q]
    list_caller_rows = q(f"""SELECT
        CASE WHEN cl.page='calling' THEN COALESCE(cb.category, 'calling') ELSE cl.page END as src,
        cl.caller, COUNT(*)
        FROM call_log cl
        LEFT JOIN calling_batch cb ON cb.reg_number=cl.reg_number AND cb.active=1
        WHERE cl.timestamp>=? AND cl.timestamp<=? {cal_filter} AND cl.caller IS NOT NULL
        GROUP BY src, cl.caller""", (ts_from, ts_to)+cal_params) or []
    list_caller = {}
    for row in list_caller_rows:
        pg = CALLING_CATEGORIES.get(row[0], row[0] or "Unknown"); cr = row[1]; cnt = row[2]
        if pg not in list_caller: list_caller[pg] = {}
        list_caller[pg][cr] = cnt
    # ── Follow-ups ───────────────────────────────────────────────────────────
    followups = (q(f"SELECT COUNT(*) FROM follow_ups WHERE created_at>=? AND created_at<=? {cal_filter}",
                   (ts_from, ts_to)+cal_params) or [[0]])[0][0]
    # ── Newsletter sign-ups ───────────────────────────────────────────────────
    nl_filter = cal_filter.replace("caller","updated_by") if cal_filter else ""
    newsletter = (q(f"SELECT COUNT(*) FROM newsletter_flags WHERE flag=1 AND updated_at>=? AND updated_at<=? {nl_filter}",
                    (ts_from, ts_to)+cal_params) or [[0]])[0][0]
    # ── Stage counts ─────────────────────────────────────────────────────────
    stage_rows = q(f"SELECT status, COUNT(*) FROM statuses WHERE status!='' {cal_filter} GROUP BY status ORDER BY COUNT(*) DESC", cal_params)
    stages = [{"stage":r[0],"count":r[1]} for r in (stage_rows or [])]
    # ── RC call time (hours) ──────────────────────────────────────────────────
    hours_on_rc = round(total_dur_sec / 3600, 2)

    # ══ CEO: CONVERSION FUNNEL ════════════════════════════════════════════════
    connected = (q(f"SELECT COUNT(*) {base} AND outcome IN ('Connected','Engaged','Interested','Meeting secured','Not interested')", bp) or [[0]])[0][0]
    interested = (q(f"SELECT COUNT(*) {base} AND outcome IN ('Interested','Meeting secured')", bp) or [[0]])[0][0]
    clients_won = (q(f"SELECT COUNT(*) FROM statuses WHERE status='Client' {cal_filter}", cal_params) or [[0]])[0][0]
    funnel = [
        {"stage":"Calls Made",          "count": calls_made},
        {"stage":"Connected",           "count": connected},
        {"stage":"Conversations >2min", "count": conv_2min},
        {"stage":"Interested+",         "count": interested},
        {"stage":"Meetings Booked",     "count": meetings},
        {"stage":"Clients Won",         "count": clients_won},
    ]

    # ══ CEO: PIPELINE £ VALUE (by stage) ══════════════════════════════════════
    # Pull income from calling_batch.data JSON for charities at each warm stage.
    # Estimated fundraising value = 2% of total income (rough heuristic; configurable).
    PIPELINE_PCT = 0.02
    pipeline = {"Interested":{"count":0,"income":0.0,"value":0.0},
                "Meeting secured":{"count":0,"income":0.0,"value":0.0},
                "Client":{"count":0,"income":0.0,"value":0.0}}
    try:
        st_rows = q(f"SELECT key, status FROM statuses WHERE status IN ('Interested','Meeting secured','Client') {cal_filter}", cal_params) or []
        # Build reg→income map from calling_batch data
        inc_rows = q("SELECT reg_number, data FROM calling_batch WHERE active=1") or []
        inc_map = {}
        for r in inc_rows:
            try:
                obj = json.loads(r[1] or "{}")
                inc = obj.get("total_income") or obj.get("latest_income") or 0
                inc_map[r[0]] = float(inc) if inc else 0.0
            except: pass
        for r in st_rows:
            key = r[0]; status = r[1]
            reg = key.split("|")[-1] if "|" in key else key
            inc = inc_map.get(reg, 0.0)
            if status in pipeline:
                pipeline[status]["count"] += 1
                pipeline[status]["income"] += inc
                pipeline[status]["value"] += inc * PIPELINE_PCT
    except Exception as e:
        print(f"  pipeline calc error: {e}")
    pipeline_total_value = round(sum(p["value"] for p in pipeline.values()), 0)

    # ══ COO: WEEKLY TRENDS (calls + connect rate per day) ═════════════════════
    trend_rows = q(f"""SELECT substr(timestamp,1,10) as day, COUNT(*),
        SUM(CASE WHEN outcome IN ('Connected','Engaged','Interested','Meeting secured','Not interested') THEN 1 ELSE 0 END),
        COALESCE(SUM(duration_sec),0)
        {base} GROUP BY day ORDER BY day ASC""", bp) or []
    daily_trend = [{"day":r[0], "calls":r[1], "connected":r[2],
                    "connect_rate": round(r[2]/r[1]*100) if r[1] else 0,
                    "hours": round(r[3]/3600,1)} for r in trend_rows]

    # ══ COO: BEST TIME OF DAY (connect rate by hour) ══════════════════════════
    hour_rows = q(f"""SELECT substr(timestamp,12,2) as hr, COUNT(*),
        SUM(CASE WHEN outcome IN ('Connected','Engaged','Interested','Meeting secured','Not interested') THEN 1 ELSE 0 END)
        {base} GROUP BY hr ORDER BY hr ASC""", bp) or []
    by_hour = [{"hour":r[0], "calls":r[1], "connected":r[2],
                "connect_rate": round(r[2]/r[1]*100) if r[1] else 0} for r in hour_rows if r[0]]

    # ══ COO: FOLLOW-UP DISCIPLINE ═════════════════════════════════════════════
    today = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
    fu_total = (q(f"SELECT COUNT(*) FROM follow_ups WHERE 1=1 {cal_filter}", cal_params) or [[0]])[0][0]
    fu_done  = (q(f"SELECT COUNT(*) FROM follow_ups WHERE done=1 {cal_filter}", cal_params) or [[0]])[0][0]
    fu_overdue = (q(f"SELECT COUNT(*) FROM follow_ups WHERE done=0 AND follow_up_at<? {cal_filter}",
                    (today,)+cal_params) or [[0]])[0][0]
    fu_completion = round(fu_done/fu_total*100) if fu_total else 0

    # ══ FUNDRAISER: HOT LEADS (Interested / Meeting secured) ══════════════════
    hot_leads = []
    try:
        hl_rows = q(f"SELECT key, status, updated_at FROM statuses WHERE status IN ('Interested','Meeting secured') {cal_filter} ORDER BY updated_at DESC LIMIT 50", cal_params) or []
        # reg→name/phone from calling_batch data
        meta_rows = q("SELECT reg_number, data FROM calling_batch WHERE active=1") or []
        meta = {}
        for r in meta_rows:
            try:
                o = json.loads(r[1] or "{}")
                meta[r[0]] = {"name": o.get("name") or o.get("charity_name") or "",
                              "phone": o.get("phone") or o.get("contact_phone") or "",
                              "income": o.get("total_income") or o.get("latest_income") or 0}
            except: pass
        for r in hl_rows:
            key=r[0]; reg=key.split("|")[-1] if "|" in key else key
            m=meta.get(reg,{})
            hot_leads.append({"reg":reg, "name":m.get("name",""), "phone":m.get("phone",""),
                              "income":m.get("income",0), "stage":r[1], "updated":r[2]})
    except Exception as e:
        print(f"  hot leads error: {e}")

    # ══ FUNDRAISER: FOLLOW-UPS DUE (next 7 days + overdue) ════════════════════
    due_rows = q(f"""SELECT reg_number, name, follow_up_at, caller, notes
        FROM follow_ups WHERE done=0 {cal_filter}
        ORDER BY follow_up_at ASC LIMIT 50""", cal_params) or []
    followups_due = [{"reg":r[0], "name":r[1], "at":r[2], "caller":r[3], "notes":r[4] or "",
                      "overdue": (r[2] or "") < today} for r in due_rows]

    # ══ FUNDRAISER: COMPLETED FOLLOW-UPS (successful) ═════════════════════════
    done_rows = q(f"""SELECT reg_number, name, follow_up_at, caller, notes, completed_at, days_late
        FROM follow_ups WHERE done=1 {cal_filter}
        ORDER BY completed_at DESC LIMIT 50""", cal_params) or []
    followups_completed = [{"reg":r[0], "name":r[1], "at":r[2], "caller":r[3],
                            "notes":r[4] or "", "completed_at": r[5] or "",
                            "days_late": r[6] if r[6] is not None else 0} for r in done_rows]
    # On-time vs late stats
    fu_ontime = sum(1 for f in followups_completed if (f["days_late"] or 0)==0)
    fu_late   = sum(1 for f in followups_completed if (f["days_late"] or 0)>0)
    fu_avg_late = round(sum(f["days_late"] or 0 for f in followups_completed)/len(followups_completed),1) if followups_completed else 0

    return jsonify({
        "ok": True,
        "period": {"from": date_from, "to": date_to},
        "caller": caller or "All callers",
        "calls_made": calls_made,
        "actual_calls": actual_calls,
        "email_followups": email_followups,
        "total_duration_sec": total_dur_sec,
        "hours_on_rc": hours_on_rc,
        "conversations_2min": conv_2min,
        "meaningful_interactions": meaningful,
        "meetings_booked": meetings,
        "a_grade_calls": a_grade,
        "voicemail": voicemail,
        "no_answer": no_answer,
        "engaged_count": engaged_ct,
        "connected": connected,
        "interested": interested,
        "clients_won": clients_won,
        "followups_set": followups,
        "newsletter_signups": newsletter,
        "outcome_breakdown": outcomes,
        "lists_used": lists_used,
        "all_callers": all_callers,
        "list_caller": list_caller,
        "stage_totals": stages,
        # CEO
        "funnel": funnel,
        "pipeline": pipeline,
        "pipeline_total_value": pipeline_total_value,
        "pipeline_pct": PIPELINE_PCT,
        # COO
        "daily_trend": daily_trend,
        "by_hour": by_hour,
        "followup_total": fu_total,
        "followup_done": fu_done,
        "followup_overdue": fu_overdue,
        "followup_completion": fu_completion,
        # Fundraiser
        "hot_leads": hot_leads,
        "followups_due": followups_due,
        "followups_completed": followups_completed,
        "fu_ontime": fu_ontime,
        "fu_late": fu_late,
        "fu_avg_late": fu_avg_late,
    })

@app.route("/api/newsletter/bulk", methods=["GET"])
def newsletter_bulk():
    return jsonify({reg: 1 for reg in _newsletter})

@app.route("/api/newsletter", methods=["POST"])
def newsletter_set():
    data = request.get_json(force=True) or {}
    reg = str(data.get("reg_number", "")).strip()
    on  = bool(data.get("on", False))
    caller = str(data.get("caller", "")).strip() or "Unknown"
    if not reg:
        return jsonify({"ok": False, "error": "Missing reg_number"}), 400
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    if on:
        _newsletter[reg] = 1
        db_exec("INSERT OR REPLACE INTO newsletter_flags(reg_number,flag,updated_by,updated_at) VALUES(?,?,?,?)",
                (reg, 1, caller, ts))
    else:
        _newsletter.pop(reg, None)
        db_exec("DELETE FROM newsletter_flags WHERE reg_number=?", (reg,))
    return jsonify({"ok": True, "on": on})

@app.route("/api/report/export", methods=["POST"])
def report_export():
    """Export a dashboard view to a 2-sheet Excel file:
    Sheet 1 = the dashboard summary (named after the view),
    Sheet 2 = 'Called list' with every call in the date range."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"ok": False, "error": "openpyxl not installed. Add 'openpyxl' to requirements.txt and redeploy."}), 500

    data = request.get_json(force=True) or {}
    view = str(data.get("view", "Overview")).strip() or "Overview"
    callers = [c.strip() for c in (data.get("callers") or []) if c.strip()]
    date_from = str(data.get("from", "")).strip()
    date_to   = str(data.get("to", "")).strip()
    promo     = int(data.get("promo", 0) or 0)
    if not date_from or not date_to:
        return jsonify({"ok": False, "error": "Missing date range"}), 400
    ts_from = date_from + " 00:00:00"; ts_to = date_to + " 23:59:59"

    # Reuse the report logic by calling the same internal data via a fresh request context
    # Simpler: re-query what each sheet needs directly.
    if len(callers)==1:
        cf="AND caller=?"; cp=(callers[0],)
    elif len(callers)>1:
        ph=','.join(['?']*len(callers)); cf=f"AND caller IN ({ph})"; cp=tuple(callers)
    else:
        cf=""; cp=()
    who = ", ".join(callers) if callers else "All callers"
    base = f"FROM call_log WHERE timestamp>=? AND timestamp<=? {cf}"
    bp = (ts_from, ts_to) + cp
    def one(sql, p=bp):
        r = db_query(sql, p); return (r or [[0]])[0][0]

    # ── Styling helpers ──
    NAVY="1E3A5F"; GOLD="E8B339"; GREEN="1D7A3A"; LIGHT="F2F4F7"
    thin=Side(style="thin", color="D5DAE0")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    def style_header(cell):
        cell.font=Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill=PatternFill("solid", fgColor=NAVY)
        cell.alignment=Alignment(horizontal="left", vertical="center")
        cell.border=border
    def style_title(cell):
        cell.font=Font(name="Arial", bold=True, color=NAVY, size=16)
    def style_label(cell):
        cell.font=Font(name="Arial", bold=True, color="555555", size=10)
    def style_val(cell):
        cell.font=Font(name="Arial", size=11, color="000000")

    wb = Workbook()
    ws = wb.active
    # Sheet name max 31 chars, no special chars
    safe_view = view.replace("&","and")[:31]
    ws.title = safe_view

    # ── SHEET 1: Dashboard summary ──
    ws["A1"] = "9 Mountains — " + view
    style_title(ws["A1"]); ws.merge_cells("A1:D1")
    ws["A2"] = f"Period: {date_from} to {date_to}    |    Caller(s): {who}"
    ws["A2"].font = Font(name="Arial", italic=True, color="777777", size=10)
    ws.merge_cells("A2:D2")

    row = 4
    def kpi(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label); style_label(ws.cell(row=row,column=1))
        ws.cell(row=row, column=2, value=value); style_val(ws.cell(row=row,column=2))
        row += 1

    # Common metrics
    total       = one(f"SELECT COUNT(*) {base}")
    total_dur   = one(f"SELECT COALESCE(SUM(duration_sec),0) {base}")
    conv2       = one(f"SELECT COUNT(*) {base} AND duration_sec>=120")
    connected   = one(f"SELECT COUNT(*) {base} AND outcome IN ('Connected','Engaged','Interested','Meeting secured','Not interested')")
    meetings    = one(f"SELECT COUNT(*) {base} AND outcome='Meeting secured'")
    interested  = one(f"SELECT COUNT(*) {base} AND outcome='Interested'")
    voicemail   = one(f"SELECT COUNT(*) {base} AND outcome='Voicemail'")
    noans       = one(f"SELECT COUNT(*) {base} AND outcome='No Answer'")
    meaningful  = one(f"SELECT COUNT(*) {base} AND duration_sec>=120 AND outcome IN ('Connected','Engaged','Interested','Meeting secured')")
    a_grade     = one(f"SELECT COUNT(*) {base} AND duration_sec>=120 AND outcome IN ('Interested','Meeting secured')")
    hours = round(total_dur/3600,2)
    connect_rate = round(connected/total*100) if total else 0

    ws.cell(row=row, column=1, value="KEY METRICS"); ws.cell(row=row,column=1).font=Font(name="Arial",bold=True,color=NAVY,size=12); row+=1
    kpi("Calls Made", total)
    kpi("Hours on RingCentral", hours)
    kpi("Connect Rate", f"{connect_rate}%")
    kpi("Conversations >2min", conv2)
    kpi("Meaningful Interactions", meaningful)
    kpi("A-Grade Calls", a_grade)
    kpi("Meetings Booked", meetings)
    kpi("Interested", interested)
    kpi("Voicemail", voicemail)
    kpi("No Answer", noans)
    kpi("Promo Emails Sent (manual)", promo)
    row += 1

    # View-specific section
    if view.lower().startswith("funnel") or "pipeline" in view.lower():
        ws.cell(row=row,column=1,value="CONVERSION FUNNEL"); ws.cell(row=row,column=1).font=Font(name="Arial",bold=True,color=NAVY,size=12); row+=1
        clients = one(f"SELECT COUNT(*) FROM statuses WHERE status='Client' {cf}", cp)
        for stg,val in [("Calls Made",total),("Connected",connected),("Conversations >2min",conv2),
                        ("Interested+",a_grade),("Meetings Booked",meetings),("Clients Won",clients)]:
            kpi(stg, val)
    elif view.lower().startswith("efficiency"):
        ws.cell(row=row,column=1,value="EFFICIENCY"); ws.cell(row=row,column=1).font=Font(name="Arial",bold=True,color=NAVY,size=12); row+=1
        cph = round(total/hours) if hours>0 else 0
        kpi("Calls per Hour", cph)
        fu_total=one(f"SELECT COUNT(*) FROM follow_ups WHERE 1=1 {cf}", cp)
        fu_done=one(f"SELECT COUNT(*) FROM follow_ups WHERE done=1 {cf}", cp)
        kpi("Follow-ups Total", fu_total)
        kpi("Follow-ups Completed", fu_done)
    elif view.lower().startswith("hot"):
        ws.cell(row=row,column=1,value="HOT LEADS (Interested / Meeting secured)"); ws.cell(row=row,column=1).font=Font(name="Arial",bold=True,color=NAVY,size=12); row+=1
        hdr=["Charity Reg","Stage","Last Updated"]
        for ci,h in enumerate(hdr,1): style_header(ws.cell(row=row,column=ci,value=h))
        row+=1
        hl=db_query(f"SELECT key, status, updated_at FROM statuses WHERE status IN ('Interested','Meeting secured') {cf} ORDER BY updated_at DESC LIMIT 100", cp) or []
        for k,st,ua in hl:
            reg=k.split("|")[-1] if "|" in k else k
            ws.cell(row=row,column=1,value=reg); ws.cell(row=row,column=2,value=st); ws.cell(row=row,column=3,value=(ua or "")[:16])
            for ci in range(1,4): ws.cell(row=row,column=ci).border=border
            row+=1

    # Outcome breakdown table (all views)
    row += 1
    ws.cell(row=row,column=1,value="OUTCOME BREAKDOWN"); ws.cell(row=row,column=1).font=Font(name="Arial",bold=True,color=NAVY,size=12); row+=1
    for ci,h in enumerate(["Outcome","Count"],1): style_header(ws.cell(row=row,column=ci,value=h))
    row+=1
    for oc,cnt in (db_query(f"SELECT outcome, COUNT(*) {base} AND outcome!='' GROUP BY outcome ORDER BY COUNT(*) DESC", bp) or []):
        ws.cell(row=row,column=1,value=oc); ws.cell(row=row,column=2,value=cnt)
        ws.cell(row=row,column=1).border=border; ws.cell(row=row,column=2).border=border
        row+=1

    ws.column_dimensions["A"].width=34
    ws.column_dimensions["B"].width=22
    ws.column_dimensions["C"].width=20
    ws.column_dimensions["D"].width=18

    # ── SHEET 2: Called list ──
    ws2 = wb.create_sheet("Called list")
    headers = ["Date/Time","Caller","Charity Name","Reg Number","Source List",
               "Phone","Outcome","Duration","Stage","Notes"]
    for ci,h in enumerate(headers,1): style_header(ws2.cell(row=1,column=ci,value=h))
    # Pull every call in range, join calling_batch for source + name
    call_rows = db_query(f"""SELECT cl.timestamp, cl.caller, cl.reg_number, cl.phone_used,
        cl.outcome, cl.duration_sec, cl.notes,
        CASE WHEN cl.page='calling' THEN COALESCE(cb.category,'calling') ELSE cl.page END as src,
        cb.data
        FROM call_log cl
        LEFT JOIN calling_batch cb ON cb.reg_number=cl.reg_number AND cb.active=1
        WHERE cl.timestamp>=? AND cl.timestamp<=? {cf}
        ORDER BY cl.timestamp DESC""", (ts_from, ts_to)+cp) or []
    r2 = 2
    for ts, caller_n, reg, phone, outcome, dur, notes, src, cbdata in call_rows:
        name=""
        try:
            if cbdata:
                o=json.loads(cbdata); name=o.get("name") or o.get("charity_name") or ""
        except Exception: pass
        key=f"calling|{reg}"
        stage=_statuses.get(key,"")
        dn=int(dur or 0)
        ws2.cell(row=r2,column=1,value=(ts or "")[:16])
        ws2.cell(row=r2,column=2,value=caller_n or "")
        ws2.cell(row=r2,column=3,value=name)
        ws2.cell(row=r2,column=4,value=reg)
        ws2.cell(row=r2,column=5,value=CALLING_CATEGORIES.get(src, src or ""))
        ws2.cell(row=r2,column=6,value=phone or "")
        ws2.cell(row=r2,column=7,value=outcome or "")
        ws2.cell(row=r2,column=8,value=f"{dn//60}:{dn%60:02d}")
        ws2.cell(row=r2,column=9,value=stage)
        ws2.cell(row=r2,column=10,value=(notes or "")[:500])
        for ci in range(1,11): ws2.cell(row=r2,column=ci).font=Font(name="Arial",size=10)
        r2+=1
    widths=[18,14,30,12,18,16,16,10,16,50]
    for ci,w in enumerate(widths,1):
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.freeze_panes="A2"
    ws.freeze_panes="A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tstamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    fname = f"{safe_view.replace(' ','_')}_report_{tstamp}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.route("/api/calling/export", methods=["POST"])
def calling_export():
    """Export the active Daily Calling list to CSV — all fields, status, and full history."""
    data = request.get_json(force=True) or {}
    regs_filter = data.get("regs")
    regs_set = set(str(r) for r in regs_filter) if regs_filter else None

    rows = db_query("SELECT reg_number, data, completed FROM calling_batch WHERE active=1 ORDER BY position ASC")
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Reg Number","Charity Name","Priority Band","Priority Score",
                "Phone","Email","Website","Contact Person","Job Role","Contact Phone","Contact Email",
                "Total Income","Statutory Income","Source","Stage","Called","Called By","Called At",
                "Newsletter","Current Comment","Full History"])
    for reg, d, completed in rows:
        if regs_set is not None and str(reg) not in regs_set:
            continue
        try: c = json.loads(d) if d else {}
        except Exception: c = {}
        ov = _contact_overrides.get(reg, {})
        key = f"calling|{reg}"
        status = _statuses.get(key, "")
        called = _called_log.get(key)
        called_yes = "Yes" if (called or completed) else "No"
        called_by = (called or {}).get("called_by", "")
        called_at = (called or {}).get("timestamp", "")
        newsletter = "Yes" if _newsletter.get(reg) else "No"
        cur_comment = _comments.get(key, "")
        hist = []
        try:
            for ts, cname, text in db_query("SELECT timestamp, caller, text FROM comment_history WHERE reg_number=? ORDER BY id ASC", (reg,)):
                hist.append(f"[{ts}] NOTE by {cname or '?'}: {text or ''}")
        except Exception: pass
        try:
            for ts, cname, outcome, dur, notes in db_query("SELECT timestamp, caller, outcome, duration_sec, notes FROM call_log WHERE reg_number=? ORDER BY id ASC", (reg,)):
                dn = int(dur or 0)
                line = f"[{ts}] CALL: {outcome or ''} ({dn//60}:{dn%60:02d}) by {cname or '?'}"
                if notes: line += f" — {notes}"
                hist.append(line)
        except Exception: pass
        full_hist = "\n".join(hist)
        phone   = ov.get("phone")   or c.get("phone", "")
        email   = ov.get("email")   or c.get("email", "")
        website = ov.get("website") or c.get("website", "")
        w.writerow([reg, c.get("name",""), c.get("band",""), c.get("priority",""),
                    phone, email, website, ov.get("name",""), ov.get("role",""), ov.get("cphone",""), ov.get("cemail",""),
                    c.get("total_income",""), c.get("statutory",""), c.get("source",""),
                    status, called_yes, called_by, called_at, newsletter, cur_comment, full_hist])
    out = "\ufeff" + buf.getvalue()   # BOM so Excel reads UTF-8 (£, emojis) correctly
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
    return Response(out, mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=daily_calling_{ts}.csv"})

@app.route("/api/ai/find_contact", methods=["POST"])
def api_ai_find_contact():
    """Use a web-search-grounded model to find the most senior contactable person
    at a charity (name, role, phone, email). Persists results + a Maximizer note."""
    data = request.get_json(force=True) or {}
    reg  = (data.get("reg_number") or "").strip()
    name = (data.get("name") or "").strip()
    caller = (data.get("caller") or "AI lookup")
    if not reg and not name:
        return jsonify({"ok": False, "error": "Missing charity"}), 400
    if not OPENAI_API_KEY:
        return jsonify({"ok": False, "error": "AI is not configured. Add OPENAI_API_KEY in Render."}), 400

    search_model = os.environ.get("OPENAI_SEARCH_MODEL", "gpt-4o-mini-search-preview")
    prompt = (
        "You are helping a UK charity-fundraising team find the best people to call at a charity.\n"
        f"Charity name: {name}\n"
        f"Charity Commission registration number: {reg}\n\n"
        "Find up to THREE most senior contactable people, ranked by seniority (most senior first), "
        "using this strict priority order: "
        "1) CEO / Chief Executive (highest); 2) Founder; 3) COO / Chief Operating Officer; "
        "4) Executive Director / Managing Director; 5) Director of Fundraising / Head of Fundraising; "
        "6) other senior Directors or VPs; 7) Chair of Trustees; 8) the charity's main named contact. "
        "List the highest job grade first and work downward. Only include people who genuinely exist. "
        "Check the charity's official website and the Charity Commission register "
        "(register-of-charities.charitycommission.gov.uk).\n\n"
        "Return ONLY a compact JSON object, no prose, with this exact shape: "
        '{"contacts": [{"name":"", "role":"", "phone":"", "email":""}, ...]}. '
        "Include 1 to 3 contacts in the contacts array, most senior first. "
        "Use an empty string for anything you cannot find on a real, current web page. "
        "NEVER guess or invent a phone number or email — only include them if they actually appear on a real page. "
        "A general charity phone/email is acceptable if no personal one is published."
    )
    try:
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
            json={"model": search_model,
                  "web_search_options": {"user_location": {"type": "approximate",
                                                           "approximate": {"country": "GB"}}},
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=90)
        if resp.status_code != 200:
            return jsonify({"ok": False, "error": f"OpenAI error {resp.status_code}: {resp.text[:200]}"}), 502
        content = ((resp.json().get("choices") or [{}])[0].get("message", {}) or {}).get("content", "") or ""
        m = re.search(r"\{.*\}", content, re.DOTALL)
        parsed = {}
        if m:
            try: parsed = json.loads(m.group(0))
            except Exception: parsed = {}
        # Support both new {contacts:[...]} and legacy flat shape
        raw_contacts = parsed.get("contacts") if isinstance(parsed.get("contacts"), list) else None
        if raw_contacts is None:
            raw_contacts = [parsed] if any(parsed.get(k) for k in ("name","role","phone","email")) else []
        contacts = []
        for c in raw_contacts[:3]:
            cc = {"name":  str(c.get("name")  or "").strip()[:CONTACT_LIMITS["name"]],
                  "role":  str(c.get("role")  or "").strip()[:CONTACT_LIMITS["role"]],
                  "phone": str(c.get("phone") or "").strip()[:CONTACT_LIMITS["cphone"]],
                  "email": str(c.get("email") or "").strip()[:CONTACT_LIMITS["cemail"]]}
            if any(cc.values()): contacts.append(cc)
        if not contacts:
            return jsonify({"ok": True, "found": False, "contacts": [],
                            "note": "The AI couldn't find a verifiable contact for this charity.",
                            "name":"","role":"","phone":"","email":""})
        # Top contact is stored in the row's contact fields
        out = contacts[0]
        if reg:
            existing = db_query("SELECT phone,email,website,contact_name,contact_role,contact_phone,contact_email FROM phone_overrides WHERE reg_number=?", (reg,))
            e0 = existing[0] if existing else ("","","","","","","")
            def _g(i): return (e0[i] if len(e0)>i else "") or ""
            cur = {"phone":_g(0),"email":_g(1),"website":_g(2),"name":_g(3),"role":_g(4),"cphone":_g(5),"cemail":_g(6)}
            if out["name"]:  cur["name"]   = out["name"]
            if out["role"]:  cur["role"]   = out["role"]
            if out["phone"]: cur["cphone"] = out["phone"]
            if out["email"]: cur["cemail"] = out["email"]
            ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            contacts_json = json.dumps(contacts)
            db_exec("""INSERT OR REPLACE INTO phone_overrides
                       (reg_number,phone,email,website,contact_name,contact_role,contact_phone,contact_email,contacts_json,updated_by,updated_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                    (reg, cur["phone"], cur["email"], cur["website"], cur["name"], cur["role"], cur["cphone"], cur["cemail"], contacts_json, caller, ts))
            _contact_overrides[reg] = {k:v for k,v in cur.items() if v}
            _contact_overrides[reg]["contacts"] = contacts
            try:
                found = mx_find_by_org_number(reg)
                if found and found.get("key"):
                    lines = ["🤖 AI-suggested contacts (most senior first):"]
                    for i,c in enumerate(contacts,1):
                        seg = f"{i}. {c['name'] or '(name not found)'}" + (f" — {c['role']}" if c['role'] else "")
                        if c["phone"]: seg += f"  ☎ {c['phone']}"
                        if c["email"]: seg += f"  ✉ {c['email']}"
                        lines.append(seg)
                    lines.append("(verify before use)")
                    mx_create_note(found["key"], "  |  ".join(lines), caller, timestamp=ts)
            except Exception as e:
                print(f"  find_contact→mx note error: {e}")
        return jsonify({"ok": True, "found": True, "contacts": contacts, **out})
    except Exception as e:
        return jsonify({"ok": False, "error": f"AI lookup failed: {str(e)[:200]}"}), 502

@app.route("/api/contact", methods=["POST"])
def contact_override():
    """Save a manual phone/email/website override → Turso + Maximizer (if entry exists)."""
    data = request.json or {}
    return _save_contact(
        reg=str(data.get("reg_number","")).strip(),
        field=str(data.get("field","")).strip().lower(),
        val=str(data.get("value","")).strip(),
        caller=str(data.get("caller","")).strip() or "Unknown",
    )

# ── Call log ──────────────────────────────────────────────────────────────────
@app.route("/api/call", methods=["POST"])
def save_call():
    """Save a completed call, log notes as a comment, store any follow-up,
    then run a full Maximizer sync (create/update + push notes & call)."""
    data = request.json or {}
    reg     = str(data.get("reg_number","")).strip()
    page    = str(data.get("page","")).strip()
    phone   = str(data.get("phone","")).strip()
    outcome = str(data.get("outcome","")).strip()[:40]
    notes   = str(data.get("notes","")).strip()[:800]
    caller  = str(data.get("caller","")).strip() or "Unknown"
    status  = str(data.get("status","")).strip()
    fu_at   = str(data.get("follow_up_at","")).strip()   # "YYYY-MM-DD HH:MM"
    retry_at_in = str(data.get("retry_at","")).strip()   # optional custom retry time
    name    = str(data.get("name","")).strip()
    try: duration = max(0, int(data.get("duration_sec", 0) or 0))
    except: duration = 0
    if not reg or not outcome:
        return jsonify({"ok": False, "error": "Missing reg_number or outcome"}), 400
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    # Flag: should we try to fetch duration from RC call log?
    fetch_rc_duration = (duration == 0 and phone and _rc_configured())

    # Retry cooldown: for "didn't get through" outcomes, work out when this
    # charity becomes eligible to be called again — the caller's own custom time
    # if they set one, otherwise the default cooldown from now.
    retry_at = ""
    if outcome in RETRY_OUTCOMES:
        if retry_at_in:
            retry_at = retry_at_in if len(retry_at_in) > 10 else retry_at_in + " 09:00"
        else:
            retry_at = (datetime.utcnow() + timedelta(hours=DEFAULT_RETRY_HOURS)).strftime("%Y-%m-%d %H:%M:%S")
        print(f"  [save_call] reg={reg} outcome={outcome!r} server_now_utc={ts!r} "
              f"retry_at_from_client={retry_at_in!r} retry_at_stored={retry_at!r}")

    # 1) Log the call (synced=0 — the sync below will push it)
    db_exec("""INSERT INTO call_log
        (reg_number,page,phone_used,outcome,notes,duration_sec,caller,timestamp,synced_to_maximizer,retry_at)
        VALUES(?,?,?,?,?,?,?,?,0,?)""",
        (reg, page, phone, outcome, notes, duration, caller, ts, retry_at or None))

    # 2) Capture the call notes as a dashboard comment (so it shows in comment history
    #    AND gets pushed to Maximizer as a Note by the sync below)
    if notes:
        db_exec("INSERT INTO comment_history(reg_number,page,text,caller,timestamp,synced_to_maximizer) VALUES(?,?,?,?,?,0)",
                (reg, page, notes, caller, ts))

    # 3) Store a follow-up if one was scheduled
    if fu_at:
        db_exec("""INSERT INTO follow_ups(reg_number,name,page,follow_up_at,caller,notes,created_at,done)
                   VALUES(?,?,?,?,?,?,?,0)""",
                (reg, name, page, fu_at, caller, notes, ts))

    # 4) Mark called
    key = f"{page}|{reg}"
    _called_log[key] = {"reg_number": reg, "page": page, "name": name,
                        "called_by": caller, "timestamp": ts}
    db_exec("INSERT OR REPLACE INTO called_log(key,reg_number,page,name,called_by,timestamp,data) VALUES(?,?,?,?,?,?,?)",
            (key, reg, page, name, caller, ts, "{}"))
    # Mark this charity completed in the active calling batch (if it's in one)
    db_exec("UPDATE calling_batch SET completed=1 WHERE active=1 AND reg_number=?", (reg,))

    # 5) Full Maximizer sync — create/update entry + push the comment & call note
    synced = 0; sync_error = ""
    try:
        charity = dict(data.get("charity") or {})
        charity["reg_number"] = reg
        charity["_called"] = True
        charity["_caller"] = caller
        charity["_status"] = status
        do_sync_one(reg, charity, caller, page)
        synced = 1
    except Exception as e:
        sync_error = str(e)[:200]
        print(f"  call→full sync error: {e}")

    # 6) If duration wasn't captured in browser (RC desktop app calling),
    #    fetch the real duration from RC call log in a background thread
    if fetch_rc_duration:
        def _bg_duration(phone_n, reg_n, page_n, ts_n):
            rc_dur, err = _rc_fetch_call_duration(phone_n, max_wait=20)
            if rc_dur and rc_dur > 0:
                db_exec("UPDATE call_log SET duration_sec=? WHERE reg_number=? AND page=? AND timestamp=?",
                        (rc_dur, reg_n, page_n, ts_n))
                print(f"  RC duration fetched: {rc_dur}s for {reg_n}")
            else:
                print(f"  RC duration fetch failed: {err}")
        threading.Thread(target=_bg_duration, args=(phone, reg, page, ts), daemon=True).start()

    return jsonify({"ok": True, "synced_to_maximizer": synced, "sync_error": sync_error,
                    "rc_duration_pending": fetch_rc_duration})

# ── Retry cooldown for "didn't get through" outcomes ─────────────────────────
# No Answer / Voicemail / Busy / Line Dropped all mean the charity was never
# actually reached, so — rather than marking them permanently done — they're
# eligible for a fresh attempt after a cooling-off period, up to a capped
# number of attempts, then treated as exhausted (needs another channel, e.g.
# email) rather than falsely marked "done" as if a real conversation happened.
RETRY_OUTCOMES = {"No Answer", "Voicemail", "Busy", "Line Dropped", "Engaged"}
DEFAULT_RETRY_HOURS = 24
MAX_RETRY_ATTEMPTS = 3

_retry_backfill_done = False

def _backfill_retry_at_if_needed():
    """One-time correction for existing call_log rows: recompute retry_at from
    each call's OWN timestamp + the standard cooldown. This fixes rows whose
    retry_at was stored by the earlier browser-local-time-based calculation
    (since removed), which could produce an incorrect — sometimes already-past
    — retry time. Runs once per server lifetime, triggered lazily so it never
    touches the early app-startup sequence."""
    global _retry_backfill_done
    if _retry_backfill_done: return
    _retry_backfill_done = True
    try:
        placeholders = ",".join("?" for _ in RETRY_OUTCOMES)
        rows = db_query(f"SELECT id, timestamp, outcome, retry_at FROM call_log WHERE outcome IN ({placeholders})",
                        tuple(RETRY_OUTCOMES)) or []
        fixed = 0
        for rid, ts, oc, retry_at in rows:
            if not ts: continue
            try:
                call_dt = datetime.strptime(str(ts)[:19], "%Y-%m-%d %H:%M:%S")
            except Exception:
                continue
            correct = (call_dt + timedelta(hours=DEFAULT_RETRY_HOURS)).strftime("%Y-%m-%d %H:%M:%S")
            if str(retry_at or "") != correct:
                db_exec("UPDATE call_log SET retry_at=? WHERE id=?", (correct, rid))
                fixed += 1
        print(f"  [retry backfill] corrected {fixed} of {len(rows)} existing call_log retry_at values")
    except Exception as e:
        print(f"  [retry backfill] error: {e}")

def _retry_states_map():
    """reg_number -> {retry_count, retry_at, retry_due, retry_exhausted}, computed
    from the most recent CONSECUTIVE retry-eligible calls per charity (a real
    conversation outcome resets the streak). Groups and sorts in Python rather
    than relying on the database's multi-column ORDER BY, so grouping is always
    correct regardless of how reg_number is stored/typed."""
    _backfill_retry_at_if_needed()
    rows = db_query("SELECT reg_number, timestamp, outcome, retry_at FROM call_log") or []
    now_full = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    by_reg = {}
    for reg, ts, oc, retry_at in rows:
        reg = str(reg or "").strip()
        if not reg: continue
        by_reg.setdefault(reg, []).append((str(ts or ""), str(oc or "").strip(), retry_at))
    out = {}
    for reg, calls in by_reg.items():
        calls.sort(key=lambda x: x[0], reverse=True)  # most recent call first, guaranteed
        streak = 0
        latest_retry_at = None
        for ts, oc, retry_at in calls:
            if oc in RETRY_OUTCOMES:
                streak += 1
                if latest_retry_at is None:
                    latest_retry_at = retry_at or ts  # most recent row for this reg
                if streak >= MAX_RETRY_ATTEMPTS:
                    break
            else:
                break  # a real outcome (or blank) — streak ends here
        if streak > 0:
            exhausted = streak >= MAX_RETRY_ATTEMPTS
            due = (not exhausted) and bool(latest_retry_at) and str(latest_retry_at) <= now_full
            out[reg] = {"retry_count": streak, "retry_at": latest_retry_at,
                        "retry_due": due, "retry_exhausted": exhausted}
    return out

def _latest_outcomes_map():
    """reg_number -> most recent call outcome, from call_log. Used to show a Call
    Outcome column/filter on both Daily Calling and Call History."""
    out = {}
    for reg, ts, oc in (db_query(
            "SELECT reg_number, timestamp, outcome FROM call_log ORDER BY timestamp ASC") or []):
        reg = str(reg or "").strip()
        if reg and oc: out[reg] = oc  # ascending scan, last write = most recent
    return out

def _charity_snapshots_for_regs(regs):
    """Build full charity row dicts (same shape as Daily Calling rows) for a specific
    set of reg numbers, using whatever data we already have on file (calling_batch
    snapshot, contact overrides, called_log name, call_log for last-dialled number).
    Used to enrich the follow-up list so the caller always has a phone number to
    dial and can see full charity info, even if the charity isn't in today's
    active 100."""
    regset = {str(r) for r in regs if r}
    if not regset: return {}
    batch_data = {}
    for reg, data in (db_query("SELECT reg_number, data FROM calling_batch") or []):
        if reg not in regset or reg in batch_data: continue
        try:
            o = json.loads(data) if data else {}
            if o: batch_data[reg] = o
        except Exception: pass
    names = {}
    for reg, nm in (db_query("SELECT reg_number, name FROM called_log") or []):
        if reg in regset and nm: names[reg] = nm
    # Last dialled number per charity — most recent call_log entry that has a number
    last_dialled = {}
    for reg, ts, phone in (db_query(
            "SELECT reg_number, timestamp, phone_used FROM call_log ORDER BY timestamp ASC") or []):
        if reg in regset and phone:
            last_dialled[reg] = phone  # ascending order → ends up holding the most recent
    outcomes = _latest_outcomes_map()
    retry_states = _retry_states_map()
    out = {}
    for reg in regset:
        base = dict(batch_data.get(reg, {}))
        base["reg_number"] = reg
        if not base.get("name"): base["name"] = names.get(reg, reg)
        ov = _contact_overrides.get(reg) or {}
        if ov.get("phone"):   base["phone"]   = ov["phone"]
        if ov.get("email"):   base["email"]   = ov["email"]
        if ov.get("website"): base["website"] = ov["website"]
        base["completed"]    = 1 if _called_today(reg) else 0
        base["called_today"] = _called_today(reg)
        base["ever_called"]  = _ever_called(reg)
        base["last_dialled"] = last_dialled.get(reg, "") or base.get("phone","")
        base["outcome"]      = outcomes.get(str(reg), "")
        rs = retry_states.get(str(reg)) or {}
        base["retry_count"]     = rs.get("retry_count", 0)
        base["retry_at"]        = rs.get("retry_at", "")
        base["retry_due"]       = rs.get("retry_due", False)
        base["retry_exhausted"] = rs.get("retry_exhausted", False)
        # Closest available proxy for "who we last spoke to" — the contact on file
        contacts = ov.get("contacts") or []
        if contacts:
            base["last_contact_name"] = contacts[0].get("name","")
            base["last_contact_role"] = contacts[0].get("role","")
        else:
            base["last_contact_name"] = ov.get("name","")
            base["last_contact_role"] = ov.get("role","")
        if not base.get("source_detail"):
            base["source_detail"] = "📞 " + (base.get("source") or "Follow-up")
        out[reg] = base
    return out

@app.route("/api/followups")
def followups_list():
    """Upcoming follow-ups. ?scope=today (default), ?scope=all (today + future),
    or ?from=YYYY-MM-DD&to=YYYY-MM-DD for the calendar (any range, includes past)."""
    scope = str(request.args.get("scope","today")).strip()
    dfrom = str(request.args.get("from","")).strip()
    dto   = str(request.args.get("to","")).strip()
    today = datetime.utcnow().strftime("%Y-%m-%d")
    now_full = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    try:
        if dfrom or dto:
            where=["done=0"]; params=[]
            if dfrom: where.append("substr(follow_up_at,1,10)>=?"); params.append(dfrom)
            if dto:   where.append("substr(follow_up_at,1,10)<=?"); params.append(dto)
            rows = db_query("SELECT id,reg_number,name,page,follow_up_at,caller,notes FROM follow_ups WHERE "
                            + " AND ".join(where) + " ORDER BY follow_up_at ASC", tuple(params))
        elif scope == "all":
            rows = db_query("""SELECT id,reg_number,name,page,follow_up_at,caller,notes
                               FROM follow_ups WHERE done=0 AND substr(follow_up_at,1,10)>=?
                               ORDER BY follow_up_at ASC""", (today,))
        else:
            # "today" now also includes OVERDUE follow-ups (dates before today that
            # are still not done), so the caller is nudged to clear them.
            rows = db_query("""SELECT id,reg_number,name,page,follow_up_at,caller,notes
                               FROM follow_ups WHERE done=0 AND substr(follow_up_at,1,10)<=?
                               ORDER BY follow_up_at ASC""", (today,))
        followup_rows = rows or []
        snapshots = _charity_snapshots_for_regs([r[1] for r in followup_rows])
        out = []
        for r in followup_rows:
            fid, reg, nm, pg, fat, caller, notes = r
            snap = dict(snapshots.get(str(reg)) or {})
            # follow_ups.name is the reliable fallback if no snapshot name is on file
            if not snap.get("name"): snap["name"] = nm or str(reg)
            item = {"id": fid, "reg_number": reg, "name": snap.get("name"), "page": pg,
                    "follow_up_at": fat, "caller": caller, "notes": notes,
                    "overdue": bool(fat) and str(fat) < now_full}
            item.update(snap)  # phone, email, website, source, source_detail, income, etc.
            # Re-assert the follow-up's OWN fields after the merge — this specific
            # follow-up's id/date/caller/note must never be silently overwritten by
            # data from the charity snapshot (which reflects the charity in general,
            # not this particular scheduled follow-up).
            item["id"] = fid
            item["reg_number"] = reg
            item["page"] = pg
            item["follow_up_at"] = fat
            item["caller"] = caller
            item["notes"] = notes
            item["fu_notes"] = notes  # explicit alias so the UI can label it clearly
            item["overdue"] = bool(fat) and str(fat) < now_full
            out.append(item)
        return jsonify({"followups": out})
    except Exception as e:
        return jsonify({"followups": [], "error": str(e)[:120]})

@app.route("/api/followup/done", methods=["POST"])
def followup_done():
    """Mark a follow-up as completed, recording when and how late vs the scheduled date."""
    data = request.json or {}
    fid = data.get("id")
    done = data.get("done", True)
    if not fid: return jsonify({"ok": False}), 400
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    if done:
        # Compute days late (completed date minus scheduled date)
        days_late = 0
        try:
            row = db_query("SELECT follow_up_at FROM follow_ups WHERE id=?", (fid,))
            if row and row[0][0]:
                sched = str(row[0][0])[:10]
                comp  = ts[:10]
                from datetime import datetime as _dt
                d1 = _dt.strptime(sched, "%Y-%m-%d")
                d2 = _dt.strptime(comp, "%Y-%m-%d")
                days_late = max(0, (d2 - d1).days)
        except Exception as e:
            print(f"  days_late calc error: {e}")
        db_exec("UPDATE follow_ups SET done=1, completed_at=?, days_late=? WHERE id=?", (ts, days_late, fid))
        return jsonify({"ok": True, "days_late": days_late})
    else:
        db_exec("UPDATE follow_ups SET done=0, completed_at=NULL, days_late=NULL WHERE id=?", (fid,))
        return jsonify({"ok": True})

@app.route("/api/followup/reschedule", methods=["POST"])
def followup_reschedule():
    """Change the date/time of an existing follow-up."""
    data = request.json or {}
    fid = data.get("id")
    new_at = str(data.get("follow_up_at", "")).strip()  # 'YYYY-MM-DD HH:MM'
    if not fid or not new_at:
        return jsonify({"ok": False, "error": "Missing id or date/time"}), 400
    db_exec("UPDATE follow_ups SET follow_up_at=? WHERE id=?", (new_at, fid))
    return jsonify({"ok": True, "follow_up_at": new_at})

@app.route("/api/followup/delete", methods=["POST"])
def followup_delete():
    """Permanently delete a follow-up (e.g. created during testing)."""
    data = request.json or {}
    fid = data.get("id")
    if not fid: return jsonify({"ok": False, "error": "Missing id"}), 400
    db_exec("DELETE FROM follow_ups WHERE id=?", (fid,))
    return jsonify({"ok": True})

# ── Email sending (SMTP) ──────────────────────────────────────────────────────
SMTP_HOST      = os.environ.get("SMTP_HOST", "")
SMTP_PORT      = int(os.environ.get("SMTP_PORT", "587") or 587)
SMTP_USER      = os.environ.get("SMTP_USER", "")
SMTP_PASS      = os.environ.get("SMTP_PASS", "")
SMTP_FROM      = os.environ.get("SMTP_FROM", "") or SMTP_USER
SMTP_FROM_NAME = os.environ.get("SMTP_FROM_NAME", "9 Mountains")

def _smtp_configured():
    return bool(SMTP_HOST and SMTP_USER and SMTP_PASS)

@app.route("/api/email/config")
def email_config():
    """Tell the frontend whether sending is configured (so it can show a helpful message)."""
    return jsonify({"configured": _smtp_configured(),
                    "from": SMTP_FROM if _smtp_configured() else ""})

@app.route("/api/email/send", methods=["POST"])
def email_send():
    """Send an email directly from the dashboard via SMTP, and log it to
    the charity's history + Maximizer Notes."""
    if not _smtp_configured():
        return jsonify({"ok": False, "error":
            "Email sending isn't configured yet. Add SMTP_HOST, SMTP_USER, SMTP_PASS "
            "(and SMTP_FROM) as environment variables in Render."}), 400
    data = request.json or {}
    to      = str(data.get("to", "")).strip()
    subject = str(data.get("subject", "")).strip()
    body    = str(data.get("body", "")).strip()
    reg     = str(data.get("reg", "")).strip()
    page    = str(data.get("page", "")).strip()
    caller  = str(data.get("caller", "")).strip() or "Unknown"
    if not to or not body:
        return jsonify({"ok": False, "error": "Missing recipient or message body"}), 400
    try:
        msg = MIMEMultipart()
        msg["From"]    = formataddr((SMTP_FROM_NAME, SMTP_FROM))
        msg["To"]      = to
        msg["Subject"] = subject or "(no subject)"
        msg.attach(MIMEText(body, "plain", "utf-8"))
        if SMTP_PORT == 465:
            srv = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30)
        else:
            srv = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
            srv.ehlo(); srv.starttls(); srv.ehlo()
        srv.login(SMTP_USER, SMTP_PASS)
        srv.sendmail(SMTP_FROM, [a.strip() for a in to.split(",") if a.strip()], msg.as_string())
        srv.quit()
    except Exception as e:
        return jsonify({"ok": False, "error": f"Send failed: {str(e)[:220]}"}), 500

    # Log to dashboard history + Maximizer Note
    logged = False; synced = 0
    if reg:
        ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        log = f"✉️ Email sent to {to}\nSubject: {subject}\n\n{body}"
        try:
            found = mx_find_by_org_number(reg)
            if found and found.get("key"):
                resp = mx_create_note(found["key"], log, caller, timestamp=ts)
                if resp and resp.get("Code") == 0: synced = 1
        except Exception as e:
            print(f"  email→mx note error: {e}")
        db_exec("INSERT INTO comment_history(reg_number,page,text,caller,timestamp,synced_to_maximizer) VALUES(?,?,?,?,?,?)",
                (reg, page, log, caller, ts, synced))
        logged = True
    return jsonify({"ok": True, "logged": logged, "mx_synced": synced})

@app.route("/api/call_history_all")
def call_history_all():
    """All calls with optional date-range / caller / outcome filters (history viewer)."""
    dfrom  = str(request.args.get("from","")).strip()
    dto    = str(request.args.get("to","")).strip()
    caller = str(request.args.get("caller","")).strip()
    outcome= str(request.args.get("outcome","")).strip()
    where=[]; params=[]
    if dfrom:  where.append("substr(timestamp,1,10)>=?"); params.append(dfrom)
    if dto:    where.append("substr(timestamp,1,10)<=?"); params.append(dto)
    if caller: where.append("caller=?"); params.append(caller)
    if outcome:where.append("outcome=?"); params.append(outcome)
    sql="SELECT reg_number,page,phone_used,outcome,notes,duration_sec,caller,timestamp FROM call_log"
    if where: sql+=" WHERE "+" AND ".join(where)
    sql+=" ORDER BY timestamp DESC LIMIT 2000"
    try:
        rows=db_query(sql, tuple(params))
    except Exception as e:
        return jsonify({"calls":[], "count":0, "error":str(e)[:120], "callers":[], "outcomes":[]})
    names={}
    for reg,nm in db_query("SELECT reg_number,name FROM called_log"):
        if nm: names[reg]=nm
    calls=[{"reg_number":r[0],"page":r[1],"phone":r[2],"outcome":r[3],"notes":r[4],
            "duration_sec":r[5] or 0,"caller":r[6],"timestamp":r[7],
            "name":names.get(r[0],"")} for r in rows]
    callers=[r[0] for r in db_query("SELECT DISTINCT caller FROM call_log WHERE caller IS NOT NULL AND caller<>'' ORDER BY caller")]
    outcomes=[r[0] for r in db_query("SELECT DISTINCT outcome FROM call_log WHERE outcome IS NOT NULL AND outcome<>'' ORDER BY outcome")]
    return jsonify({"calls":calls,"count":len(calls),"callers":callers,"outcomes":outcomes})

@app.route("/api/call_history_all/export")
def call_history_export():
    """Export full call history as CSV, including all Daily Calling main-page fields
    (charity details, contacts, income, source, stage) joined to each call."""
    dfrom  = str(request.args.get("from","")).strip()
    dto    = str(request.args.get("to","")).strip()
    caller = str(request.args.get("caller","")).strip()
    outcome= str(request.args.get("outcome","")).strip()
    where=[]; params=[]
    if dfrom:  where.append("cl.timestamp>=?"); params.append(dfrom+" 00:00:00")
    if dto:    where.append("cl.timestamp<=?"); params.append(dto+" 23:59:59")
    if caller: where.append("cl.caller=?"); params.append(caller)
    if outcome:where.append("cl.outcome=?"); params.append(outcome)
    sql = """SELECT cl.timestamp, cl.caller, cl.reg_number, cl.phone_used, cl.outcome,
                    cl.notes, cl.duration_sec, cl.synced_to_maximizer,
                    CASE WHEN cl.page='calling' THEN COALESCE(cb.category,'calling') ELSE cl.page END as src,
                    cb.data
             FROM call_log cl
             LEFT JOIN calling_batch cb ON cb.reg_number=cl.reg_number AND cb.active=1"""
    if where: sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY cl.timestamp DESC LIMIT 10000"
    try:
        rows = db_query(sql, tuple(params)) or []
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500
    # name lookup fallback from called_log
    names = {}
    for reg, nm in (db_query("SELECT reg_number,name FROM called_log") or []):
        if nm: names[reg] = nm

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "Call Date/Time","Caller","Charity Name","Reg Number","Source List","Stage",
        "Main Phone","Main Email","Website",
        "Contact Person","Job Role","Contact Phone","Contact Email",
        "Total Income (£)","Statutory (£)",
        "Call Outcome","Call Duration","Synced to Maximizer","Call Notes"
    ])
    for r in rows:
        ts, cllr, reg, phone_used, oc, notes, dur, synced, src, cbdata = r
        c = {}
        try:
            if cbdata: c = json.loads(cbdata)
        except Exception: c = {}
        name = c.get("name") or c.get("charity_name") or names.get(reg,"") or ""
        ov = _contact_overrides.get(reg, {}) or {}
        main_phone = ov.get("phone") or c.get("phone") or ""
        main_email = ov.get("email") or c.get("email") or ""
        website    = ov.get("website") or c.get("website") or ""
        # primary contact (first of the 3 if present)
        contacts = ov.get("contacts") or []
        if contacts:
            cp = contacts[0]
            con_name, con_role = cp.get("name",""), cp.get("role","")
            con_phone, con_email = cp.get("phone",""), cp.get("email","")
            # append extra contacts inline
            extras = []
            for x in contacts[1:]:
                seg = (x.get("name","") or "")
                if x.get("role"): seg += f" ({x['role']})"
                if x.get("phone"): seg += f" {x['phone']}"
                if x.get("email"): seg += f" {x['email']}"
                if seg.strip(): extras.append(seg.strip())
            if extras:
                con_name = con_name + "  ||  " + "  ||  ".join(extras)
        else:
            con_name  = ov.get("name","")
            con_role  = ov.get("role","")
            con_phone = ov.get("cphone","")
            con_email = ov.get("cemail","")
        stage = _statuses.get(f"calling|{reg}","") or _statuses.get(f"{src}|{reg}","")
        income = c.get("total_income","") or c.get("latest_income","")
        statutory = c.get("statutory","") or (c.get("financials",{}) or {}).get("statutory","")
        dn = int(dur or 0)
        w.writerow([
            (ts or "")[:16], cllr or "", name, reg, CALLING_CATEGORIES.get(src, src or ""), stage,
            main_phone, main_email, website,
            con_name, con_role, con_phone, con_email,
            income, statutory,
            oc or "", f"{dn//60}:{dn%60:02d}", ("Yes" if synced else "No"), (notes or "")
        ])
    ts_now = datetime.utcnow().strftime("%Y%m%d_%H%M")
    return Response(buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=call_history_{ts_now}.csv"})

@app.route("/api/call/history")
def call_history():
    """Get all calls for a charity, most recent first."""
    reg = str(request.args.get("reg","")).strip()
    if not reg: return jsonify({"calls": []})
    rows = db_query("""SELECT phone_used, outcome, notes, duration_sec, caller, timestamp, synced_to_maximizer
                       FROM call_log WHERE reg_number=? ORDER BY id DESC""", (reg,))
    return jsonify({"reg_number": reg, "calls": [
        {"phone": r[0], "outcome": r[1], "notes": r[2],
         "duration_sec": r[3], "caller": r[4],
         "timestamp": r[5], "synced": bool(r[6])}
        for r in rows
    ]})

@app.route("/api/comment/edit", methods=["POST"])
def comment_edit():
    """Edit an existing dashboard comment_history note. Maximizer notes cannot be edited."""
    data = request.json or {}
    note_id = data.get("id")
    new_text = str(data.get("text", "")).strip()
    if not note_id:
        return jsonify({"ok": False, "error": "Missing note id"}), 400
    if not new_text:
        return jsonify({"ok": False, "error": "Note text cannot be empty"}), 400
    # Verify the note exists and is a dashboard note (has an id in our table)
    row = db_query("SELECT reg_number FROM comment_history WHERE id=?", (note_id,))
    if not row:
        return jsonify({"ok": False, "error": "Note not found"}), 404
    # Mark as not-synced since the text changed (so it re-syncs to Maximizer)
    db_exec("UPDATE comment_history SET text=?, synced_to_maximizer=0 WHERE id=?", (new_text[:2000], note_id))
    return jsonify({"ok": True})

@app.route("/api/comment/history")
def comment_history():
    import re as _re
    reg = str(request.args.get("reg","")).strip()
    if not reg: return jsonify({"history": []})

    # Local DB history
    rows = db_query("SELECT id, text, caller, timestamp, synced_to_maximizer FROM comment_history WHERE reg_number=? ORDER BY id DESC", (reg,))
    local = [
        {"id": r[0], "text": r[1], "caller": r[2], "timestamp": r[3],
         "synced": bool(r[4]), "source": "dashboard"}
        for r in rows
    ]

    # Maximizer notes
    mx_notes = []
    try:
        found = mx_find_by_org_number(reg)
        if found and found.get("key"):
            notes = mx_read_notes(found["key"], top=50)
            for n in notes:
                rich = n.get("RichText","") or ""
                # Strip HTML for clean display
                clean = _re.sub(r"<br\s*/?>", "\n", rich)
                clean = _re.sub(r"</p>\s*<p[^>]*>", "\n", clean)
                clean = _re.sub(r"<[^>]+>", "", clean).strip()
                dt = n.get("DateTime","")
                # Try to extract caller name from "<strong>NAME</strong>" pattern
                caller_match = _re.search(r"<strong>([^<]+)</strong>", rich)
                caller = caller_match.group(1) if caller_match else (n.get("Creator","") or "Maximizer User")
                mx_notes.append({
                    "text": clean,
                    "caller": caller,
                    "timestamp": dt.replace("T"," ")[:19] if dt else "",
                    "synced": True,
                    "source": "maximizer",
                    "category": n.get("Category","")
                })
    except Exception as e:
        print(f"  history maximizer fetch error: {e}")

    # Merge - dashboard notes already in DB are duplicates of synced Maximizer notes
    # Deduplicate by matching text content
    seen_texts = set()
    merged = []
    # Show Maximizer notes first (they're the authoritative source for synced items)
    for item in mx_notes:
        key = item["text"].strip()[:100]
        seen_texts.add(key)
        merged.append(item)
    # Then any dashboard notes that aren't already represented
    for item in local:
        key = item["text"].strip()[:100]
        if key not in seen_texts:
            merged.append(item)
            seen_texts.add(key)

    # Sort by timestamp descending
    merged.sort(key=lambda x: x.get("timestamp",""), reverse=True)

    return jsonify({"reg_number": reg, "history": merged,
                    "dashboard_count": len(local),
                    "maximizer_count": len(mx_notes)})

@app.route("/api/status",methods=["POST"])
def save_status():
    data=request.json or {}; reg=str(data.get("reg_number","")).strip(); page=str(data.get("page","")).strip()
    status=str(data.get("status","")).strip()
    caller=str(data.get("caller","")).strip() or "Unknown"
    if not reg or not page: return jsonify({"ok":False}),400
    key=f"{page}|{reg}"
    if status:
        _statuses[key]=status
        db_exec("INSERT OR REPLACE INTO statuses(key,status) VALUES(?,?)", (key, status))
        if key not in _called_log:
            ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
            _called_log[key]={"reg_number":reg,"page":page,"called_by":caller,"timestamp":ts}
            db_exec("INSERT OR REPLACE INTO called_log(key,reg_number,page,name,called_by,timestamp,data) VALUES(?,?,?,?,?,?,?)",
                    (key, reg, page, "", caller, ts, "{}"))
    else:
        _statuses.pop(key,None)
        db_exec("DELETE FROM statuses WHERE key=?", (key,))
    return jsonify({"ok":True})

@app.route("/api/db/test_write")
def db_test_write():
    """Test writing and reading to confirm DB works end-to-end."""
    import time
    test_name = f"TEST_{int(time.time())}"
    result = {"test_name": test_name}

    # Write
    result["write_result"] = db_exec("INSERT OR IGNORE INTO users(name) VALUES(?)", (test_name,))

    # Read back
    rows = db_query("SELECT name FROM users WHERE name=?", (test_name,))
    result["read_back"] = [r[0] for r in rows] if rows else "NOT FOUND"

    # All users
    all_rows = db_query("SELECT name, added_at FROM users")
    result["all_users"] = [{"name": r[0], "added_at": r[1]} for r in all_rows]
    result["all_count"] = len(all_rows)

    # Cleanup
    db_exec("DELETE FROM users WHERE name=?", (test_name,))

    return jsonify(result)

@app.route("/api/db/stats")
def db_stats():
    """Show row counts for all DB tables."""
    if not TURSO_URL or not TURSO_TOKEN:
        return jsonify({"connected": False, "error": "TURSO_DATABASE_URL/TOKEN env vars not set"})
    stats = {"connected": True, "url": TURSO_URL[:60] + "…"}
    for tbl in ("users","called_log","comments","statuses","saved_searches"):
        try:
            r = db_query(f"SELECT COUNT(*) FROM {tbl}")
            stats[tbl] = r[0][0] if r else 0
        except Exception as e:
            stats[tbl] = f"err: {e}"
    return jsonify(stats)

@app.route("/api/db/export/<table>")
def db_export(table):
    """Export a table as JSON for backup or viewing."""
    if table not in ("users","called_log","comments","statuses","saved_searches"):
        return jsonify({"error": "Invalid table"}), 400
    # Get columns + rows via libsql_client ResultSet
    try:
        client = _new_conn()
        if not client:
            return jsonify({"error": "No DB connection"}), 500
        rs = client.execute(f"SELECT * FROM {table}")
        cols = list(rs.columns) if hasattr(rs, 'columns') else []
        rows = [dict(zip(cols, tuple(r))) for r in rs.rows]
        client.close()
        return jsonify({"table": table, "columns": cols,
                        "row_count": len(rows), "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e), "table": table}), 500

@app.route("/api/saved_searches",methods=["GET"])
def get_saved_searches(): return jsonify(list(_saved_searches.keys()))

@app.route("/api/saved_searches",methods=["POST"])
def save_search():
    data=request.json or {}; name=str(data.get("name","")).strip(); criteria=data.get("criteria",{})
    if not name: return jsonify({"ok":False}),400
    _saved_searches[name]=criteria
    db_exec("INSERT OR REPLACE INTO saved_searches(name,criteria) VALUES(?,?)", (name, json.dumps(criteria)))
    return jsonify({"ok":True})

# ── Column layout presets (Daily Calling / Call History — shareable) ────────────
@app.route("/api/column_presets", methods=["GET"])
def list_column_presets():
    """Return presets visible to the given caller: their own private ones, all
    public ones, and any 'selected' ones that name them."""
    caller = str(request.args.get("caller","")).strip()
    rows = db_query("SELECT id,name,config_json,created_by,created_at,visibility,allowed_users_json FROM column_presets") or []
    out = []
    for pid, name, cfg, creator, created_at, vis, allowed_json in rows:
        vis = vis or "private"
        if vis == "public":
            visible = True
        elif vis == "private":
            visible = (creator == caller)
        else:  # 'selected'
            try: allowed = json.loads(allowed_json) if allowed_json else []
            except Exception: allowed = []
            visible = (creator == caller) or (caller in allowed)
        if not visible: continue
        try: config = json.loads(cfg) if cfg else {}
        except Exception: config = {}
        out.append({"id": pid, "name": name, "config": config, "created_by": creator,
                    "created_at": created_at, "visibility": vis,
                    "allowed_users": (json.loads(allowed_json) if allowed_json else []),
                    "mine": (creator == caller)})
    out.sort(key=lambda x: x["created_at"] or "", reverse=True)
    return jsonify(out)

@app.route("/api/column_presets", methods=["POST"])
def save_column_preset():
    data = request.json or {}
    name = str(data.get("name","")).strip()
    config = data.get("config") or {}
    creator = str(data.get("created_by","")).strip() or "Unknown"
    visibility = str(data.get("visibility","private")).strip().lower()
    if visibility not in ("private","public","selected"): visibility = "private"
    allowed_users = data.get("allowed_users") or []
    if not name: return jsonify({"ok": False, "error": "Name is required"}), 400
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    db_exec("""INSERT INTO column_presets(name,config_json,created_by,created_at,visibility,allowed_users_json)
               VALUES(?,?,?,?,?,?)""",
            (name, json.dumps(config), creator, now, visibility, json.dumps(allowed_users)))
    return jsonify({"ok": True})

@app.route("/api/column_presets/<int:pid>", methods=["DELETE"])
def delete_column_preset(pid):
    caller = str(request.args.get("caller","")).strip()
    row = db_query("SELECT created_by FROM column_presets WHERE id=?", (pid,))
    if not row: return jsonify({"ok": False, "error": "Not found"}), 404
    if row[0][0] != caller:
        return jsonify({"ok": False, "error": "Only the creator can delete this preset"}), 403
    db_exec("DELETE FROM column_presets WHERE id=?", (pid,))
    return jsonify({"ok": True})

@app.route("/api/saved_searches/<name>",methods=["GET"])
def get_saved_search(name):
    if name not in _saved_searches: return jsonify({"error":"Not found"}),404
    return jsonify(_saved_searches[name])

@app.route("/api/saved_searches/<name>",methods=["DELETE"])
def delete_saved_search(name):
    _saved_searches.pop(name,None)
    db_exec("DELETE FROM saved_searches WHERE name=?", (name,))
    return jsonify({"ok":True})

CONTACT_HDRS=["Phone","Email","Address 1","Address 2","Town","County","Postcode",
              "Activities","Call Status","Comment","Called By","Call Timestamp"]

def get_contact_details(reg):
    COLS={"registered_charity_number","charity_contact_phone","charity_contact_email",
          "charity_contact_address1","charity_contact_address2","charity_contact_address3",
          "charity_contact_address4","charity_contact_postcode","charity_activities"}
    for row in stream_zip_csv(CHARITY_URL,COLS):
        if row.get("registered_charity_number","").strip()==str(reg).strip():
            return {"phone":row.get("charity_contact_phone",""),"email":row.get("charity_contact_email",""),
                    "address1":row.get("charity_contact_address1",""),"address2":row.get("charity_contact_address2",""),
                    "town":row.get("charity_contact_address3",""),"county":row.get("charity_contact_address4",""),
                    "postcode":row.get("charity_contact_postcode",""),"activities":row.get("charity_activities","")}
    return {}

def contact_row(reg,page):
    d=get_contact_details(reg); key=f"{page}|{reg}"; cal=_called_log.get(key,{})
    return[d.get("phone",""),d.get("email",""),d.get("address1",""),d.get("address2",""),
           d.get("town",""),d.get("county",""),d.get("postcode",""),d.get("activities",""),
           _statuses.get(key,""),_comments.get(key,""),cal.get("called_by",""),cal.get("timestamp","")]

def csv_response(rows,headers,filename):
    def gen():
        buf=io.StringIO(); w=csv.writer(buf)
        w.writerow(headers); yield buf.getvalue(); buf.seek(0); buf.truncate()
        for row in rows:
            w.writerow(row); yield buf.getvalue(); buf.seek(0); buf.truncate()
    return Response(stream_with_context(gen()),mimetype="text/csv",
                    headers={"Content-Disposition":f"attachment; filename={filename}"})

@app.route("/download/prospects")
def dl_prospects():
    lf=request.args.get("list","all"); cached=cache_get("prospects")
    if not cached: return "Load Prospects tab first.",400
    map_={"best":"Best Immediate","readiness":"Readiness Package","contract":"Contract Growth","retention":"Retention/Replacement"}
    items=[]; seen=set()
    if lf=="all":
        for k,l in map_.items():
            for c in cached["lists"].get(k,[]):
                if c["reg_number"] not in seen: seen.add(c["reg_number"]); items.append((c,l))
    else:
        l=map_.get(lf,lf)
        for c in cached["lists"].get(lf,[]): items.append((c,l))
    hdrs=["Charity Number","Charity Name","Website","Prospect List","Total Income","Total Expenditure",
          "Gov Grants","Gov Contracts","Statutory Total","Statutory %","Prev Statutory","Financial Year"]+CONTACT_HDRS
    rows=[]
    for c,label in items:
        f=c.get("financials",{}); reg=c.get("reg_number","")
        rows.append([reg,c.get("name",""),c.get("website",""),label,f.get("total_income",0),
                     f.get("total_expenditure",0),f.get("gov_grants",0),f.get("gov_contracts",0),
                     f.get("statutory",0),f.get("stat_pct",0),f.get("prev_statutory",0),f.get("fin_year","")]
                    +contact_row(reg,"prospects"))
    return csv_response(rows,hdrs,f"Statutory_Prospects_{datetime.utcnow().strftime('%Y%m%d')}.csv")

@app.route("/download/new_charities")
def dl_new():
    days=request.args.get("days","30"); cached=cache_get(f"new_{days}")
    if not cached: return "Load New Charities tab first.",400
    hdrs=["Charity Number","Charity Name","Date Registered","Website","Status"]+CONTACT_HDRS
    rows=[[c["reg_number"],c["name"],c["date_registered"],c["website"],c["status"]]
          +contact_row(c["reg_number"],"new") for c in cached["charities"]]
    return csv_response(rows,hdrs,f"New_Charities_{days}days_{datetime.utcnow().strftime('%Y%m%d')}.csv")

@app.route("/download/late_accounts")
def dl_late():
    cached=cache_get("late")
    if not cached: return "Load Late Accounts tab first.",400
    def sev(m): return "Severe" if m>12 else "Moderate" if m>=6 else "Recent"
    hdrs=["Charity Number","Charity Name","Due Date","Months Late","Latest Income","Severity"]+CONTACT_HDRS
    rows=[[c["reg_number"],c["name"],c["due_date"],c["months_late"],c["latest_income"],sev(c["months_late"])]
          +contact_row(c["reg_number"],"late") for c in cached["charities"]]
    return csv_response(rows,hdrs,f"Late_Accounts_{datetime.utcnow().strftime('%Y%m%d')}.csv")

@app.route("/download/old_charities")
def dl_old():
    cached=cache_get("old_charities")
    if not cached: return "Load Old Charities tab first.",400
    hdrs=["Charity Number","Charity Name","Website","Date Registered","Age (Years)","Latest Income","Financial Year"]+CONTACT_HDRS
    rows=[[c["reg_number"],c["name"],c["website"],c["date_registered"],c["age_years"],c["total_income"],c["fin_year"]]
          +contact_row(c["reg_number"],"old") for c in cached["charities"]]
    return csv_response(rows,hdrs,f"Old_Charities_{datetime.utcnow().strftime('%Y%m%d')}.csv")

@app.route("/api/debug")
def debug(): return jsonify({"status":dict(_bg_status),"cache":list(_cache.keys())})

# ── Users / Callers API ────────────────────────────────────────────────────────
@app.route("/api/users", methods=["GET"])
def get_users():
    """Always fetch fresh from DB so manual DB edits show immediately."""
    rows = db_query("SELECT name FROM users ORDER BY added_at")
    if rows:
        # Update in-memory cache too so other code stays in sync
        global _users
        _users = [r[0] for r in rows]
    return jsonify(_users)

@app.route("/api/db/refresh", methods=["GET","POST"])
def db_refresh():
    """Reload all in-memory state from DB. Call this after editing the DB manually."""
    try:
        db_load_state()
        return jsonify({"ok": True, "users": len(_users),
                        "statuses": len(_statuses),
                        "comments": len(_comments),
                        "called_log": len(_called_log),
                        "saved_searches": len(_saved_searches)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/users", methods=["POST"])
def add_user():
    data = request.json or {}
    name = str(data.get("name","")).strip()
    if not name or len(name) > 60:
        return jsonify({"ok":False,"error":"Invalid name"}), 400
    if name not in _users:
        _users.append(name)
        db_exec("INSERT OR IGNORE INTO users(name) VALUES(?)", (name,))
    return jsonify({"ok":True,"users":_users})

@app.route("/api/users/<name>", methods=["DELETE"])
def delete_user(name):
    if name in _users and name != "Muhanna":
        _users.remove(name)
        db_exec("DELETE FROM users WHERE name=?", (name,))
    return jsonify({"ok":True,"users":_users})

# ── Called log now stores caller name ─────────────────────────────────────────
@app.route("/api/mark_called_by", methods=["POST"])
def mark_called_by():
    """Mark called with a specific caller name. toggle=False means always set (JS handles toggle)."""
    data   = request.json or {}
    reg    = str(data.get("reg_number","")).strip()
    page   = str(data.get("page","")).strip()
    name   = str(data.get("name","")).strip()
    caller = str(data.get("caller","")).strip() or "Unknown"
    do_toggle = data.get("toggle", True)
    set_called = data.get("called", None)   # explicit desired state, if provided
    if not reg or not page:
        return jsonify({"ok":False,"error":"Missing reg_number or page"}), 400
    key = f"{page}|{reg}"
    # Un-mark: explicit called=False, or legacy toggle of an already-called row
    if set_called is False or (set_called is None and do_toggle and key in _called_log):
        if key in _called_log:
            del _called_log[key]
            db_exec("DELETE FROM called_log WHERE key=?", (key,))
        try:
            db_exec("UPDATE calling_batch SET completed=0 WHERE active=1 AND reg_number=?", (reg,))
        except Exception: pass
        return jsonify({"ok":True,"marked":False})
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    _called_log[key] = {
        "reg_number": reg, "name": name, "page": page,
        "called_by": caller, "timestamp": ts
    }
    db_exec("INSERT OR REPLACE INTO called_log(key,reg_number,page,name,called_by,timestamp,data) VALUES(?,?,?,?,?,?,?)",
            (key, reg, page, name, caller, ts, "{}"))
    # Also flag the active Daily-Calling batch row as completed so the X/100 progress
    # count and the "✓ done" chip stay consistent with the green called-row state.
    try:
        db_exec("UPDATE calling_batch SET completed=1 WHERE active=1 AND reg_number=?", (reg,))
    except Exception as e:
        print(f"  mark_called_by batch-complete error: {e}")
    return jsonify({"ok":True,"marked":True,"caller":caller})

# ── Milestone charities (approaching age anniversary with zero statutory) ──────
def compute_milestones(milestone_years):
    """Find charities reaching milestone_years around now.
    Normally: anniversary in the next 12 months.
    For the 1-year milestone we also include charities whose anniversary passed in the
    last 7 days (so callers can catch 'just turned 1') — i.e. a small look-back window."""
    today      = datetime.utcnow().date()
    # Upcoming window: registered between (today - years - 1yr) and (today - years)
    cutoff_hi = today.replace(year=today.year - milestone_years)
    cutoff_lo = today.replace(year=today.year - milestone_years - 1)
    # Small look-back: also catch anniversaries that passed in the last 7 days.
    # A charity that turned `years` exactly N days ago was registered `years` years
    # before (today - N). So extend the high cutoff forward by 7 days.
    from datetime import timedelta as _td
    cutoff_hi = cutoff_hi + _td(days=7)

    PC={"registered_charity_number","income_from_government_grants",
        "income_from_government_contracts","total_gross_income","fin_period_end_date"}
    CC={"registered_charity_number","charity_name","date_of_registration",
        "charity_registration_status","charity_contact_web",
        "charity_contact_phone","charity_contact_email",
        "charity_contact_address3","charity_contact_address4"}

    # Get parta latest per charity
    pl = {}
    for row in stream_zip_csv(PARTA_URL, PC):
        reg = row.get("registered_charity_number","").strip()
        if not reg: continue
        yr = row.get("fin_period_end_date","")[:10]
        if reg not in pl or yr > pl[reg].get("fin_period_end_date",""): pl[reg] = row

    results = []
    for row in stream_zip_csv(CHARITY_URL, CC):
        if row.get("charity_registration_status","").upper().strip() not in ("REGISTERED","R"): continue
        dt = row.get("date_of_registration","")[:10]
        if not dt or len(dt) < 10: continue
        try:
            reg_date = datetime.strptime(dt, "%Y-%m-%d").date()
        except: continue
        # Check if registration date falls in the milestone window
        if not (cutoff_lo < reg_date <= cutoff_hi): continue
        reg = row.get("registered_charity_number","").strip()
        if not reg: continue
        fin = pl.get(reg,{})
        gg  = float(fin.get("income_from_government_grants","") or 0)
        gc  = float(fin.get("income_from_government_contracts","") or 0)
        if gg + gc > 0: continue  # skip if has statutory income
        ti  = float(fin.get("total_gross_income","") or 0)
        # Calculate exact anniversary date
        try:
            anniversary = reg_date.replace(year=reg_date.year + milestone_years)
        except ValueError:
            anniversary = reg_date.replace(year=reg_date.year + milestone_years, day=28)
        months_to_go = round((anniversary - today).days / 30.4, 1)
        days_to_go = (anniversary - today).days
        web = row.get("charity_contact_web","")
        results.append({
            "reg_number":   reg,
            "name":         row.get("charity_name","") or reg,
            "website":      web if web not in ("nan","None") else "",
            "phone":        row.get("charity_contact_phone",""),
            "email":        row.get("charity_contact_email",""),
            "town":         row.get("charity_contact_address3",""),
            "county":       row.get("charity_contact_address4",""),
            "date_registered": dt,
            "age_years":    today.year - reg_date.year,
            "milestone":    milestone_years,
            "anniversary_date": str(anniversary),
            "months_to_anniversary": months_to_go,
            "days_to_anniversary": days_to_go,
            "just_completed": (-7 <= days_to_go < 0),   # anniversary passed in last 7 days
            "within_week":    (0 <= days_to_go <= 7),    # anniversary within next 7 days
            "total_income": ti,
            "fin_year":     fin.get("fin_period_end_date","")[:4],
        })
    results.sort(key=lambda x: x.get("anniversary_date",""))
    return {"charities": results, "count": len(results),
            "milestone": milestone_years, "generated": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}

# ── Daily Calling — top 100 priority-ranked contacts ────────────────────────────
import math as _math

def _called_today(reg):
    """True if this charity was called today (UTC), per call_log timestamps."""
    reg = str(reg)
    today = datetime.utcnow().strftime("%Y-%m-%d")
    for v in _called_log.values():
        if str(v.get("reg_number"))==reg and str(v.get("timestamp","")).startswith(today):
            return True
    return False

def _ever_called(reg):
    reg = str(reg)
    return any(str(v.get("reg_number"))==reg for v in _called_log.values())

def priority_score(c):
    """
    Composite 0–100 score blending Opportunity (40), Urgency (30), Freshness (30).
    Adapts to whatever fields a charity has — missing fields use neutral defaults
    so no list type is unfairly penalised. Returns (score, breakdown).
    """
    ti   = flt(c.get("total_income"))
    stat = flt(c.get("statutory"))
    pstat= flt(c.get("prev_statutory"))
    reg  = c.get("reg_number","")

    # ── Opportunity (0–40): deal size via log-scaled income + statutory exposure ──
    # £50k→~10pts, £500k→~25pts, £3m→~38pts (log curve). Statutory adds up to +12.
    opp = 0.0
    if ti > 0:
        opp = min(28.0, (_math.log10(max(ti,1)) - 4.0) * 14.0)   # log10(10k)=4 → 0
        opp = max(0.0, opp)
    if stat > 0:
        opp += min(12.0, (_math.log10(max(stat,1)) - 4.0) * 8.0)
        opp = min(40.0, opp)
    if ti <= 0 and stat <= 0:
        opp = 14.0   # neutral default for lists without financials (e.g. Newly Registered)

    # ── Urgency (0–30): declining trend, lateness, approaching anniversary ────────
    urg = 0.0
    if pstat > 0 and stat < pstat:                       # statutory funding declining
        drop = (pstat - stat) / pstat
        urg += min(16.0, drop * 16.0)
    ml = flt(c.get("months_late"))
    if ml > 0:                                           # late accounts
        urg += min(16.0, ml * 1.2)
    mta = c.get("months_to_anniversary")
    if mta is not None and mta != "":
        try:
            m = float(mta)
            if 0 <= m <= 12: urg += (12.0 - m) / 12.0 * 14.0   # closer = more urgent
        except: pass
    if c.get("spend_over_income"):
        urg += 8.0
    urg = min(30.0, urg)
    if urg == 0.0:
        urg = 6.0    # small neutral baseline

    # ── Freshness (0–30): uncalled floats to top; called-today drops ─────────────
    if _called_today(reg):
        fresh = 0.0
    elif _ever_called(reg):
        fresh = 12.0
    else:
        fresh = 30.0

    total = round(opp + urg + fresh, 1)
    return total, {"opportunity": round(opp,1), "urgency": round(urg,1), "freshness": fresh}

def _normalize_for_calling(c, source):
    """Map any list's charity dict into the common shape the calling page uses."""
    fin = c.get("financials",{}) or {}
    ti   = c.get("total_income", fin.get("total_income", c.get("latest_income","")))
    stat = fin.get("statutory","")
    pstat= fin.get("prev_statutory","")
    spend_over = any(f.get("label","").startswith("Spend") for f in fin.get("flags",[]))
    # Human-readable "which data from the source" label, shown in its own column
    # and fed to the AI insight so the caller knows why this charity is on the list.
    detail = ""
    try:
        if source == "Anniversary Watch" or c.get("milestone"):
            ms = c.get("milestone")
            if ms:
                yr = int(ms)
                yrtxt = "1 year" if yr == 1 else f"{yr} years"
                dta = c.get("days_to_anniversary")
                when = ""
                if c.get("within_week"): when = " · due within a week"
                elif c.get("just_completed"): when = " · just reached"
                elif isinstance(c.get("months_to_anniversary"), (int, float)):
                    m = c.get("months_to_anniversary")
                    if m is not None and m >= 0: when = f" · in ~{round(m)} mo"
                detail = f"🎂 {yrtxt} anniversary{when}"
        elif source == "Late Accounts" or c.get("months_late") not in ("", None):
            ml = c.get("months_late")
            try:
                mlv = float(ml)
                band = "Over 12 mo late" if mlv >= 12 else ("6–12 mo late" if mlv >= 6 else "Under 6 mo late")
                detail = f"⏰ {band} ({round(mlv)} mo)"
            except Exception:
                if ml: detail = f"⏰ {ml} months late"
        elif source == "Newly Registered":
            dr = str(c.get("date_registered","") or "")[:10]
            if dr:
                try:
                    from datetime import datetime as _dt
                    days = (_dt.utcnow().date() - _dt.strptime(dr, "%Y-%m-%d").date()).days
                    detail = f"📰 Registered {days} days ago"
                except Exception:
                    detail = f"📰 Registered {dr}"
        elif source == "40+ Yrs Zero Statutory":
            detail = "🏛️ 40+ yrs · no statutory income"
        elif source == "Prospect Lists":
            detail = "📋 Prospect list"
        elif source == "Charity Search":
            detail = "🔍 Saved search"
    except Exception:
        detail = ""
    return {
        "reg_number": c.get("reg_number",""),
        "name": c.get("name",""),
        "phone": c.get("phone",""),
        "email": c.get("email",""),
        "website": c.get("website",""),
        "town": c.get("town",""),
        "county": c.get("county",""),
        "total_income": ti,
        "statutory": stat,
        "prev_statutory": pstat,
        "months_late": c.get("months_late",""),
        "months_to_anniversary": c.get("months_to_anniversary",""),
        "milestone": c.get("milestone",""),
        "date_registered": c.get("date_registered",""),
        "days_to_anniversary": c.get("days_to_anniversary",""),
        "within_week": c.get("within_week", False),
        "just_completed": c.get("just_completed", False),
        "spend_over_income": spend_over,
        "removed": bool(c.get("removed", False)),
        "source": source,
        "source_detail": detail,
    }

def _collect_from_cache():
    """Gather charities from all cached lists, tagged by source. Returns dict source→list."""
    pools = {}
    # Prospects (nested lists)
    p = cache_get("prospects", max_age=86400)
    if p and p.get("lists"):
        merged = {}
        for sub in ("best","readiness","contract","retention"):
            for c in p["lists"].get(sub,[]):
                merged[c.get("reg_number")] = c   # dedupe across sublists
        pools["Prospect Lists"] = list(merged.values())
    # Newly Registered (default 30d window)
    for k in ("new_30","new_60","new_90"):
        n = cache_get(k, max_age=86400)
        if n and n.get("charities"):
            pools["Newly Registered"] = n["charities"]; break
    # Late Accounts
    l = cache_get("late", max_age=86400)
    if l and l.get("charities"):
        pools["Late Accounts"] = l["charities"]
    # 40+ Zero Statutory
    o = cache_get("old_charities", max_age=86400)
    if o and o.get("charities"):
        pools["40+ Yrs Zero Statutory"] = o["charities"]
    # Anniversary Watch — merge any cached milestone windows (incl. 1-year)
    anniv = {}
    for k in ("milestones_1","milestones_5","milestones_10","milestones_15","milestones_20",
              "milestones_25","milestones_30","milestones_40","milestones_45","milestones_50"):
        m = cache_get(k, max_age=86400)
        if m and m.get("charities"):
            for c in m["charities"]:
                anniv[c.get("reg_number")] = c
    if anniv:
        pools["Anniversary Watch"] = list(anniv.values())
    return pools

CALLING_CATEGORIES = {
    "all": "All combined",
    "prospects": "Prospect Lists",
    "new": "Newly Registered",
    "late": "Late Accounts",
    "old": "40+ Yrs Zero Statutory",
    "milestones": "Anniversary Watch",
    "search": "Charity Search (saved)",
}

def _rank_calling_pool(category, filters=None):
    """Return (scored_sorted_list, error_response_or_None) for a calling category."""
    filters = filters or {}
    pools = _collect_from_cache()
    if category == "search":
        saved_name = str(filters.get("saved_search","")).strip()
        if not saved_name:
            names = list(_saved_searches.keys())
            return [], jsonify({"charities":[], "count":0, "category":category,
                "category_label": CALLING_CATEGORIES.get(category,category),
                "note": ("Pick a saved search from the 'Saved Search' filter above, then Generate."
                         if names else
                         "No saved searches yet. Save one from the Charity Search (Advanced) page first.")})
        crit = _saved_searches.get(saved_name)
        if not isinstance(crit, dict):
            return [], jsonify({"charities":[], "count":0, "category":category,
                "category_label": CALLING_CATEGORIES.get(category,category),
                "note": f"Saved search '{saved_name}' was not found — it may have been deleted."})
        # Run the saved search live (saved searches only store filter criteria, not
        # results, so results always reflect the current Charity Commission data).
        # Daily Calling only ever uses the top 100, so a small bounded limit here is
        # plenty and keeps a broad saved search (e.g. a full year of registrations)
        # from pulling in thousands of rows unnecessarily.
        rows, _truncated = _run_advanced_search(crit, limit=300)
        if not rows:
            return [], jsonify({"charities":[], "count":0, "category":category,
                "category_label": CALLING_CATEGORIES.get(category,category),
                "note": f"Saved search '{saved_name}' returned no charities."})
        src_map = {f"Charity Search (saved: {saved_name})": rows}
    elif category == "all":
        src_map = pools
    else:
        label = CALLING_CATEGORIES.get(category)
        if label and label in pools:
            src_map = {label: pools[label]}
        else:
            return [], jsonify({"charities":[], "count":0, "category":category,
                "category_label": CALLING_CATEGORIES.get(category,category), "status":"not_loaded",
                "note":f"Open the '{CALLING_CATEGORIES.get(category,category)}' tab once to load its data, then return here."})
    if not src_map:
        return [], jsonify({"charities":[], "count":0, "category":category,
            "category_label": CALLING_CATEGORIES.get(category,category), "status":"not_loaded",
            "note":"No list data is loaded yet. Open the list tabs once to load them, then return here."})
    seen = {}
    for source, rows in src_map.items():
        for c in rows:
            reg = c.get("reg_number","")
            if not reg or reg in seen: continue
            if c.get("removed"): continue
            seen[reg] = _normalize_for_calling(c, source)
    scored = []
    for reg, c in seen.items():
        sc, breakdown = priority_score(c)
        c["priority"] = sc
        c["priority_breakdown"] = breakdown
        c["called_today"] = _called_today(reg)
        c["ever_called"] = _ever_called(reg)
        c["band"] = "high" if sc>=65 else ("med" if sc>=45 else "low")
        scored.append(c)
    scored.sort(key=lambda x: x["priority"], reverse=True)
    return scored, None

@app.route("/api/daily_calling")
def api_daily_calling():
    category = request.args.get("category","all").strip().lower()
    filters = {"saved_search": request.args.get("saved_search","")}
    scored, err = _rank_calling_pool(category, filters)
    if err is not None:
        return err
    top = scored[:100]
    called_today_count = sum(1 for c in top if c["called_today"])
    return jsonify({
        "charities": top,
        "count": len(top),
        "total_pool": len(scored),
        "called_today": called_today_count,
        "category": category,
        "category_label": CALLING_CATEGORIES.get(category, category),
        "generated": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
    })

def _load_active_batch():
    """Read the active batch from DB, freshen called/contact state, return list."""
    rows = db_query("SELECT reg_number,position,completed,data FROM calling_batch WHERE active=1 ORDER BY position ASC")
    outcomes = _latest_outcomes_map()
    retry_states = _retry_states_map()
    out = []
    for reg, pos, completed, data in rows:
        try: c = json.loads(data) if data else {}
        except Exception: c = {}
        c["reg_number"] = reg
        ov = _contact_overrides.get(reg) or {}
        if ov.get("phone"):   c["phone"]   = ov["phone"]
        if ov.get("email"):   c["email"]   = ov["email"]
        if ov.get("website"): c["website"] = ov["website"]
        c["completed"]    = int(completed or 0)
        c["called_today"] = _called_today(reg)
        c["ever_called"]  = _ever_called(reg)
        c["outcome"]      = outcomes.get(str(reg), "")
        rs = retry_states.get(str(reg)) or {}
        c["retry_count"]     = rs.get("retry_count", 0)
        c["retry_at"]        = rs.get("retry_at", "")
        c["retry_due"]       = rs.get("retry_due", False)
        c["retry_exhausted"] = rs.get("retry_exhausted", False)
        # Safe, additive-only repair: if the ground-truth call log shows this
        # charity WAS genuinely called today but the batch row's completed flag
        # doesn't reflect that (e.g. left over from an earlier bug), restore it.
        # This only ever SETS completed — it never clears/resets it — so it can
        # only repair missing data, never remove genuine progress.
        if c["called_today"] and not c["completed"]:
            c["completed"] = 1
            db_exec("UPDATE calling_batch SET completed=1 WHERE active=1 AND reg_number=?", (reg,))
        out.append(c)
    return out

@app.route("/api/call_history_list/export", methods=["POST"])
def api_call_history_list_export():
    """Export selected (or all) previously-called charities to Excel, with their
    charity details, contacts, last-called date, follow-up date, latest outcome and stage."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({"error": "openpyxl not installed. Add 'openpyxl' to requirements.txt and redeploy."}), 500
    data = request.get_json(force=True) or {}
    regs = [str(r).strip() for r in (data.get("regs") or []) if str(r).strip()]
    mode = str(data.get("mode","")).strip()  # "all" = use filters, else use regs
    # "All items across all pages" → recompute the full filtered list from the same
    # filters the History view is using, so the export isn't limited to the current page.
    if mode == "all" or not regs:
        flt = data.get("filters") or {}
        full = _build_call_history_rows(
            called_from=str(flt.get("called_from","")).strip(),
            called_to=str(flt.get("called_to","")).strip(),
            fu_from=str(flt.get("fu_from","")).strip(),
            fu_to=str(flt.get("fu_to","")).strip(),
            q=str(flt.get("q","")).strip(),
            outcome_filter=str(flt.get("outcome","")).strip())
        regs = [r["reg_number"] for r in full]
    if not regs:
        return jsonify({"error": "No charities to export"}), 400
    regset = set(regs)

    # Gather supporting data
    last_called, last_outcome = {}, {}
    for reg, ts, oc in (db_query("SELECT reg_number, timestamp, outcome FROM call_log ORDER BY timestamp ASC") or []):
        if reg in regset:
            last_called[reg] = str(ts)
            if oc: last_outcome[reg] = oc
    for reg, ts in (db_query("SELECT reg_number, MAX(timestamp) FROM called_log GROUP BY reg_number") or []):
        if reg in regset and ts and str(ts) > last_called.get(reg, ""):
            last_called[reg] = str(ts)
    followup = {}
    for reg, fat, done in (db_query("SELECT reg_number, follow_up_at, done FROM follow_ups ORDER BY follow_up_at ASC") or []):
        if reg in regset and not done and reg not in followup:
            followup[reg] = str(fat)
    batch_data, names = {}, {}
    for reg, d in (db_query("SELECT reg_number, data FROM calling_batch") or []):
        if reg in regset and reg not in batch_data:
            try:
                o = json.loads(d) if d else {}
                if o: batch_data[reg] = o
            except Exception: pass
    for reg, nm in (db_query("SELECT reg_number, name FROM called_log") or []):
        if reg in regset and nm: names[reg] = nm

    NAVY="1E3A5F"; thin=Side(style="thin", color="D5DAE0")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    wb = Workbook(); ws = wb.active; ws.title = "Call History"
    headers = ["Charity Name","Reg Number","Source","Source Detail","Last Called","Latest Outcome","Follow-up Due",
               "Stage","Main Phone","Main Email","Website",
               "Contact Person","Job Role","Contact Phone","Contact Email",
               "Total Income","Statutory"]
    for ci,h in enumerate(headers,1):
        c=ws.cell(row=1,column=ci,value=h)
        c.font=Font(name="Arial",bold=True,color="FFFFFF",size=11)
        c.fill=PatternFill("solid",fgColor=NAVY); c.border=border
        c.alignment=Alignment(horizontal="left",vertical="center")
    # Preserve the order the user is viewing (regs list order)
    # Reg → reconstructed source_detail (reuse the same logic the History view uses)
    sd_map = {}
    try:
        for r in _build_call_history_rows():
            sd_map[r["reg_number"]] = r.get("source_detail","")
    except Exception:
        pass
    r=2
    for reg in regs:
        base=batch_data.get(reg,{}) or {}
        name=base.get("name") or names.get(reg,"") or reg
        ov=_contact_overrides.get(reg,{}) or {}
        contacts=ov.get("contacts") or []
        if contacts:
            cp=contacts[0]; con_name,con_role,con_phone,con_email=cp.get("name",""),cp.get("role",""),cp.get("phone",""),cp.get("email","")
            extras=[]
            for x in contacts[1:]:
                seg=x.get("name","") or ""
                if x.get("role"): seg+=f" ({x['role']})"
                if x.get("phone"): seg+=f" {x['phone']}"
                if x.get("email"): seg+=f" {x['email']}"
                if seg.strip(): extras.append(seg.strip())
            if extras: con_name=con_name+"  ||  "+"  ||  ".join(extras)
        else:
            con_name,con_role,con_phone,con_email=ov.get("name",""),ov.get("role",""),ov.get("cphone",""),ov.get("cemail","")
        stage=_statuses.get(f"calling|{reg}","")
        vals=[name, reg, base.get("source",""), sd_map.get(reg,""), (last_called.get(reg,"") or "")[:16],
              last_outcome.get(reg,""), (followup.get(reg,"") or "")[:16], stage,
              ov.get("phone") or base.get("phone",""), ov.get("email") or base.get("email",""),
              ov.get("website") or base.get("website",""),
              con_name, con_role, con_phone, con_email,
              base.get("total_income",""), base.get("statutory","")]
        for ci,v in enumerate(vals,1):
            cell=ws.cell(row=r,column=ci,value=v)
            cell.font=Font(name="Arial",size=10); cell.border=border
        r+=1
    widths=[30,12,16,22,17,20,17,16,16,24,22,30,18,16,24,14,12]
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A2"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    tstamp=datetime.utcnow().strftime("%Y%m%d_%H%M")
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=call_history_{tstamp}.xlsx"})

def _build_call_history_rows(called_from="", called_to="", fu_from="", fu_to="", q="", outcome_filter=""):
    """Return the full filtered + sorted list of previously-called charities (no paging).
    q filters by charity name or reg number (case-insensitive substring), applied
    across the WHOLE history — not just whatever page happens to be loaded — so a
    search always finds a match regardless of the chosen page size."""
    # Last-called timestamp per charity
    last_called = {}
    for reg, ts in (db_query("SELECT reg_number, MAX(timestamp) FROM call_log GROUP BY reg_number") or []):
        if reg and ts: last_called[reg] = str(ts)
    for reg, ts in (db_query("SELECT reg_number, MAX(timestamp) FROM called_log GROUP BY reg_number") or []):
        if reg and ts:
            cur = last_called.get(reg, "")
            if str(ts) > cur: last_called[reg] = str(ts)
    # Follow-ups
    followups = {}
    for reg, fat, done in (db_query("SELECT reg_number, follow_up_at, done FROM follow_ups ORDER BY follow_up_at ASC") or []):
        if not reg: continue
        followups.setdefault(reg, {"next": None, "any": None})
        followups[reg]["any"] = str(fat)
        if not done and followups[reg]["next"] is None:
            followups[reg]["next"] = str(fat)
    regs = set(last_called.keys())
    if not regs: return []
    batch_data = {}
    for reg, data in (db_query("SELECT reg_number, data FROM calling_batch") or []):
        if reg in batch_data: continue
        try:
            o = json.loads(data) if data else {}
            if o: batch_data[reg] = o
        except Exception: pass
    names = {}
    for reg, nm in (db_query("SELECT reg_number, name FROM called_log") or []):
        if nm: names[reg] = nm
    outcomes = _latest_outcomes_map()
    retry_states = _retry_states_map()
    rows = []
    for reg in regs:
        lc = last_called.get(reg, "")
        fu = followups.get(reg, {})
        fu_next = fu.get("next")
        if called_from and (not lc or lc[:10] < called_from): continue
        if called_to   and (not lc or lc[:10] > called_to):   continue
        if fu_from or fu_to:
            fdate = (fu_next or fu.get("any") or "")[:10]
            if not fdate: continue
            if fu_from and fdate < fu_from: continue
            if fu_to   and fdate > fu_to:   continue
        base = dict(batch_data.get(reg, {}))
        base["reg_number"] = reg
        if not base.get("name"): base["name"] = names.get(reg, reg)
        ov = _contact_overrides.get(reg) or {}
        if ov.get("phone"):   base["phone"]   = ov["phone"]
        if ov.get("email"):   base["email"]   = ov["email"]
        if ov.get("website"): base["website"] = ov["website"]
        base["completed"]    = 1
        base["called_today"] = _called_today(reg)
        base["ever_called"]  = True
        base["last_called"]  = lc
        base["follow_up_at"] = fu_next or ""
        base["outcome"]        = outcomes.get(str(reg), "")
        base["latest_outcome"] = outcomes.get(str(reg), "")
        rs = retry_states.get(str(reg)) or {}
        base["retry_count"]     = rs.get("retry_count", 0)
        base["retry_at"]        = rs.get("retry_at", "")
        base["retry_due"]       = rs.get("retry_due", False)
        base["retry_exhausted"] = rs.get("retry_exhausted", False)
        # Ensure a Source Detail: use the stored one, else reconstruct from stored
        # fields (milestone / months_late / date_registered), else a light fallback.
        if not base.get("source_detail"):
            sd = ""
            try:
                ms = base.get("milestone")
                if ms:
                    yr = int(ms)
                    sd = f"🎂 {'1 year' if yr==1 else str(yr)+' years'} anniversary"
                elif base.get("months_late") not in ("", None):
                    mlv = float(base.get("months_late"))
                    bandl = "Over 12 mo late" if mlv>=12 else ("6–12 mo late" if mlv>=6 else "Under 6 mo late")
                    sd = f"⏰ {bandl}"
                elif base.get("date_registered"):
                    dr = str(base.get("date_registered"))[:10]
                    from datetime import datetime as _dt
                    days = (_dt.utcnow().date() - _dt.strptime(dr, "%Y-%m-%d").date()).days
                    if 0 <= days <= 400: sd = f"📰 Registered {days} days ago"
            except Exception:
                sd = ""
            base["source_detail"] = sd or ("📞 " + (base.get("source") or "Previously called"))
        # Name / reg number search — matches across the FULL history, not a page
        if q:
            ql = q.strip().lower()
            if ql not in str(base.get("name","")).lower() and ql not in str(reg).lower():
                continue
        # Outcome filter
        if outcome_filter:
            oc = base.get("outcome","")
            if outcome_filter == "__none__":
                if oc: continue
            elif oc != outcome_filter:
                continue
        rows.append(base)
    rows.sort(key=lambda x: x.get("last_called",""), reverse=True)
    return rows

@app.route("/api/call_history_list")
def api_call_history_list():
    """Charities that have EVER been called, in the same row shape as Daily Calling,
    so the History sub-tab can show them with full call functionality. Supports
    optional date filters (last-called range, follow-up range), a page size and offset."""
    try:
        called_from = str(request.args.get("called_from","")).strip()
        called_to   = str(request.args.get("called_to","")).strip()
        fu_from     = str(request.args.get("fu_from","")).strip()
        fu_to       = str(request.args.get("fu_to","")).strip()
        q           = str(request.args.get("q","")).strip()
        outcome_f   = str(request.args.get("outcome","")).strip()
        try: limit = min(1000, max(10, int(request.args.get("limit", "100"))))
        except: limit = 100
        try: offset = max(0, int(request.args.get("offset", "0")))
        except: offset = 0

        rows = _build_call_history_rows(called_from, called_to, fu_from, fu_to, q, outcome_f)
        total = len(rows)
        page_rows = rows[offset:offset+limit]
        return jsonify({"charities": page_rows, "count": len(page_rows),
                        "total_called": total, "total_filtered": total,
                        "offset": offset, "limit": limit})
    except Exception as e:
        print(f"  /api/call_history_list error: {e}")
        import traceback; traceback.print_exc()
        return jsonify({"charities": [], "count": 0, "total_called": 0,
                        "error": f"Could not load call history: {str(e)[:200]}"}), 200

@app.route("/api/calling_batch")
def api_calling_batch():
    """Return the current active calling batch (persisted, stable until regenerated)."""
    try:
        batch = _load_active_batch()
        if not batch:
            return jsonify({"charities": [], "count": 0, "has_batch": False,
                            "note": "No active calling list yet. Pick a source and generate your first 100."})
        meta = db_query("SELECT batch_no,created_at,category FROM calling_batch WHERE active=1 LIMIT 1")
        batch_no, created_at, category = (meta[0] if meta else (1, "", "all"))
        # A charity counts as "done" if it's been called TODAY, in this batch —
        # NOT simply "ever called", since a retry-due charity can be legitimately
        # ever_called (from days ago) while still being pending a fresh attempt
        # today. Using ever_called here would wrongly count retry-due charities
        # as done and make the X/Y counter (and green highlighting) misleading.
        # (The completed flag itself is kept in sync by the safe, additive-only
        # repair inside _load_active_batch — no further self-heal needed here.)
        done = sum(1 for c in batch if c.get("completed") or c.get("called_today"))
        return jsonify({
            "charities": batch, "count": len(batch), "has_batch": True,
            "batch_no": batch_no, "created_at": created_at, "category": category,
            "completed": done, "remaining": len(batch) - done,
            "category_label": CALLING_CATEGORIES.get(category, category),
        })
    except Exception as e:
        print(f"  /api/calling_batch error: {e}")
        import traceback; traceback.print_exc()
        return jsonify({"charities": [], "count": 0, "has_batch": False,
                        "error": f"Could not load calling list: {str(e)[:200]}"}), 200

@app.route("/api/calling_batch/status")
def api_calling_batch_status():
    rows = db_query("SELECT completed FROM calling_batch WHERE active=1")
    total = len(rows)
    done = sum(1 for r in rows if r[0])
    return jsonify({"has_batch": total > 0, "total": total,
                    "completed": done, "remaining": total - done,
                    "all_done": total > 0 and done >= total})

@app.route("/api/calling_batch/generate", methods=["POST"])
def api_calling_batch_generate():
    """Generate a new batch of 100. mode: 'fresh' (ignore leftovers),
    'leftovers' (carry uncompleted forward first), or 'new' (all done → next 100)."""
    data = request.json or {}
    category = str(data.get("category", "all")).strip().lower() or "all"
    mode = str(data.get("mode", "fresh")).strip().lower()
    filters = data.get("filters") or {}

    # When Anniversary Watch is the source AND a specific milestone year is chosen,
    # compute that milestone window fresh (like the Anniversary tab does) so the data
    # is always available even if it wasn't pre-cached. This fixes "1 year + within a
    # week returns nothing in Daily Calling" when milestones_1 hadn't been cached.
    if category == "milestones" and filters.get("years"):
        try:
            yr = int(filters.get("years"))
            fresh = compute_milestones(yr)
            rows = (fresh or {}).get("charities", []) if isinstance(fresh, dict) else (fresh or [])
            if rows:
                cache_set(f"milestones_{yr}", {"charities": rows,
                    "milestone": yr, "generated": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")})
        except Exception as e:
            print(f"  fresh milestone compute error: {e}")

    scored, err = _rank_calling_pool(category, filters)
    if err is not None:
        return err

    # Name-based exclusions (Women's Institute, schools, colleges, men's shed, Rotary, …)
    EXCLUDE_PATTERNS = {
        "wi":        ["womens institute", "woman institute"],
        "schools":   ["school"],
        "colleges":  ["college"],
        "mens_shed": ["mens shed", "man shed"],
        "rotary":    ["rotary"],
        "childcare": ["childcare", "child care", "nursery", "pre-school", "preschool", "playgroup", "play group"],
        "churches":  ["church", "parochial", "parish", "chapel", "cathedral", "methodist", "baptist", "catholic"],
        "trusts":    ["trust"],
        "school_pta":["pta", "parent teacher", "friends of"],
        "pfa":       ["pfa", "parents and friends", "parent and friends"],
        "foundation":["foundation"],
    }
    exclude_keys = data.get("exclude") or []
    excl_kw = []
    for k in exclude_keys:
        excl_kw += EXCLUDE_PATTERNS.get(str(k), [])
    def _norm(s):
        return (s or "").lower().replace("'", "").replace("\u2019", "")
    def _name_excluded(c):
        if not excl_kw: return False
        nm = _norm(c.get("name"))
        return any(kw in nm for kw in excl_kw)

    leftover_rows = db_query("SELECT reg_number,data FROM calling_batch WHERE active=1 AND completed=0 ORDER BY position ASC")
    leftovers = []
    for reg, d in leftover_rows:
        try: c = json.loads(d) if d else {}
        except Exception: c = {}
        c["reg_number"] = reg
        if _name_excluded(c): continue   # drop excluded leftovers too
        leftovers.append(c)
    leftover_regs = set(c["reg_number"] for c in leftovers)

    # A charity is eligible for a fresh list if it's never been called, OR its
    # retry cooldown has passed (retry_due). Charities still cooling down, or
    # that have exhausted their 3 retry attempts, stay excluded — exhausted ones
    # are shown separately (in light red) rather than silently vanishing.
    retry_states = _retry_states_map()
    def _retry_ok(reg):
        if not _ever_called(reg): return True
        rs = retry_states.get(str(reg))
        return bool(rs and rs.get("retry_due"))

    pool = [c for c in scored if _retry_ok(c["reg_number"]) and not _name_excluded(c) and c["reg_number"] not in _calling_excl]

    # ── Source-specific filters (Anniversary milestone/timing, Late months, New days) ──
    def _num(v, default=None):
        try: return float(v)
        except: return default
    def _passes_filters(c):
        # Anniversary milestone (exact years) + timing (months to anniversary)
        yrs = filters.get("years")
        if yrs:
            try:
                if int(c.get("milestone", 0) or 0) != int(yrs): return False
            except: pass
        timing = filters.get("timing")
        if timing:
            if timing == "within_week":
                if not c.get("within_week"):
                    # fallback to days_to_anniversary if flag not present
                    dta = _num(c.get("days_to_anniversary"), None)
                    if dta is None or not (0 <= dta <= 7): return False
            elif timing == "just_completed":
                if not c.get("just_completed"):
                    dta = _num(c.get("days_to_anniversary"), None)
                    if dta is None or not (-7 <= dta < 0): return False
            else:
                mta = _num(c.get("months_to_anniversary"), None)
                if mta is None or mta > _num(timing, 999): return False
        # Late accounts: months_late bands
        latem = filters.get("latemonths")
        if latem:
            ml = _num(c.get("months_late"), None)
            if ml is None: return False
            if latem == "over12" and ml < 12: return False
            if latem == "6to12" and not (6 <= ml < 12): return False
            if latem == "under6" and ml >= 6: return False
        # Newly registered: days since registration
        newdays = filters.get("newdays")
        if newdays:
            dr = str(c.get("date_registered","") or "")[:10]
            if not dr: return False
            try:
                from datetime import datetime as _dt
                days = (_dt.utcnow().date() - _dt.strptime(dr, "%Y-%m-%d").date()).days
                if days > int(newdays): return False
            except: return False
        return True
    if filters:
        pool = [c for c in pool if _passes_filters(c)]

    selected = []; used = set()
    if mode == "leftovers":
        for c in leftovers:
            if c["reg_number"] not in used:
                selected.append(c); used.add(c["reg_number"])
    excl = set(used)
    # "fresh" no longer excludes uncalled leftovers — they get included in the new 100
    # (only truly called items are excluded via _ever_called in the pool filter above)
    for c in pool:
        if len(selected) >= 100: break
        reg = c["reg_number"]
        if reg in excl or reg in used: continue
        selected.append(c); used.add(reg)

    selected = selected[:100]
    if not selected:
        return jsonify({"ok": False, "error": "No charities available to generate a new list for this source."}), 200

    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    bno_rows = db_query("SELECT COALESCE(MAX(batch_no),0) FROM calling_batch")
    batch_no = (bno_rows[0][0] if bno_rows else 0) + 1
    db_exec("UPDATE calling_batch SET active=0 WHERE active=1")
    for i, c in enumerate(selected):
        db_exec("INSERT INTO calling_batch(batch_no,reg_number,position,completed,created_at,category,active,data) VALUES(?,?,?,?,?,?,1,?)",
                (batch_no, c["reg_number"], i, 0, now, category, json.dumps(c)))
    return jsonify({"ok": True, "batch_no": batch_no, "count": len(selected),
                    "category": category, "carried_leftovers": (mode == "leftovers")})

@app.route("/api/milestones")
def api_milestones():
    years = int(request.args.get("years","40"))
    key   = f"milestones_{years}"
    r = bg_check(key, {"charities":[],"count":0,"milestone":years})
    if r: return r
    bg_run(key, lambda: compute_milestones(years))
    return jsonify({"status":"loading","message":f"Loading {years}-year milestone charities…",
                    "charities":[],"count":0,"milestone":years}), 202

@app.route("/download/milestones")
def dl_milestones():
    years = int(request.args.get("years","40"))
    cached = cache_get(f"milestones_{years}")
    if not cached: return "Load milestones tab first.", 400
    hdrs = ["Charity Number","Charity Name","Website","Date Registered","Current Age (Yrs)",
            f"Milestone ({years} yrs)","Anniversary Date","Months Until Anniversary",
            "Latest Income","Fin Year"] + CONTACT_HDRS
    rows = []
    for c in cached["charities"]:
        rows.append([c["reg_number"],c["name"],c["website"],c["date_registered"],
                     c["age_years"],f"{years} years",c["anniversary_date"],
                     c["months_to_anniversary"],c["total_income"],c["fin_year"]]
                    + contact_row(c["reg_number"],"milestones"))
    return csv_response(rows, hdrs, f"Milestones_{years}yr_{datetime.utcnow().strftime('%Y%m%d')}.csv")

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0",port=port,debug=False)


if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0",port=port,debug=False)

# ═══════════════════════════════════════════════════════════════════════
# MAXIMIZER CRM SYNC — Octopus API
# ═══════════════════════════════════════════════════════════════════════
MX_TOKEN = os.environ.get("MAXIMIZER_TOKEN",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiJteHB5ZjV6MGFwbWpub29jO"
    "XM3NCIsImlhdCI6MTc3ODUyNzEyMywiZXhwIjoxODczMDY1NjAwLCJteC1jaWQiOiJEMzIx"
    "RDMxRS04QzRBLTQyRTMtQUU5Ny03NjMzRDk5Qjk5QjIiLCJteC13c2lkIjoiOENDNkVBRkY"
    "tQkFERi00RDY5LTgyRTktNzQxREFERUU2QjU0IiwibXgtZGIiOiJjZmM0MWQwY2Y0OWQ0MjV"
    "iYjQ0ZTQ0YzA3ZDdiMTBmZCIsIm14LXVpZCI6Ik1BSEVTSCIsIm14LXBsIjoiY2xvdWQifQ"
    ".hEwCX0Yg35m_QoPa1yXiCoumJ-_9tP1OL3YM8MgOjMU"
)
MX_DB   = "cfc41d0cf49d425bb44e44c07d7b10fd"
MX_BASE = os.environ.get("MAXIMIZER_BASE","https://api.maximizer.com/octopus")

# ── UDF option maps from UDFOptionList.xml ────────────────────────────────────
# TYPEID(111)=What, (112)=Who, (113)=How, (107)=Country, (108)=LocalAuth,
# (109)=Region, (261)=Policy, (264)=IsSync

WHAT_MAP = {
    "accommodation/housing":"1","amateur sport":"2","animals":"3",
    "armed forces/emergency service efficiency":"4",
    "arts/culture/heritage/science":"5","disability":"6",
    "economic/community development/employment":"7","education/training":"8",
    "environment/conservation/heritage":"9","general charitable purposes":"10",
    "human rights/religious or racial harmony/equality or diversity":"11",
    "other charitable purposes":"12","overseas aid/famine relief":"13",
    "recreation":"14","religious activities":"15",
    "the advancement of health or saving of lives":"16",
    "the prevention or relief of poverty":"17",
}
WHO_MAP = {
    "children/young people":"1","elderly/old people":"2",
    "other charities or voluntary bodies":"3","other defined groups":"4",
    "people of a particular ethnic or racial origin":"5",
    "people with disabilities":"6","the general public/mankind":"7",
}
HOW_MAP = {
    "acts as an umbrella or resource body":"1",
    "makes grants to individuals":"2","makes grants to organisations":"3",
    "other charitable activities":"4",
    "provides advocacy/advice/information":"5",
    "provides buildings/facilities/open space":"6",
    "provides human resources":"7","provides other finance":"8",
    "provides services":"9","sponsors or undertakes research":"10",
}
REGION_MAP = {
    "throughout england":"1","throughout england and wales":"2",
    "throughout london":"3","throughout wales":"4",
}
LOCAL_AUTH_MAP = {
    "barking and dagenham":"1","barnet":"2","barnsley":"3",
    "bath and north east somerset":"4","bedford":"5","bexley":"6",
    "birmingham city":"7","blackburn with darwen":"8","blackpool":"9",
    "blaenau gwent":"10","bolton":"11","bournemouth":"12",
    "bracknell forest":"13","bradford city":"14","brent":"15",
    "bridgend":"16","brighton and hove":"17","bristol city":"18",
    "bromley":"19","buckinghamshire":"20","bury":"21","caerphilly":"22",
    "calderdale":"23","cambridgeshire":"24","camden":"25","cardiff":"26",
    "carmarthenshire":"27","central bedfordshire":"28","ceredigion":"29",
    "cheshire east":"30","cheshire west & chester":"31",
    "city of london":"32","city of swansea":"33","city of wakefield":"34",
    "city of westminster":"35","city of york":"36","conwy":"37",
    "cornwall":"38","coventry city":"39","croydon":"40","cumbria":"41",
    "darlington":"42","denbighshire":"43","derby city":"44",
    "derbyshire":"45","devon":"46","doncaster":"47","dorset":"48",
    "dudley":"49","durham":"50","ealing":"51",
    "east riding of yorkshire":"52","east sussex":"53","enfield":"54",
    "essex":"55","flintshire":"56","gateshead":"57","gloucestershire":"58",
    "greenwich":"59","gwynedd":"60","hackney":"61","halton":"62",
    "hammersmith and fulham":"63","hampshire":"64","haringey":"65",
    "harrow":"66","hartlepool":"67","havering":"68","herefordshire":"69",
    "hertfordshire":"70","hillingdon":"71","hounslow":"72",
    "isle of anglesey":"73","isle of wight":"74","isles of scilly":"75",
    "islington":"76","kensington and chelsea":"77","kent":"78",
    "kingston upon hull city":"79","kingston upon thames":"80",
    "kirklees":"81","knowsley":"82","lambeth":"83","lancashire":"84",
    "leeds city":"85","leicester city":"86","leicestershire":"87",
    "lewisham":"88","lincolnshire":"89","liverpool city":"90","luton":"91",
    "manchester city":"92","medway":"93","merthyr tydfil":"94",
    "merton":"95","middlesbrough":"96","milton keynes":"97",
    "monmouthshire":"98","neath port talbot":"99",
    "newcastle upon tyne city":"100","newham":"101","newport city":"102",
    "norfolk":"103","north east lincolnshire":"104",
    "north lincolnshire":"105","north somerset":"106",
    "north tyneside":"107","north yorkshire":"108",
    "northamptonshire":"109","northumberland":"110",
    "nottingham city":"111","nottinghamshire":"112","oldham":"113",
    "oxfordshire":"114","pembrokeshire":"115","peterborough city":"116",
    "plymouth city":"117","poole":"118","portsmouth city":"119",
    "powys":"120","reading":"121","redbridge":"122",
    "redcar and cleveland":"123","rhondda cynon taff":"124",
    "richmond upon thames":"125","rochdale":"126","rotherham":"127",
    "rutland":"128","salford city":"129","sandwell":"130","sefton":"131",
    "sheffield city":"132","shropshire":"133","slough":"134",
    "solihull":"135","somerset":"136","south gloucestershire":"137",
    "south tyneside":"138","southampton city":"139",
    "southend-on-sea":"140","southwark":"141","st helens":"142",
    "staffordshire":"143","stockport":"144","stockton-on-tees":"145",
    "stoke-on-trent city":"146","suffolk":"147","sunderland":"148",
    "surrey":"149","sutton":"150","swindon":"151","tameside":"152",
    "telford & wrekin":"153","thurrock":"154","torbay":"155",
    "torfaen":"156","tower hamlets":"157","trafford":"158",
    "vale of glamorgan":"159","walsall":"160","waltham forest":"161",
    "wandsworth":"162","warrington":"163","warwickshire":"164",
    "west berkshire":"165","west sussex":"166","wigan":"167",
    "wiltshire":"168","windsor and maidenhead":"169","wirral":"170",
    "wokingham":"171","wolverhampton":"172","worcestershire":"173",
    "wrexham":"174",
}
POLICY_MAP = {
    "bullying and harassment policy and procedures":"1",
    "conflicting interests":"2",
    "financial reserves policy and procedures":"3",
    "internal charity financial controls policy and procedures":"4",
    "internal risk management policy and procedures":"5",
    "investing charity funds policy and procedures":"6","investment":"7",
    "risk management":"8","safeguarding policy and procedures":"9",
    "safeguarding vulnerable beneficiaries":"10",
    "serious incident reporting policy and procedures":"11",
    "trustee conflicts of interest policy and procedures":"12",
    "volunteer management":"13","complaints handling":"14",
    "paying staff":"15",
    "campaigns and political activity policy and procedures":"16",
    "complaints policy and procedures":"17",
    "engaging external speakers at charity events policy and procedures":"18",
    "social media policy and procedures":"19",
    "trustee expenses policy and procedures":"20",
}

COUNTRY_MAP = {
    "abu dhabi":"1","afghanistan":"2","ajman":"3",
    "akrotiri":"4","aland islands":"5","albania":"6",
    "algeria":"7","american samoa":"8","andorra":"9",
    "angola":"10","anguilla":"11","antarctica":"12",
    "antigua and barbuda":"13","argentina":"14","armenia":"15",
    "aruba":"16","ascension":"17","australia":"18",
    "austria":"19","azerbaijan":"20","bahrain":"21",
    "baker island":"22","bangladesh":"23","barbados":"24",
    "belarus":"25","belgium":"26","belize":"27",
    "benin":"28","bermuda":"29","bhutan":"30",
    "bolivia":"31","bonaire":"32","bosnia and herzegovina":"33",
    "botswana":"34","bouvet island":"35","brazil":"36",
    "british antarctic territory":"37","british indian ocean territory":"38","british virgin islands":"39",
    "brunei":"40","bulgaria":"41","burkina faso":"42",
    "burma":"43","burundi":"44","cambodia":"45",
    "cameroon":"46","canada":"47","cape verde":"48",
    "cayman islands":"49","central african republic":"50","ceuta":"51",
    "chad":"52","chile":"53","china":"54",
    "christmas island":"55","cocos (keeling) islands":"56","colombia":"57",
    "comoros":"58","congo":"59","congo (democratic republic)":"60",
    "cook islands":"61","costa rica":"62","croatia":"63",
    "cuba":"64","cyprus":"65","czech republic":"66",
    "denmark":"67","dhekelia":"68","djibouti":"69",
    "dominica":"70","dominican republic":"71","dubai":"72",
    "east timor":"73","easter island":"74","ecuador":"75",
    "egypt":"76","el salvador":"77","equatorial guinea":"78",
    "eritrea":"79","estonia":"80","eswatini":"81",
    "ethiopia":"82","falkland islands":"83","faroe islands":"84",
    "fiji":"85","finland":"86","france":"87",
    "french guiana":"88","french polynesia":"89","french southern territories":"90",
    "fujairah":"91","gabon":"92","georgia":"93",
    "germany":"94","ghana":"95","gibraltar":"96",
    "greece":"97","greenland":"98","grenada":"99",
    "guadeloupe":"100","guam":"101","guatemala":"102",
    "guernsey":"103","guinea":"104","guinea-bissau":"105",
    "guyana":"106","haiti":"107","heard island and mcdonald islands":"108",
    "honduras":"109","hong kong":"110","howland island":"111",
    "hungary":"112","iceland":"113","india":"114",
    "indonesia":"115","iran":"116","iraq":"117",
    "ireland":"118","isle of man":"119","israel":"120",
    "italy":"121","ivory coast":"122","jamaica":"123",
    "japan":"124","jarvis island":"125","jersey":"126",
    "johnston atoll":"127","jordan":"128","kazakhstan":"129",
    "kenya":"130","kingman reef":"131","kiribati":"132",
    "kosovo":"133","kuwait":"134","kyrgyzstan":"135",
    "laos":"136","latvia":"137","lebanon":"138",
    "lesotho":"139","liberia":"140","libya":"141",
    "liechtenstein":"142","lithuania":"143","luxembourg":"144",
    "macau":"145","macedonia":"146","madagascar":"147",
    "malawi":"148","malaysia":"149","maldives":"150",
    "mali":"151","malta":"152","marshall islands":"153",
    "martinique":"154","mauritania":"155","mauritius":"156",
    "mayotte":"157","melilla":"158","mexico":"159",
    "micronesia":"160","midway islands":"161","moldova":"162",
    "monaco":"163","mongolia":"164","montenegro":"165",
    "montserrat":"166","morocco":"167","mozambique":"168",
    "namibia":"169","nauru":"170","navassa island":"171",
    "nepal":"172","netherlands":"173","netherlands antilles":"174",
    "new caledonia":"175","new zealand":"176","nicaragua":"177",
    "niger":"178","nigeria":"179","niue":"180",
    "norfolk island":"181","north korea":"182","northern ireland":"183",
    "northern mariana islands":"184","norway":"185","occupied palestinian territories":"186",
    "oman":"187","pakistan":"188","palau":"189",
    "palmyra atoll":"190","panama":"191","papua new guinea":"192",
    "paraguay":"193","peru":"194","philippines":"195",
    "pitcairn, henderson, ducie and oeno islands":"196","poland":"197","portugal":"198",
    "puerto rico":"199","qatar":"200","reunion":"201",
    "ras al-khaimah":"202","romania":"203","russia":"204",
    "rwanda":"205","são tomé and principe":"206","saba":"207",
    "saint barthélemy":"208","saint helena":"209","saint pierre and miquelon":"210",
    "saint vincent":"211","saint-martin":"212","samoa":"213",
    "san marino":"214","saudi arabia":"215","scotland":"216",
    "senegal":"217","serbia":"218","seychelles":"219",
    "sierra leone":"220","singapore":"221","sint eustatius":"222",
    "sint maarten":"223","slovakia":"224","slovenia":"225",
    "solomon islands":"226","somalia":"227","south africa":"228",
    "south georgia and south sandwich islands":"229","south korea":"230","south sudan":"231",
    "spain":"232","sri lanka":"233","st kitts and nevis":"234",
    "st lucia":"235","sudan":"236","suriname":"237",
    "svalbard and jan mayen":"238","sweden":"239","switzerland":"240",
    "syria":"241","taiwan":"242","tajikistan":"243",
    "tanzania":"244","thailand":"245","the bahamas":"246",
    "the gambia":"247","togo":"248","tokelau":"249",
    "tonga":"250","trinidad and tobago":"251","tristan da cunha":"252",
    "tunisia":"253","turkey":"254","turkmenistan":"255",
    "turks and caicos islands":"256","tuvalu":"257","uganda":"258",
    "ukraine":"259","umm al-quwain":"260","united arab emirates":"261",
    "united states":"262","united states virgin islands":"263","uruguay":"264",
    "uzbekistan":"265","vanuatu":"266","vatican city":"267",
    "venezuela":"268","vietnam":"269","wake island":"270",
    "wallis and futuna":"271","western sahara":"272","yemen":"273",
    "zambia":"274","zimbabwe":"275","ducie and oeno islands":"276",
    "henderson":"277","pitcairn":"278","réunion":"279",
    "saint barthélemy":"280",
}

GIFT_AID_MAP = {"no":"1","yes":"2"}

def _lookup(mapping, text):
    if not text: return None
    return mapping.get(str(text).lower().strip())

def _multi_keys(mapping, text_or_list):
    """Return list of option keys for multiple CC values."""
    if not text_or_list: return []
    if isinstance(text_or_list, str):
        items = [x.strip() for x in text_or_list.replace(';',',').split(',') if x.strip()]
    else:
        items = list(text_or_list)
    return [k for k in (_lookup(mapping, i) for i in items) if k]

# CC classification cache — loaded lazily
_classif_cache = {}
_classif_loaded = False

def load_classif_cache():
    """Load charity classification (what/who/how) from CC bulk file into memory."""
    global _classif_cache, _classif_loaded
    if _classif_loaded: return
    try:
        cols = {"registered_charity_number","classification_code","classification_type"}
        for row in stream_zip_csv(CLASSIF_URL, cols):
            reg  = row.get("registered_charity_number","").strip()
            code = row.get("classification_code","").strip()
            ctype = row.get("classification_type","").strip().lower()
            if not reg or not code: continue
            if reg not in _classif_cache:
                _classif_cache[reg] = {"what":[],"who":[],"how":[]}
            if ctype == "what":
                _classif_cache[reg]["what"].append(code)
            elif ctype == "who":
                _classif_cache[reg]["who"].append(code)
            elif ctype == "how":
                _classif_cache[reg]["how"].append(code)
        _classif_loaded = True
        print(f"Classif cache loaded: {len(_classif_cache)} charities")
    except Exception as e:
        print(f"load_classif_cache error: {e}")
        _classif_loaded = True  # Don't retry on error



# ── CC API — full charity detail lookup ──────────────────────────────────────
_cc_detail_cache = {}

def fetch_cc_charity(reg_no):
    """Fetch full charity details from CC API: who/what/how, address, dates, etc."""
    reg_no = str(reg_no)
    if reg_no in _cc_detail_cache:
        return _cc_detail_cache[reg_no]
    hdrs = {"Ocp-Apim-Subscription-Key": API_KEY}
    try:
        r = requests.get(f"{CC_API_BASE}/allcharitydetails/{reg_no}/0",
                         headers=hdrs, timeout=15)
        if r.status_code == 200:
            data = r.json()
            _cc_detail_cache[reg_no] = data
            return data
        print(f"  fetch_cc_charity({reg_no}): HTTP {r.status_code}")
    except Exception as e:
        print(f"  fetch_cc_charity({reg_no}): {e}")
    return {}

_cc_policy_cache = {}  # per-charity live policy lookup cache

def fetch_cc_policy(reg_no):
    """Fetch a single charity's policies from the CC API (a separate endpoint from
    allcharitydetails). Used for one charity at a time — e.g. Daily Calling AI
    insight or a Maximizer sync trigger — never for bulk search exports, which use
    the bounded bulk-file scan instead. The exact response shape isn't confirmed,
    so we defensively look for text anywhere in the response that matches a known
    policy label, the same value-matching approach used for the bulk export."""
    reg_no = str(reg_no)
    if reg_no in _cc_policy_cache:
        return _cc_policy_cache[reg_no]
    hdrs = {"Ocp-Apim-Subscription-Key": API_KEY}
    found = []
    try:
        r = requests.get(f"{CC_API_BASE}/charitypolicyinformation/{reg_no}/0",
                         headers=hdrs, timeout=15)
        if r.status_code == 200:
            data = r.json()
            items = data if isinstance(data, list) else ([data] if isinstance(data, dict) else [])
            for item in items:
                if not isinstance(item, dict): continue
                for val in item.values():
                    if isinstance(val, str) and val.strip().lower() in KNOWN_POLICY_LABELS:
                        if val.strip() not in found: found.append(val.strip())
        else:
            print(f"  fetch_cc_policy({reg_no}): HTTP {r.status_code}")
    except Exception as e:
        print(f"  fetch_cc_policy({reg_no}): {e}")
    _cc_policy_cache[reg_no] = found
    return found

def enrich_charity_from_cc(c):
    """Enrich a charity dict with full CC API data."""
    reg_no = str(c.get("reg_number","")).strip()
    if not reg_no:
        return c
    cc = fetch_cc_charity(reg_no)
    if not cc or not isinstance(cc, dict):
        return c
    # Basic contact fields
    if not c.get("phone")   and cc.get("phone"):   c["phone"]   = cc["phone"]
    if not c.get("email")   and cc.get("email"):   c["email"]   = cc["email"]
    if not c.get("website") and cc.get("web"):     c["website"] = cc["web"]
    # Address lines
    c["address1"]  = cc.get("address_line_one") or ""
    c["address2"]  = cc.get("address_line_two") or ""
    c["address3"]  = cc.get("address_line_three") or ""
    c["city"]      = (cc.get("address_line_five") or "") or (cc.get("address_line_four") or "")
    c["county"]    = cc.get("address_line_four") or ""
    c["postcode"]  = cc.get("address_post_code") or ""
    # Registration details
    dor = cc.get("date_of_registration") or ""
    if dor: c["date_of_registration"] = dor[:10]
    c["organisation_number"] = str(cc.get("organisation_number") or "")
    c["linked_charity"] = "Yes" if (cc.get("group_subsid_suffix") or 0) != 0 else "No"
    # Financials
    c["latest_income"]      = cc.get("latest_income") or ""
    c["latest_expenditure"] = cc.get("latest_expenditure") or ""
    c["fin_year_start"]     = (cc.get("latest_acc_fin_year_start_date") or "")[:10]
    c["fin_year_end"]       = (cc.get("latest_acc_fin_year_end_date") or "")[:10]
    # What / Who / How from who_what_where array
    wwh = cc.get("who_what_where") or []
    if not isinstance(wwh, list): wwh = []
    def _wwh(kind):
        out=[]
        for x in wwh:
            if isinstance(x, dict) and str(x.get("classification_type","")).lower()==kind:
                d=x.get("classification_desc")
                if d: out.append(str(d))
        return out
    whats, whos, hows = _wwh("what"), _wwh("who"), _wwh("how")
    if whats: c["what"] = ",".join(whats)
    if whos:  c["who"]  = ",".join(whos)
    if hows:  c["how"]  = ",".join(hows)
    # Local Authority / Region / Country — embedded arrays in the same response
    la_list = cc.get("CharityAoOLocalAuthority") or []
    if isinstance(la_list, list) and la_list and isinstance(la_list[0], dict):
        c["local_authority"] = la_list[0].get("local_authority","")
    region_list = cc.get("CharityAoORegion") or []
    if isinstance(region_list, list) and region_list:
        names = [r.get("region","") for r in region_list if isinstance(r, dict) and r.get("region")]
        if names: c["region"] = ", ".join(dict.fromkeys(names))
    country_list = cc.get("CharityAoOCountryContinent") or []
    if isinstance(country_list, list) and country_list:
        names = [r.get("country","") for r in country_list if isinstance(r, dict) and r.get("country")]
        if names: c["country"] = ", ".join(dict.fromkeys(names))
    # Gift Aid / Land & property — try a few plausible field names defensively,
    # since these aren't confirmed in the live API response the way others are.
    ga = cc.get("gift_aid")
    if ga is None: ga = cc.get("charity_gift_aid")
    if ga is not None: c["gift_aid"] = "Yes" if ga in (True,"true","True","Y","1",1) else "No"
    hl = cc.get("has_land")
    if hl is None: hl = cc.get("charity_has_land")
    if hl is not None: c["has_land"] = "Yes" if hl in (True,"true","True","Y","1",1) else "No"
    # Policy — separate live endpoint (not part of allcharitydetails)
    pol = fetch_cc_policy(reg_no)
    if pol: c["policy"] = ", ".join(pol)
    print(f"  CC enriched: what={(c.get('what') or '')[:30]} who={(c.get('who') or '')[:20]} "
          f"la={c.get('local_authority') or ''} region={c.get('region') or ''} "
          f"country={c.get('country') or ''} policy={(c.get('policy') or '')[:30]} "
          f"addr={(c.get('address1') or '')[:20]}")
    return c

# ── CC classification cache (what/who/how) ────────────────────────────────────
_classif_cache = {}
_classif_loaded = False

def load_classif_cache():
    global _classif_cache, _classif_loaded
    if _classif_loaded: return
    try:
        cols = {"registered_charity_number","classification_code","classification_type"}
        for row in stream_zip_csv(CLASSIF_URL, cols):
            reg   = row.get("registered_charity_number","").strip()
            code  = row.get("classification_code","").strip()
            ctype = row.get("classification_type","").strip().lower()
            if not reg or not code: continue
            if reg not in _classif_cache:
                _classif_cache[reg] = {"what":[],"who":[],"how":[]}
            if "what" in ctype:   _classif_cache[reg]["what"].append(code)
            elif "who" in ctype:  _classif_cache[reg]["who"].append(code)
            elif "how" in ctype:  _classif_cache[reg]["how"].append(code)
        _classif_loaded = True
        print(f"Classif cache: {len(_classif_cache)} charities")
    except Exception as e:
        print(f"load_classif_cache: {e}")
        _classif_loaded = True

# CC classification code → name (from CC bulk schema)
CC_CODE_MAP = {
    # What — per Charity Commission API reference data (classification_type='What')
    "101":"general charitable purposes","102":"education/training",
    "103":"the advancement of health or saving of lives","104":"disability",
    "105":"the prevention or relief of poverty","106":"overseas aid/famine relief",
    "107":"accommodation/housing","108":"religious activities",
    "109":"arts/culture/heritage/science","110":"amateur sport","111":"animals",
    "112":"environment/conservation/heritage",
    "113":"economic/community development/employment",
    "114":"armed forces/emergency service efficiency",
    "115":"human rights/religious or racial harmony/equality or diversity",
    "116":"recreation","117":"other charitable purposes",
    # Who
    "201":"children/young people","202":"elderly/old people",
    "203":"people with disabilities",
    "204":"people of a particular ethnic or racial origin",
    "205":"other charities or voluntary bodies","206":"other defined groups",
    "207":"the general public/mankind",
    # How
    "301":"makes grants to individuals","302":"makes grants to organisations",
    "303":"provides other finance","304":"provides human resources",
    "305":"provides buildings/facilities/open space","306":"provides services",
    "307":"provides advocacy/advice/information","308":"sponsors or undertakes research",
    "309":"acts as an umbrella or resource body","310":"other charitable activities",
}

def get_charity_classification(reg_no):
    """Get what/who/how strings for a charity reg number."""
    load_classif_cache()
    entry = _classif_cache.get(str(reg_no), {})
    def codes_to_str(codes):
        return ",".join(CC_CODE_MAP.get(c, c) for c in codes if c)
    return {
        "what": codes_to_str(entry.get("what",[])),
        "who":  codes_to_str(entry.get("who",[])),
        "how":  codes_to_str(entry.get("how",[])),
    }

# ── Bounded, on-demand enrichment for CSV export only ────────────────────────────
# IMPORTANT: earlier versions of this code cached the ENTIRE national Region/Local
# Authority/Country and Policy datasets permanently in memory (hundreds of thousands
# of charities), triggered the first time any search ran. Combined with the existing
# classification cache, this pushed the server past Render's 512MB free-tier limit
# and caused repeated out-of-memory crashes. This version instead does a single
# bounded, targeted pass per export — filtered down to just the charities actually
# being exported (capped) — and never retains data for charities outside that set.
# This only bounds an extreme edge case (e.g. exporting the entire national
# register with no filters at all). The retained memory for a normal export scales
# with the export size itself (a few bytes per charity), not with this ceiling, so
# it's set generously high rather than tightly — the earlier 1,000 cap was overly
# cautious for a mechanism that was already safe well beyond that.
EXPORT_ENRICH_MAX = 25000  # hard cap on how many charities one export can enrich

def _scan_classification_for(reg_set):
    """What / Who / How — from the classification bulk file, filtered to reg_set."""
    out = {}
    try:
        cols = {"registered_charity_number","classification_code","classification_type"}
        tmp = {}
        for row in stream_zip_csv(CLASSIF_URL, cols):
            reg = row.get("registered_charity_number","").strip()
            if reg not in reg_set: continue
            ctype = row.get("classification_type","").strip().lower()
            code = row.get("classification_code","").strip()
            if not code: continue
            tmp.setdefault(reg, {"what":[],"who":[],"how":[]})
            if ctype in ("what","who","how"):
                tmp[reg][ctype].append(CC_CODE_MAP.get(code, code))
        for reg, d in tmp.items():
            out[reg] = {"what": ",".join(dict.fromkeys(d["what"])),
                        "who":  ",".join(dict.fromkeys(d["who"])),
                        "how":  ",".join(dict.fromkeys(d["how"]))}
    except Exception as e:
        print(f"enrich_regs_for_export classification error: {e}")
    return out

def _scan_area_for(reg_set):
    """Region / Local Authority / Country — from the area-of-operation bulk file."""
    out = {}
    try:
        cols = {"registered_charity_number","geographic_area_type","geographic_area_description"}
        tmp = {}
        for row in stream_zip_csv(AREA_URL, cols):
            reg = row.get("registered_charity_number","").strip()
            if reg not in reg_set: continue
            atype = row.get("geographic_area_type","").strip().lower()
            adesc = row.get("geographic_area_description","").strip()
            if not adesc: continue
            tmp.setdefault(reg, {"region":[], "local_authority":[], "country":[]})
            bucket = "region" if atype=="region" else ("local_authority" if atype=="local authority" else ("country" if atype=="country" else None))
            if bucket and adesc not in tmp[reg][bucket]:
                tmp[reg][bucket].append(adesc)
        for reg, d in tmp.items():
            out[reg] = {"region": ", ".join(d["region"]),
                        "local_authority": ", ".join(d["local_authority"]),
                        "country": ", ".join(d["country"])}
    except Exception as e:
        print(f"enrich_regs_for_export area error: {e}")
    return out

# The Charity Commission's fixed set of policy category labels (confirmed from a
# charity's own XML annual-return export). Used to detect which column in the
# policy bulk file actually holds this data, since the column NAME itself was
# never confirmed — the exact column can vary, but the VALUES are a fixed set.
KNOWN_POLICY_LABELS = {
    "bullying and harassment policy and procedures",
    "conflicting interests",
    "financial reserves policy and procedures",
    "internal charity financial controls policy and procedures",
    "internal risk management policy and procedures",
    "investing charity funds policy and procedure",
    "investing charity funds policy and procedures",
    "investment",
    "risk management",
    "safeguarding policy and procedures",
    "safeguarding vulnerable beneficiaries",
    "serious incident reporting policy and procedures",
    "trustee conflicts of interest policy and procedures",
    "volunteer management",
    "complaints handling",
    "paying staff",
    "campaigns and political activity policy and procedures",
    "complaints policy and procedures",
    "engaging external speakers at charity events policy and procedures",
    "social media policy and procedures",
    "trustee expenses policy and procedures",
}

def _scan_policy_for(reg_set):
    """Policy — from the policy bulk file. The exact column name was never
    confirmed, so instead of guessing a name, we request every column and detect
    which one actually holds policy text by matching its values against the known
    set of Charity Commission policy labels."""
    out = {}
    try:
        tmp = {}
        policy_col = None
        rows_seen = 0
        sample_row = None
        for row in stream_zip_csv(POLICY_URL, None):  # None = keep every column
            rows_seen += 1
            if sample_row is None: sample_row = row
            reg = row.get("registered_charity_number","").strip()
            if reg not in reg_set: continue
            desc = ""
            # Once we've identified the right column, use it directly (fast path)
            if policy_col and row.get(policy_col, "").strip().lower() in KNOWN_POLICY_LABELS:
                desc = row[policy_col].strip()
            else:
                # Find the column whose value matches a known policy label
                for col, val in row.items():
                    if val and val.strip().lower() in KNOWN_POLICY_LABELS:
                        desc = val.strip()
                        if not policy_col:
                            policy_col = col
                            print(f"  [policy] detected policy column: {col!r}")
                        break
            if not desc: continue
            tmp.setdefault(reg, [])
            if desc not in tmp[reg]: tmp[reg].append(desc)
        for reg, lst in tmp.items():
            out[reg] = {"policy": ", ".join(lst)}
        if not tmp:
            print(f"  [policy] WARNING: no matches found among {rows_seen} rows scanned. "
                  f"Sample row columns/values: {sample_row}")
    except Exception as e:
        print(f"enrich_regs_for_export policy error: {e}")
    return out

def enrich_regs_for_export(regs):
    """Look up What/Who/How, Region/Local Authority/Country, and Policy for a
    bounded list of specific charities (used only when exporting search results to
    CSV — never for a live/on-screen search). Streams each CC bulk file once and
    keeps data only for the requested reg numbers, so memory stays proportional to
    the export size rather than the whole national register. The three source
    files are independent, so they're scanned in parallel to keep total wait time
    close to the slowest single file rather than the sum of all three."""
    reg_set = {str(r).strip() for r in regs if str(r).strip()}
    out = {r: {"what":"","who":"","how":"","region":"","local_authority":"","country":"","policy":""} for r in reg_set}
    if not reg_set:
        return out
    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=3) as ex:
        fut_cls = ex.submit(_scan_classification_for, reg_set)
        fut_area = ex.submit(_scan_area_for, reg_set)
        fut_pol = ex.submit(_scan_policy_for, reg_set)
        for reg, d in fut_cls.result().items(): out[reg].update(d)
        for reg, d in fut_area.result().items(): out[reg].update(d)
        for reg, d in fut_pol.result().items(): out[reg].update(d)
    return out

def mx_hdrs():
    return {"Authorization":f"Bearer {MX_TOKEN}",
            "Content-Type":"application/json","Accept":"application/json"}

def mx_call(endpoint, body):
    url = f"{MX_BASE.rstrip('/')}/{endpoint}"
    # Add Compatibility header required by Octopus API
    if "Compatibility" not in body:
        body["Compatibility"] = {"AbEntryKey": "2.0"}
    r = requests.post(url, headers=mx_hdrs(), json=body, timeout=20)
    if not r.ok: raise Exception(f"{r.status_code}: {r.text[:200]}")
    return r.json()

def mx_read(search_query=None, fields=None, top=1, entry_type=None):
    """Read AbEntry — includes Compatibility header to make Company entries visible."""
    sq = dict(search_query or {})
    if entry_type:
        sq["type"] = entry_type
    body = {
        "abEntry": {
            "criteria": {"searchQuery": sq, "top": top},
            "scope": {"fields": fields or {"key":1,"companyName":1}}
        },
        "Compatibility": {"AbEntryKey": "2.0"}
    }
    return mx_call("AbEntryRead", body)

def mx_write_create(data_dict):
    resp = mx_call("AbEntryCreate",{"AbEntry":{"Data":data_dict}})
    # Extract key from response: resp["AbEntry"]["Data"]["Key"]
    try:
        key = resp.get("AbEntry",{}).get("Data",{}).get("Key","")
        if key:
            print(f"  AbEntryCreate key: {key[:30]}")
            resp["_extracted_key"] = key
    except: pass
    return resp

def _extract_key_from_create_resp(resp):
    """Try every possible location for the key in an AbEntryCreate response."""
    import json
    # Log full response for debugging
    print(f"  Create resp: {json.dumps(resp)[:300]}")
    # Try all known locations
    candidates = [
        resp.get("AbEntry",{}).get("Data",{}).get("Key",""),
        resp.get("abEntry",{}).get("Data",{}).get("Key",""),
        resp.get("Data",{}).get("Key",""),
        resp.get("Key",""),
    ]
    # Also check if Data is a list
    data = resp.get("AbEntry",resp.get("abEntry",{})).get("Data",{})
    if isinstance(data, list) and data:
        candidates.append(data[0].get("Key",""))
    for c in candidates:
        if c: return c
    return ""

def mx_write_update(data_dict):
    return mx_call("AbEntryUpdate",{"AbEntry":{"Data":data_dict}})

def title_case(s):
    if not s: return s
    minor = {"a","an","the","and","or","but","of","in","on","at","to","for",
             "by","with","from","as","its"}
    acronyms = {"uk","cio","nhs","ymca","rspca","rspb","rnli","rnib","nspcc",
                "hmrc","plc","ltd","raf","rbl"}
    words = s.strip().split()
    result = []
    for i, w in enumerate(words):
        wl = w.lower().strip("',.")
        if wl in acronyms: result.append(w.upper())
        elif i==0 or wl not in minor: result.append(w.capitalize())
        else: result.append(w.lower())
    return " ".join(result)

# Set via env var once TYPEID is discovered via /probe_org_number_typeid
MX_ORG_NUM_TYPEID = os.environ.get("MX_ORG_NUM_TYPEID","")

def mx_find_by_org_number(reg_no):
    """Find Company entry by TYPEID(114) — confirmed working format."""
    import requests as req
    try:
        reg_int = int(str(reg_no).strip())
    except:
        return None
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    body = {
        "AbEntry": {
            "Scope": {"Fields": {"Key": 1, "CompanyName": 1, "/AbEntry/Udf/$TYPEID(114)": 1}},
            "Criteria": {"SearchQuery": {"$AND": [
                {"Type": {"$EQ": "Company"}},
                {"/AbEntry/Udf/$TYPEID(114)": {"$EQ": reg_int}}
            ]}}
        },
        "Configuration": {"Drivers": {"IAbEntrySearcher": "Maximizer.Model.Access.Sql.AbEntrySearcher"}},
        "Compatibility": {"AbEntryKey": "2.0"}
    }
    try:
        r = req.post(f"{MX_BASE}/AbEntryRead", headers=hdrs, json=body, timeout=15)
        d = r.json()
        items = d.get("AbEntry",{}).get("Data",[])
        if not isinstance(items, list): items = [items] if items else []
        if items:
            key  = items[0].get("Key","")
            name = items[0].get("CompanyName","")
            print(f"  Found by TYPEID(114): {name}")
            return {"key": key, "companyName": name}
    except Exception as e:
        print(f"  mx_find error: {e}")
    return None

def _clean_email_for_mx(e):
    """Return a single valid email or '' (Maximizer rejects malformed emails)."""
    import re
    e = str(e or "").strip()
    if not e: return ""
    # take first if multiple separated by , ; / space
    first = re.split(r"[,;/\s]+", e)[0].strip()
    if re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", first):
        return first[:100]
    return ""

def _clean_phone_for_mx(p):
    """Return a cleaned single phone number or '' for Maximizer Phone field."""
    import re
    p = str(p or "").strip()
    if not p: return ""
    # If multiple numbers mashed together, take the first chunk before a double-space/comma/semicolon
    first = re.split(r"[,;]|\s{2,}|\s(?=\d{6,})", p)[0].strip()
    # keep only valid phone chars
    cleaned = re.sub(r"[^0-9+()\- ]", "", first).strip()
    return cleaned[:30]

def _clean_website_for_mx(w):
    w = str(w or "").strip()
    return w[:200]

def build_charity_data_full(c):
    """Build complete Company entry data with ALL UDFs for single create/update call."""
    name = title_case(c.get("name","") or "Unknown Charity")[:79]
    data = {"Type": "Company", "CompanyName": name}
    _ph = _clean_phone_for_mx(c.get("phone"))
    if _ph: data["Phone1"] = _ph
    _em = _clean_email_for_mx(c.get("email"))
    if _em: data["Email1"] = _em
    _web = _clean_website_for_mx(c.get("website"))
    if _web: data["WebSite"] = _web

    # Address fields — all capped at 79 chars per Maximizer limit
    # From CCToMaximizerCRM.exe: AddressLine1 = lines 1+2+3 joined, City=line4, State=line5
    a1 = str(c.get("address1","") or "")
    a2 = str(c.get("address2","") or "")
    a3 = str(c.get("address3","") or "")
    combined = " ".join(p for p in [a1, a2, a3] if p).strip()
    # Split at word boundary before 79 chars
    if len(combined) <= 79:
        line1, line2 = combined, ""
    else:
        split_at = combined.rfind(" ", 0, 79)
        if split_at == -1: split_at = 79
        line1 = combined[:split_at].strip()
        line2 = combined[split_at:].strip()[:79]
    city    = str(c.get("county","") or "")[:79]   # address_line_four
    state   = str(c.get("city","") or "")[:79]     # address_line_five
    zipcode = str(c.get("postcode","") or "")[:10]
    if line1 or zipcode or city:
        addr_dict = {
            "AddressLine1": line1,
            "City":          city,
            "StateProvince": state,
            "ZipCode":       zipcode,
            "Country":       "England"   # All CC charities are in England/Wales
        }
        if line2:
            addr_dict["AddressLine2"] = line2
        data["Address"] = addr_dict

    fin = c.get("financials") or {}
    what = c.get("what","") or fin.get("what","")
    who  = c.get("who","")  or fin.get("who","")
    how  = c.get("how","")  or fin.get("how","")
    region = str(c.get("geo_area","") or c.get("region","") or "throughout england").strip()
    la = str(c.get("local_authority","") or c.get("town","") or "").strip()

    # Organisation Number = TYPEID(114) as integer (from CCToMaximizerCRM.exe)
    # Charity Suffix = TYPEID(263) = 0 for main charity
    reg = str(c.get("reg_number","")).strip()
    if reg:
        try:
            data["/AbEntry/Udf/$TYPEID(114)"] = int(reg)
            data["/AbEntry/Udf/$TYPEID(263)"] = 0
        except: pass

    # Multi-value table UDFs
    what_keys = _multi_keys(WHAT_MAP, what)
    if what_keys: data["/AbEntry/Udf/$TYPEID(111)"] = what_keys
    who_keys = _multi_keys(WHO_MAP, who)
    if who_keys: data["/AbEntry/Udf/$TYPEID(112)"] = who_keys
    how_keys = _multi_keys(HOW_MAP, how)
    if how_keys: data["/AbEntry/Udf/$TYPEID(113)"] = how_keys
    rk = _lookup(REGION_MAP, region)
    if rk: data["/AbEntry/Udf/$TYPEID(109)"] = rk
    lk = _lookup(LOCAL_AUTH_MAP, la)
    if lk: data["/AbEntry/Udf/$TYPEID(108)"] = lk
    # Policy is multi-select, same as What/Who/How — a charity commonly has several
    policy = c.get("policy","") or fin.get("policy","")
    policy_keys = _multi_keys(POLICY_MAP, policy)
    if policy_keys: data["/AbEntry/Udf/$TYPEID(261)"] = policy_keys
    # Country (area of operation outside England/Wales, where applicable)
    country = str(c.get("country","") or "").strip()
    if country:
        # A charity can operate in several countries; take the first that maps —
        # Country appears to be single-select in Maximizer's UDF schema.
        for one in country.split(","):
            ck = _lookup(COUNTRY_MAP, one.strip())
            if ck:
                data["/AbEntry/Udf/$TYPEID(107)"] = ck
                break
    # Gift Aid (Yes/No)
    ga = str(c.get("gift_aid","") or "").strip()
    if ga:
        gk = _lookup(GIFT_AID_MAP, ga)
        if gk: data["/AbEntry/Udf/$TYPEID(243)"] = gk
    # Linked Charity: "1"=No (main charity, suffix=0), "2"=Yes (linked, suffix>0)
    linked_value = "2" if int(c.get("charity_suffix", 0) or 0) > 0 else "1"
    data["/AbEntry/Udf/$TYPEID(264)"] = linked_value

    # Caller (TYPEID 378) - string field, who is calling from dashboard
    caller = str(c.get("caller","") or "").strip()
    if caller:
        data["/AbEntry/Udf/$TYPEID(378)"] = caller[:50]

    # Caller Status (TYPEID 379) - currently Yes/No, will be text after admin change
    # Map dashboard status to value. If status is set, send as string
    status = str(c.get("status","") or "").strip()
    if status:
        data["/AbEntry/Udf/$TYPEID(379)"] = status[:50]

    return data

def build_charity_base(c):
    """Alias for full data build."""
    return build_charity_data_full(c)

def _mx_find_key_by_creation_date(company_name, created_after_iso):
    """Find recently created Individual entry by lastName=reg_number."""
    import requests as req
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    name_lower = company_name.lower()
    match_parts = [name_lower[:15], name_lower[:10], name_lower[:8]]

    bodies = [
        # ONLY working pattern: empty searchQuery, sorted by recent
        {"abEntry": {"criteria": {"searchQuery": {}, "top": 100},
                     "scope": {"fields": {"key":1,"companyName":1,"type":1}},
                     "orderBy": {"fields": [{"lastModifyDate": "DESC"}]}}},
    ]

    for i, body in enumerate(bodies):
        try:
            r = req.post(f"{MX_BASE}/AbEntryRead", headers=hdrs, json=body, timeout=12)
            d = r.json()
            if d.get("Code") != 0:
                print(f"  Strategy {i+1}: Code={d.get('Code')} Msg={str(d.get('Msg',''))[:60]}")
                continue
            items = d.get("abEntry",{}).get("Data",[])
            print(f"  Strategy {i+1}: {len(items)} items returned")
            for item in items:
                cn = item.get("companyName","").lower()
                t  = item.get("type","")
                # Log first 5 items to see what's there
                if items.index(item) < 5:
                    print(f"    [{t}] '{cn[:35]}'")
                for mp in match_parts:
                    if mp and mp in cn:
                        print(f"  ✓ MATCH strategy {i+1}: '{item.get('companyName')}'")
                        return item.get("key","")
        except Exception as e:
            print(f"  Strategy {i+1} error: {e}")
    return ""

def build_charity_udfs(c, key):
    """Build list of individual UDF update dicts — each sent separately."""
    updates = []

    def upd(typeid, val):
        if val:
            updates.append({"Key": key,
                            f"/AbEntry/Udf/$TYPEID({typeid})": str(val)})

    # Pull what/who/how from financials if not directly on c
    fin = c.get("financials") or {}
    what = c.get("what","") or fin.get("what","")
    who  = c.get("who","")  or fin.get("who","")
    how  = c.get("how","")  or fin.get("how","")

    # Organisation Number = TYPEID(114), Suffix = TYPEID(263)
    reg = str(c.get("reg_number","")).strip()
    if reg:
        try:
            upd(114, int(reg))
            upd(263, 0)
        except: pass

    # Multi-value: pass full array to table UDFs
    what_keys = _multi_keys(WHAT_MAP, what)
    if what_keys: updates.append({"Key": key, "/AbEntry/Udf/$TYPEID(111)": what_keys})

    who_keys = _multi_keys(WHO_MAP, who)
    if who_keys: updates.append({"Key": key, "/AbEntry/Udf/$TYPEID(112)": who_keys})

    how_keys = _multi_keys(HOW_MAP, how)
    if how_keys: updates.append({"Key": key, "/AbEntry/Udf/$TYPEID(113)": how_keys})

    region = str(c.get("geo_area","") or c.get("region","") or "throughout england").strip()
    rk = _lookup(REGION_MAP, region)
    if rk: upd(109, rk)

    la = str(c.get("local_authority","") or c.get("town","") or "").strip()
    lk = _lookup(LOCAL_AUTH_MAP, la)
    if lk: upd(108, lk)

    # Policy is multi-select, same as What/Who/How — a charity commonly has several
    policy = c.get("policy","") or fin.get("policy","")
    policy_keys = _multi_keys(POLICY_MAP, policy)
    if policy_keys: updates.append({"Key": key, "/AbEntry/Udf/$TYPEID(261)": policy_keys})

    # Country (area of operation outside England/Wales, where applicable)
    country = str(c.get("country","") or "").strip()
    if country:
        for one in country.split(","):
            ck = _lookup(COUNTRY_MAP, one.strip())
            if ck:
                upd(107, ck)
                break

    # Gift Aid (Yes/No)
    ga = str(c.get("gift_aid","") or "").strip()
    if ga:
        gk = _lookup(GIFT_AID_MAP, ga)
        if gk: upd(243, gk)

    # Linked Charity: "1"=No, "2"=Yes based on suffix
    linked_v = "2" if int(c.get("charity_suffix", 0) or 0) > 0 else "1"
    upd(264, linked_v)

    # Caller and Caller Status
    caller = str(c.get("caller","") or "").strip()
    if caller: upd(378, caller[:50])
    status = str(c.get("status","") or "").strip()
    if status: upd(379, status[:50])

    return updates

def mx_apply_udfs(key, c):
    """Apply each UDF update separately using {Value: v} wrapper format."""
    results = []
    for upd_data in build_charity_udfs(c, key):
        try:
            # Wrap UDF values in {"Value": v} per UpdateCompanyUdfs format
            wrapped = {"Key": upd_data["Key"]}
            for k, v in upd_data.items():
                if k != "Key":
                    wrapped[k] = {"Value": v}
            r = mx_write_update(wrapped)
            field = [k for k in upd_data if k.startswith("/AbEntry")][0]
            ok = r.get("Code") == 0
            result = {"field":field,"val":upd_data[field],"Code":r.get("Code"),"ok":ok}
            if not ok:
                result["Msg"] = str(r.get("Msg",""))[:150]
            results.append(result)
        except Exception as e:
            results.append({"error":str(e)[:100]})
    return results

def _mx_get_all_keys():
    """Get Address Book entry keys — returns Companies too via address field."""
    try:
        r = mx_call("AbEntryRead", {
            "abEntry": {
                "criteria": {"searchQuery": {}, "top": 2000},
                "scope": {"fields": {"key":1, "companyName":1,
                                     "/AbEntry/Address/Key":1}},
                "orderBy": {"fields": [{"lastModifyDate": "DESC"}]}
            }
        })
        return {item["key"]: item.get("companyName","")
                for item in r.get("abEntry",{}).get("Data",[])}
    except Exception as e:
        print(f"  _mx_get_all_keys: {e}")
        return {}

def mx_create_note(entry_key, note_text, caller_name="", timestamp=None):
    """Create a Note on a Maximizer Address Book entry.
    timestamp: optional ISO string or 'YYYY-MM-DD HH:MM:SS' — uses current UTC if omitted.
    """
    if not entry_key or not note_text: return None
    from datetime import datetime as _dt
    if timestamp:
        now_iso = str(timestamp).replace(" ", "T")[:19]
    else:
        now_iso = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    html = f"<p>{note_text.replace(chr(10),'<br/>')}</p>"
    if caller_name:
        html = f"<p><strong>{caller_name}</strong> — {now_iso[:10]}</p>" + html
    body = {
        "Note": {
            "Data": {
                "Key": None,
                "ParentKey": entry_key,
                "DateTime": now_iso,
                "RichText": html,
                "Category": "Dashboard Note"
            }
        },
        "Compatibility": {"AbEntryKey": "2.0"}
    }
    try:
        resp = mx_call("Create", body)
        print(f"  mx_create_note: Code={resp.get('Code')} for {entry_key[:25]}")
        return resp
    except Exception as e:
        print(f"  mx_create_note error: {e}")
        return None

def mx_read_notes(entry_key, top=50):
    """Read all Notes for a given Address Book entry, newest first."""
    if not entry_key: return []
    import requests as req
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    body = {
        "Note": {
            "Criteria": {
                "SearchQuery": {"ParentKey": {"$EQ": entry_key}}
            },
            "Scope": {
                "Fields": {
                    "Key": 1, "ParentKey": 1, "DateTime": 1,
                    "RichText": 1, "Category": 1, "Creator": 1
                }
            },
            "OrderBy": {"Fields": [{"DateTime": "DESC"}], "PageSize": top}
        },
        "Compatibility": {"AbEntryKey": "2.0"}
    }
    try:
        r = req.post(f"{MX_BASE}/Read", headers=hdrs, json=body, timeout=12)
        d = r.json()
        if d.get("Code") != 0:
            print(f"  mx_read_notes: Code={d.get('Code')} Msg={str(d.get('Msg',''))[:80]}")
            return []
        items = d.get("Note",{}).get("Data",[])
        if not isinstance(items, list): items = [items] if items else []
        return items
    except Exception as e:
        print(f"  mx_read_notes error: {e}")
        return []

@app.route("/api/maximizer/test_note")
def mx_test_note():
    """Test note creation directly. Pass ?reg=1202982&text=hello&caller=Mahesh."""
    reg    = str(request.args.get("reg","1202982")).strip()
    text   = str(request.args.get("text","Test note from API")).strip()
    caller = str(request.args.get("caller","TestUser")).strip()
    result = {"reg": reg, "text": text, "caller": caller}

    found = mx_find_by_org_number(reg)
    if not found or not found.get("key"):
        result["error"] = "Charity not found in Maximizer"
        return jsonify(result)

    entry_key = found["key"]
    result["entry_key"] = entry_key[:30]
    result["company_name"] = found.get("companyName","")

    # Try each endpoint and report full response
    from datetime import datetime as _dt
    now_iso = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    html = f"<p><strong>{caller}</strong> — {now_iso[:10]}</p><p>{text}</p>"
    body = {
        "Note": {
            "Data": {
                "Key": None,
                "ParentKey": entry_key,
                "DateTime": now_iso,
                "RichText": html,
                "Category": "Dashboard Note"
            }
        },
        "Compatibility": {"AbEntryKey": "2.0"}
    }

    result["attempts"] = {}
    import requests as req
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    for endpoint in ("Create", "NoteCreate", "Note/Create", "NotesCreate"):
        try:
            r = req.post(f"{MX_BASE}/{endpoint}", headers=hdrs, json=body, timeout=12)
            try: resp_data = r.json()
            except: resp_data = r.text[:200]
            result["attempts"][endpoint] = {
                "status": r.status_code,
                "response": resp_data
            }
        except Exception as e:
            result["attempts"][endpoint] = {"error": str(e)[:200]}
    return jsonify(result)

def mx_create_entry(c, caller=""):
    """Create Company entry - uses TYPEID(114) for org number, gets key from response."""
    import time

    # Enrich with full CC data
    c = enrich_charity_from_cc(c)

    # Check for existing entry before creating (prevent duplicates)
    reg = str(c.get("reg_number","")).strip()
    if reg:
        existing = mx_find_by_org_number(reg)
        if existing:
            print(f"  Entry already exists — updating instead of creating")
            r = mx_update_entry(existing.get("key",""), c, caller)
            return existing.get("key",""), r
        print(f"  No existing entry found for reg {reg} — creating new")

    # Create with ALL UDFs in one call (Table types save correctly)
    # TYPEID(114) for org number is also in create call
    data = build_charity_data_full(c)
    print(f"  mx_create: {data.get('CompanyName','?')}")
    result = mx_write_create(data)
    print(f"  Code={result.get('Code')} Msg={str(result.get('Msg',''))[:120]}")
    if result.get("Code") != 0:
        return "", result   # failed — caller will surface the error

    # Get key from response (per CCToMaximizerCRM.exe: resp.AbEntry.Data.Key)
    key = result.get("_extracted_key","") or _extract_key_from_create_resp(result)
    if not key:
        # Fallback: search by TYPEID(114) = reg number
        time.sleep(1)
        found = mx_find_by_org_number(str(c.get("reg_number","")))
        if found: key = found.get("key","")

    if key:
        print(f"  Got key: {key[:25]}")
        # Apply UDFs that couldn't go in create (one at a time)
        mx_apply_udfs(key, c)
    else:
        print(f"  No key found — entry may NOT have been created")
    return key, result

def mx_update_entry(key, c, caller=""):
    """Update existing Company entry with all fields."""
    c = enrich_charity_from_cc(c)
    data = build_charity_data_full(c)
    data["Key"] = key
    data.pop("Type", None)
    try:
        r = mx_write_update(data)
        print(f"  mx_update: Code={r.get('Code')} {str(r.get('Msg',''))[:60]}")
        return r
    except Exception as e:
        print(f"  mx_update error: {e}")
        return {"Code": -1}

def mx_search_by_name(name):
    try:
        result = mx_read(search_query={"companyName":f"%{name[:25]}%"},
                         fields={"key":1,"companyName":1,"type":1}, top=5)
        return result.get("abEntry",{}).get("Data",[])
    except: return []

def _mx_find_company_key(name):
    """Find the key of a Company entry by name — tries multiple strategies."""
    import base64 as b64
    name_part = name[:30].lower()
    # Strategy 1: search with Company type filter
    try:
        r = mx_read(search_query={"companyName": name[:30]},
                    fields={"key":1,"companyName":1,"type":1},
                    top=20, entry_type="Company")
        for item in r.get("abEntry",{}).get("Data",[]):
            if name_part[:12] in item.get("companyName","").lower():
                return item["key"]
    except Exception as e:
        print(f"  _find_co strategy1: {e}")
    # Strategy 2: get all recent and filter
    try:
        r2 = mx_read(search_query={},
                     fields={"key":1,"companyName":1,"type":1}, top=100)
        for item in r2.get("abEntry",{}).get("Data",[]):
            if name_part[:12] in item.get("companyName","").lower():
                return item["key"]
    except Exception as e:
        print(f"  _find_co strategy2: {e}")
    # Strategy 3: Read with addresses field (deprecated but returns Companies)
    try:
        r3 = mx_call("AbEntryRead", {
            "abEntry": {
                "criteria": {"searchQuery": {"companyName": name[:30]}, "top": 10},
                "scope": {"fields": {"key":1,"companyName":1,"type":1,
                                     "/AbEntry/Address/Key":1}}
            }
        })
        for item in r3.get("abEntry",{}).get("Data",[]):
            if name_part[:12] in item.get("companyName","").lower():
                return item["key"]
    except Exception as e:
        print(f"  _find_co strategy3: {e}")
    return ""


def _push_unsynced_calls_to_mx(reg_str, mx_key):
    """Push any unsynced call-log entries to Maximizer as Notes.
    Each call becomes a Note capturing outcome, duration, phone, caller and time.
    """
    try:
        rows = db_query(
            "SELECT id, phone_used, outcome, notes, duration_sec, caller, timestamp "
            "FROM call_log WHERE reg_number=? AND synced_to_maximizer=0 ORDER BY id ASC",
            (reg_str,)
        )
        if not rows:
            return 0
        pushed = 0
        for row_id, phone, outcome, notes, dur, cname, ts in rows:
            try:
                note_text = f"📞 Call: {outcome or 'Logged'}"
                try: dur = int(dur or 0)
                except: dur = 0
                if dur: note_text += f" ({dur//60}m {dur%60}s)"
                if phone: note_text += f"\nPhone: {phone}"
                if cname: note_text += f"\nCalled by: {cname}"
                # (free-text notes are captured separately as a comment Note)
                resp = mx_create_note(mx_key, note_text, cname or "Dashboard", timestamp=ts)
                if resp and resp.get("Code") == 0:
                    db_exec("UPDATE call_log SET synced_to_maximizer=1 WHERE id=?", (row_id,))
                    pushed += 1
            except Exception as e:
                print(f"  Call-note push error id={row_id}: {e}")
        if pushed:
            print(f"  Pushed {pushed} unsynced calls to Maximizer for {reg_str}")
        return pushed
    except Exception as e:
        print(f"  _push_unsynced_calls error: {e}")
        return 0

def _push_unsynced_comments_to_mx(reg_str, mx_key):
    """Push any unsynced dashboard comments to Maximizer as individual Notes.
    Each comment becomes its own Note with the original timestamp.
    """
    try:
        rows = db_query(
            "SELECT id, text, caller, timestamp FROM comment_history "
            "WHERE reg_number=? AND synced_to_maximizer=0 ORDER BY id ASC",
            (reg_str,)
        )
        if not rows:
            return 0
        pushed = 0
        for row_id, text, cname, ts in rows:
            try:
                resp = mx_create_note(mx_key, text, cname or "Dashboard", timestamp=ts)
                if resp and resp.get("Code") == 0:
                    db_exec("UPDATE comment_history SET synced_to_maximizer=1 WHERE id=?", (row_id,))
                    pushed += 1
            except Exception as e:
                print(f"  Note push error id={row_id}: {e}")
        if pushed:
            print(f"  Pushed {pushed} unsynced comments to Maximizer for {reg_str}")
        return pushed
    except Exception as e:
        print(f"  _push_unsynced_comments error: {e}")
        return 0

def do_sync_one(reg, c, caller, page):
    """Sync one charity — fetch full CC data then create/update in Maximizer."""
    # Enrich with full CC API data (what/who/how, address, etc.)
    c = enrich_charity_from_cc(c)

    # ── Merge dashboard contact overrides (user-edited phone/email/website) ──
    reg_str = str(c.get("reg_number","")).strip()
    if reg_str:
        co = _contact_overrides.get(reg_str, {})
        if co.get("phone"):   c["phone"]   = co["phone"]
        if co.get("email"):   c["email"]   = co["email"]
        if co.get("website"): c["website"] = co["website"]

    # Fallback classification from bulk file
    fin = c.get("financials") or {}
    if not c.get("what") and fin.get("what"): c["what"] = fin["what"]
    if not c.get("who")  and fin.get("who"):  c["who"]  = fin["who"]
    if not c.get("how")  and fin.get("how"):  c["how"]  = fin["how"]

    # ── Caller & Caller Status UDFs are driven by the dashboard "Called" state ──
    # The dashboard sends the ACTUAL on-screen called state in each charity
    # (_called / _caller). We use that — NOT the server-side _called_log, which
    # can hold stale entries that no longer match what the user sees.
    # Not called → both left blank in Maximizer.
    c["caller"] = ""
    c["status"] = ""
    if c.get("_called"):
        c["caller"] = c.get("_caller") or caller or ""
        # Caller Status = the Status dropdown value the user selected (e.g. "Interested").
        # Prefer the value sent from the dashboard; fall back to the stored status in
        # Turso (survives page reloads); default to "Contacted" if none set.
        _dash_page = {"old_charities": "old"}.get(page, page)
        server_status = _statuses.get(f"{_dash_page}|{reg_str}", "") if reg_str else ""
        c["status"] = (c.get("_status") or server_status or "Contacted")

    charity_name = title_case(c.get("name","") or "")
    existing = mx_find_by_org_number(reg_str) if reg_str else None
    existing_key = existing.get("key","") if existing else ""
    print(f"  {reg_str}: existing={'yes' if existing_key else 'no'} "
          f"phone={'yes' if c.get('phone') else 'no'} "
          f"caller={c.get('caller','')} status={c.get('status','')}")

    mx_key = ""
    if existing_key:
        print(f"  Updating: {charity_name[:30]}")
        result = mx_update_entry(existing_key, c, caller)
        code = result.get("Code") if isinstance(result, dict) else 0
        if code not in (0, None):
            raise Exception(f"Maximizer UPDATE failed (Code={code}): {str(result.get('Msg',''))[:400]}")
        mx_key = existing_key
        action = "updated"
    else:
        print(f"  Creating: {charity_name[:30]}")
        key, result = mx_create_entry(c, caller)
        code = result.get("Code") if isinstance(result, dict) else 0
        if code not in (0, None):
            raise Exception(f"Maximizer CREATE failed (Code={code}): {str(result.get('Msg',''))[:400]}")
        if not key:
            raise Exception(f"CREATE reported Code={code} but NO entry was saved (no Key returned). Resp: {str(result)[:140]}")
        mx_key = key
        action = "created"

    # ── Push any unsynced dashboard comments → Maximizer Notes ──────────────
    if mx_key and reg_str:
        _push_unsynced_comments_to_mx(reg_str, mx_key)
        _push_unsynced_calls_to_mx(reg_str, mx_key)

    return action, None

# ── API Routes ────────────────────────────────────────────────────────────────

@app.route("/api/maximizer/test")
def mx_test():
    server_ip = "unknown"
    try:
        r = requests.get("https://checkip.amazonaws.com",timeout=4)
        server_ip = r.text.strip()
    except: pass
    try:
        result = mx_read(fields={"key":1,"companyName":1},top=1)
        entries = result.get("abEntry",{}).get("Data",[])
        sample = entries[0].get("companyName","") if entries else ""
        return jsonify({"ok":True,"message":f"Connected! Sample: '{sample}'",
                        "database":MX_DB,"base":MX_BASE})
    except Exception as e:
        msg = str(e)
        hint = f"Render IP ({server_ip}) may not be whitelisted." if ("403" in msg or "whitelist" in msg.lower()) else ""
        return jsonify({"ok":False,"error":msg,"hint":hint,"server_ip":server_ip}), 200

@app.route("/api/maximizer/sync_one", methods=["POST"])
def mx_sync_one_row():
    """Sync a single charity to Maximizer (per-row sync button)."""
    data = request.json or {}
    c = dict(data.get("charity") or {})
    reg = str(c.get("reg_number", "")).strip()
    page = str(data.get("page", "")).strip()
    caller = str(data.get("caller", "")).strip()
    if not reg:
        return jsonify({"ok": False, "error": "Missing reg_number"}), 400
    c["reg_number"] = reg
    try:
        action, _ = do_sync_one(reg, c, caller, page)
        return jsonify({"ok": True, "action": action})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:200]}), 500

@app.route("/api/maximizer/sync",methods=["POST"])
def mx_sync():
    global MX_BASE
    data      = request.json or {}
    charities = data.get("charities",[])
    page      = data.get("page","prospects")
    caller    = data.get("caller","")
    base_url  = data.get("base_url","").strip()
    if base_url: MX_BASE = base_url.rstrip("/")
    sync_key  = f"mx_sync_{page}"
    with _bg_lock:
        if _bg_status.get(sync_key)=="loading":
            return jsonify({"ok":False,"message":"Sync already running"}),409
    def _worker():
        with _bg_lock: _bg_status[sync_key]="loading"
        res={"created":0,"updated":0,"errors":0,"mx_callers_pulled":0,
             "error":"","error_details":[]}
        try:
            for c in charities[:100]:
                reg=c.get("reg_number","")
                if not reg: continue
                try:
                    action,_=do_sync_one(reg,c,caller,page)
                    if action=="created": res["created"]+=1
                    else: res["updated"]+=1
                except Exception as e:
                    import traceback
                    err_msg = f"{reg}: {str(e)[:450]}"
                    print(f"  sync_err: {err_msg}\n{traceback.format_exc()}")
                    res["errors"]+=1
                    res["error_details"].append(err_msg)
        except Exception as e:
            res["error"]=str(e)
        finally:
            cache_set(sync_key,res)
            with _bg_lock: _bg_status[sync_key]="idle"
        print(f"Mx sync done: {res}")
    threading.Thread(target=_worker,daemon=True).start()
    return jsonify({"ok":True,"status":"loading",
                    "message":f"Syncing {len(charities)} charities…"}),202

@app.route("/api/maximizer/sync_status")
def mx_sync_status():
    page=request.args.get("page","prospects")
    sync_key=f"mx_sync_{page}"
    with _bg_lock: status=_bg_status.get(sync_key,"idle")
    result=cache_get(sync_key,max_age=60) or {}
    # Include error details in response for debugging
    return jsonify({"status":status,"result":result,
                    "error_details":result.get("error_details",[])})

@app.route("/api/maximizer/config",methods=["GET","POST"])
def mx_config():
    global MX_BASE
    if request.method=="POST":
        d=request.json or {}
        if d.get("base_url"): MX_BASE=d["base_url"].rstrip("/")
    return jsonify({"ok":True,"base_url":MX_BASE,"database":MX_DB,
                    "user":"MAHESH","token_expires":"2029-05-10"})

@app.route("/api/maximizer/find")
def mx_find():
    reg=request.args.get("reg","")
    name=request.args.get("name","")
    try:
        if reg:
            entry=mx_find_by_org_number(reg)
            if entry: return jsonify({"found":True,"entry":entry})
            if name:
                entries=mx_search_by_name(name[:30])
                return jsonify({"found":bool(entries),"entries":entries,"note":"by name"})
            return jsonify({"found":False,"note":f"No entry for reg {reg}"})
        return jsonify({"error":"Provide ?reg= parameter"}),400
    except Exception as e:
        return jsonify({"error":str(e)}),500

@app.route("/api/maximizer/read_org_number_typeid")
def mx_read_org_number_typeid():
    """Read Juvenis entry UDFs in small batches to find Organisation Number TYPEID."""
    import requests as req
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    results = {}
    juvenis_key = "Q29udGFjdAkyNDAzMjcyNTIyMzc1Nzk5MzU4MDJDCTE="

    # Read in small batches of 10 TYPEIDs to avoid errors
    all_udf_vals = {}
    for batch_start in range(1, 50, 10):
        batch = list(range(batch_start, min(batch_start+10, 50)))
        fields = {"key":1, "companyName":1}
        for i in batch:
            fields[f"/AbEntry/Udf/$TYPEID({i})"] = 1
        try:
            r = req.post(f"{MX_BASE}/AbEntryRead", headers=hdrs, json={
                "abEntry": {
                    "criteria": {"searchQuery": {"key": juvenis_key}, "top": 1},
                    "scope": {"fields": fields}
                }
            }, timeout=10)
            d = r.json()
            if d.get("Code") == 0:
                for item in d.get("abEntry",{}).get("Data",[]):
                    for k,v in item.items():
                        if k.startswith("/AbEntry/Udf") and v and v != [] and v != 0:
                            all_udf_vals[k] = v
        except Exception as e:
            results[f"batch_{batch_start}_err"] = str(e)

    results["non_null_udfs"] = all_udf_vals
    results["found_99999"] = {k:v for k,v in all_udf_vals.items()
                               if "99999" in str(v) or 99999 in (v if isinstance(v,list) else [v])}
    return jsonify(results)

@app.route("/api/maximizer/fix_org_number")
def mx_fix_org_number():
    """Find entry via TYPEID(264) search, then probe+set Organisation Number."""
    import requests as req
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    reg = request.args.get("reg","1202982")
    results = {}

    # Step 1: Find entry - try multiple search strategies
    key = ""
    c = {"reg_number": reg}
    c = enrich_charity_from_cc(c)
    name_hint = title_case(c.get("name","")).lower()[:12]

    search_bodies = [
        # Strategy 1: type=Company filter (worked in probe=1 test!)
        {"abEntry": {"criteria": {"searchQuery": {"type": "Company"}, "top": 50},
                     "scope": {"fields": {"key":1,"companyName":1,"type":1}}}},
        # Strategy 2: TYPEID(264)=2
        {"abEntry": {"criteria": {"searchQuery": {"/AbEntry/Udf/$TYPEID(264)": "2"}, "top": 20},
                     "scope": {"fields": {"key":1,"companyName":1,"type":1}}}},
        # Strategy 3: empty search
        {"abEntry": {"criteria": {"searchQuery": {}, "top": 200},
                     "scope": {"fields": {"key":1,"companyName":1,"type":1}}}},
    ]
    for i, body in enumerate(search_bodies):
        try:
            r = req.post(f"{MX_BASE}/AbEntryRead", headers=hdrs, json=body, timeout=12)
            d = r.json()
            items = d.get("abEntry",{}).get("Data",[])
            results[f"search_{i+1}"] = {"Code": d.get("Code"), "count": len(items),
                                         "first3": [x.get("companyName","") for x in items[:3]]}
            for item in items:
                cn = item.get("companyName","").lower()
                if name_hint[:8] in cn or reg in cn:
                    key = item.get("key","")
                    results["found_strategy"] = i+1
                    results["found_name"] = item.get("companyName")
                    break
            if key: break
        except Exception as e:
            results[f"search_{i+1}_err"] = str(e)

    if not key:
        results["error"] = "Entry not found via any search strategy"
        return jsonify(results)

    results["key"] = key[:30]

    # Step 2: Probe TYPEIDs to find Organisation Number field
    # Try setting the value and check which one accepts it
    found_typeid = None
    probe_results = {}
    # Skip known TYPEIDs and try numeric/alphanumeric ones
    skip = {107,108,109,111,112,113,243,261,264,2}
    for typeid in list(range(1,50)) + list(range(200,270)):
        if typeid in skip: continue
        try:
            r2 = req.post(f"{MX_BASE}/AbEntryUpdate", headers=hdrs,
                json={"AbEntry":{"Data":{"Key":key,
                      f"/AbEntry/Udf/$TYPEID({typeid})": reg}}},
                timeout=5)
            d2 = r2.json()
            if d2.get("Code") == 0:
                probe_results[f"TYPEID({typeid})"] = "✓ WORKS"
                if not found_typeid:
                    found_typeid = typeid
            elif "doesn't support" not in str(d2.get("Msg","")):
                probe_results[f"TYPEID({typeid})"] = f"diff: {str(d2.get('Msg',''))[:60]}"
        except: pass

    results["probe_results"] = probe_results
    results["found_org_typeid"] = found_typeid

    # Step 3: Also try TYPEID(2) as string (it was StringField)
    try:
        r3 = req.post(f"{MX_BASE}/AbEntryUpdate", headers=hdrs,
            json={"AbEntry":{"Data":{"Key":key,
                  "/AbEntry/Udf/$TYPEID(2)": str(reg)}}},
            timeout=5)
        d3 = r3.json()
        results["typeid2_result"] = {"Code": d3.get("Code"),
                                      "Msg": str(d3.get("Msg",""))[:100]}
    except Exception as e:
        results["typeid2_err"] = str(e)

    return jsonify(results)

@app.route("/api/maximizer/probe_org_number_typeid")
def mx_probe_org_number():
    """Find the TYPEID for Organisation Number UDF by probing."""
    import requests as req
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    # Find the existing test entry
    existing = mx_find_by_org_number_by_name("1202982")
    if not existing:
        return jsonify({"error": "No test entry found — run /create_test_charity first"})
    key = existing.get("key","")
    results = {"key": key, "found_at": existing.get("companyName")}
    # Probe TYPEIDs in range likely for numeric/alphanumeric UDFs
    # We know: 107,108,109,111,112,113,243,261,264 are Table/Yes-No
    # Organisation Number (Numeric) and Organisation Number New (Alphanumeric)
    # are likely in a different range. Probe 1-50 and 200-270
    for typeid in list(range(1,50)) + list(range(200,270)):
        if typeid in [107,108,109,111,112,113,243,261,264]:
            continue
        try:
            r = req.post(f"{MX_BASE}/AbEntryUpdate", headers=hdrs,
                json={"AbEntry": {"Data": {"Key": key,
                      f"/AbEntry/Udf/$TYPEID({typeid})": "1202982"}}},
                timeout=5)
            d = r.json()
            if d.get("Code") == 0:
                results[f"TYPEID({typeid})"] = "✓ WORKS — likely Organisation Number!"
            elif "doesn't support" not in str(d.get("Msg","")):
                results[f"TYPEID({typeid})"] = f"Different error: {str(d.get('Msg',''))[:60]}"
        except Exception as e:
            pass  # Skip timeouts
    return jsonify(results)

def mx_find_by_org_number_by_name(reg_no):
    """Find entry by name containing reg number."""
    try:
        result = mx_read(
            search_query={"companyName": f"%{reg_no}%"},
            fields={"key":1,"companyName":1,"type":1}, top=5
        )
        items = result.get("abEntry",{}).get("Data",[])
        for item in items:
            if str(reg_no) in item.get("companyName",""):
                return item
        return None
    except: return None

@app.route("/api/maximizer/create_test_charity")
def mx_create_test_charity():
    """Create or update test entry — checks for duplicates by TYPEID(114)."""
    c = {"reg_number": "1202982",
         "name": "ST GEORGE'S INDIAN ORTHODOX CHURCH, LONDON",
         "caller": "Muhanna",
         "status": "Contacted"}
    c = enrich_charity_from_cc(c)
    result = {"cc_data": {k:v for k,v in c.items() if k not in ("financials",)}}

    try:
        # STEP 1: Check for existing entry by TYPEID(114) BEFORE create
        existing = mx_find_by_org_number("1202982")
        if existing and existing.get("key"):
            existing_key = existing["key"]
            result["action"] = "UPDATE (existing entry found)"
            result["existing_key"] = existing_key[:30]
            result["existing_name"] = existing.get("companyName","")
            # Build data without Type (can't change type on update)
            data = build_charity_data_full(c)
            data.pop("Type", None)
            data["Key"] = existing_key
            result["data_sent_address"] = data.get("Address",{})
            result["data_sent_company"] = data.get("CompanyName","")
            result["udfs_in_create"] = [k for k in data if k.startswith("/AbEntry")]
            resp = mx_call("AbEntryUpdate", {"AbEntry":{"Data":data}})
            result["update_code"] = resp.get("Code")
            result["update_msg"]  = str(resp.get("Msg",""))[:100]
            result["full_create_response"] = resp
            result["SUCCESS"] = resp.get("Code") == 0
            result["note"] = "UPDATED existing entry — no duplicate created"
            result["fields_set"] = list(resp.get("AbEntry",{}).get("Data",{}).keys())
        else:
            # STEP 2: Create new entry (only when no duplicate exists)
            result["action"] = "CREATE (no existing entry)"
            data = build_charity_data_full(c)
            result["type_check"] = data.get("Type","MISSING")
            result["udfs_in_create"] = [k for k in data if k.startswith("/AbEntry")]
            result["data_sent_address"] = data.get("Address",{})
            result["data_sent_company"] = data.get("CompanyName","")
            resp = mx_write_create(data)
            result["create_code"] = resp.get("Code")
            result["create_msg"]  = str(resp.get("Msg",""))[:100]
            result["full_create_response"] = resp
            result["SUCCESS"] = resp.get("Code") == 0
            if resp.get("Code") == 0:
                result["note"] = "CREATED new entry"
            else:
                result["note"] = f"CREATE FAILED: {str(resp.get('Msg',''))[:100]}"
            result["fields_set"] = list(resp.get("AbEntry",{}).get("Data",{}).keys())

        # STEP 3: Verify find works after operation
        import time; time.sleep(1)
        found = mx_find_by_org_number("1202982")
        result["find_after_create"] = found.get("companyName","NOT FOUND") if found else "NOT FOUND"

    except Exception as e:
        result["error"] = str(e)
    return jsonify(result)

@app.route("/api/maximizer/probe_caller")
def mx_probe_caller():
    """Probe specifically near TYPEID(378) and test different value types."""
    import requests as req, time
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    results = {"attempts": {}}

    found = mx_find_by_org_number("1202982")
    if not found or not found.get("key"):
        return jsonify({"error": "No test entry."})
    key = found["key"]
    results["test_key"] = key[:30]

    # Try TYPEIDs near 378 with different value types
    # 378 is Caller (string). Caller Status was created with it - likely 377 or 379
    test_typeids = [370, 371, 372, 373, 374, 375, 376, 377, 379, 380, 381, 382, 383, 384, 385]

    for typeid in test_typeids:
        attempts = {}
        # Try string
        for val_type, val in [("str", "STATUSCHECK"), ("int", 1), ("table", ["1"]),
                               ("yn", "1"), ("yn2", "2")]:
            try:
                r = req.post(f"{MX_BASE}/AbEntryUpdate", headers=hdrs,
                    json={"AbEntry":{"Data":{"Key":key,
                          f"/AbEntry/Udf/$TYPEID({typeid})": val}},
                          "Compatibility":{"AbEntryKey":"2.0"}},
                    timeout=3)
                d = r.json()
                if d.get("Code") == 0:
                    attempts[val_type] = "WORKS"
                else:
                    msg = str(d.get("Msg",""))[:80]
                    if "Too Many" in msg: time.sleep(2)
                    elif "doesn't support" not in msg:
                        attempts[val_type] = f"err: {msg[:50]}"
            except: pass
            time.sleep(0.1)
        if attempts:
            results["attempts"][f"TYPEID({typeid})"] = attempts

    results["note"] = "Check Maximizer Caller Status field. Tell me what value it shows now."
    return jsonify(results)


@app.route("/api/maximizer/update_by_id")
def mx_update_by_id():
    """Apply UDFs to Company entry. Usage: ?id=IDENTIFICATION&reg=REG_NUMBER"""
    import base64 as b64, requests as req
    identification = request.args.get("id","").strip()
    reg = request.args.get("reg","1202982").strip()
    probe_org = request.args.get("probe","0") == "1"
    if not identification or identification in ("PASTE_ID_HERE","YOUR_IDENTIFICATION_HERE"):
        return jsonify({"error":"Paste the actual Identification number from Maximizer System Information"}), 400

    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    c = {"reg_number": reg}
    c = enrich_charity_from_cc(c)
    results = {"identification": identification, "reg": reg, "cc_fetched": bool(c.get("what"))}

    # Find real key via TYPEID(2) search (most reliable)
    found = mx_find_by_org_number(reg)
    if found:
        key = found.get("key","")
        results["key_source"] = "TYPEID(2) search"
    else:
        # Fallback: try all suffix variants with the identification
        key = ""
        for suffix in ["0","1","2","3",""]:
            try_key = b64.b64encode(f"Company	{identification}	{suffix}".encode()).decode()
            try:
                r = __import__('requests').post(
                    f"{MX_BASE}/AbEntryRead", headers=hdrs,
                    json={"abEntry":{"criteria":{"searchQuery":{},"top":1},
                                     "scope":{"fields":{"key":1,"companyName":1}}}},
                    timeout=5
                )
                # Just use suffix 0 as fallback
                key = b64.b64encode(f"Company	{identification}	0".encode()).decode()
                results["key_source"] = f"constructed suffix=0"
            except: pass
            break

    if not key:
        results["error"] = "Could not find entry"
        return jsonify(results)

    results["key"] = key[:30]+"..."

    # Apply each UDF separately
    udf_results = mx_apply_udfs(key, c)
    results["udf_results"] = udf_results

    # Probe for Organisation Number TYPEID if requested
    if probe_org:
        found_typeid = None
        for typeid in list(range(1,50)) + list(range(200,280)):
            if typeid in [107,108,109,111,112,113,243,261,264]: continue
            try:
                r = req.post(f"{MX_BASE}/AbEntryUpdate", headers=hdrs,
                    json={"AbEntry":{"Data":{"Key":key,
                          f"/AbEntry/Udf/$TYPEID({typeid})": reg}}},
                    timeout=5)
                d = r.json()
                if d.get("Code") == 0:
                    results[f"TYPEID_{typeid}"] = f"✓ WORKS for org number!"
                    found_typeid = typeid
                    break
                elif "doesn't support" not in str(d.get("Msg","")):
                    results[f"TYPEID_{typeid}"] = f"different error: {str(d.get('Msg',''))[:40]}"
            except: pass
        results["org_num_typeid"] = found_typeid
    else:
        results["hint"] = "Add &probe=1 to find the Organisation Number TYPEID"

    return jsonify(results)

@app.route("/api/maximizer/fill_latest_company")
def mx_fill_latest_company():
    """Find entry by companyName in top 100 recent, apply all UDFs."""
    import requests as req
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    reg = request.args.get("reg","1202982")
    results = {}

    try:
        r = req.post(f"{MX_BASE}/AbEntryRead", headers=hdrs, json={
            "abEntry": {
                "criteria": {"searchQuery": {}, "top": 100},
                "scope": {"fields": {"key":1,"companyName":1,"type":1}},
                "orderBy": {"fields": [{"lastModifyDate": "DESC"}]}
            }
        }, timeout=12)
        d = r.json()
        items = d.get("abEntry",{}).get("Data",[])
        results["total"] = len(items)
        results["first_5"] = [{"n":x.get("companyName",""),"t":x.get("type","")} for x in items[:5]]

        # Get CC data for matching
        c = {"reg_number": reg}
        c = enrich_charity_from_cc(c)
        name_lc = title_case(c.get("name","")).lower()[:15]

        for item in items:
            cn = item.get("companyName","").lower()
            if name_lc[:10] in cn or reg in cn:
                key = item.get("key","")
                results["found"] = item.get("companyName")
                results["key"] = key[:30]
                results["udf_results"] = mx_apply_udfs(key, c)
                return jsonify(results)

        results["note"] = f"Not found. Looking for '{name_lc[:15]}' in {len(items)} entries"
    except Exception as e:
        results["error"] = str(e)
    return jsonify(results)

@app.route("/api/maximizer/apply_udfs_to_latest")
def mx_apply_to_latest():
    """Try all key variants for the given Identification, apply UDFs to working one."""
    import requests as req, base64 as b64
    hdrs = {"Authorization": f"Bearer {MX_TOKEN}", "Content-Type": "application/json"}
    identification = request.args.get("id","").strip()
    reg = request.args.get("reg","1202982")
    results = {"identification": identification, "reg": reg}

    if not identification:
        return jsonify({"error": "Provide ?id=IDENTIFICATION&reg=REG_NUMBER"}), 400

    c = {"reg_number": reg}
    c = enrich_charity_from_cc(c)

    # Try every possible key variant, log ALL responses
    attempts = {}
    working_key = ""
    for type_str in ["Company", "Contact", "Individual"]:
        for suffix in ["0", "1", "2", "3", ""]:
            raw = f"{type_str}\t{identification}\t{suffix}"
            key = b64.b64encode(raw.encode()).decode()
            try:
                r = req.post(f"{MX_BASE}/AbEntryUpdate", headers=hdrs,
                    json={"AbEntry": {"Data": {"Key": key,
                          "/AbEntry/Udf/$TYPEID(264)": "2"}}},
                    timeout=6)
                d = r.json()
                label = f"{type_str}_{suffix or 'empty'}"
                attempts[label] = {
                    "Code": d.get("Code"),
                    "Msg": str(d.get("Msg",""))[:120]
                }
                if d.get("Code") == 0:
                    working_key = key
                    results["working_format"] = label
                    break
            except Exception as e:
                attempts[f"{type_str}_{suffix or 'empty'}"] = {"error": str(e)[:80]}
        if working_key:
            break

    results["attempts"] = attempts
    if not working_key:
        results["error"] = "No working key — see attempts for exact error messages"
        return jsonify(results)

    results["udf_results"] = mx_apply_udfs(working_key, c)
    return jsonify(results)

# ════════════════════════════════════════════════════════════════════════════════
# RINGCENTRAL WEBPHONE INTEGRATION (Phase 2 — WebRTC in-browser calls)
# ════════════════════════════════════════════════════════════════════════════════
# Flow: backend exchanges long-lived JWT → short-lived access_token → calls SIP
# provisioning endpoint to get sipInfo (one-day SIP credentials). Frontend uses
# sipInfo + the @ringcentral/web-phone SDK to register over WSS and place calls.
# Audio flows via WebRTC peer connection — no RC desktop app needed.

RC_SERVER_URL    = os.environ.get("RC_SERVER_URL", "https://platform.ringcentral.com").rstrip("/")
RC_EMBEDDABLE_CLIENT_ID = os.environ.get("RC_EMBEDDABLE_CLIENT_ID", "") or RC_CLIENT_ID
RC_CLIENT_SECRET = os.environ.get("RC_CLIENT_SECRET", "")
RC_JWT           = os.environ.get("RC_JWT", "")
RC_FROM_EXT      = os.environ.get("RC_FROM_EXT", "")
RC_FROM_NUMBER   = os.environ.get("RC_FROM_NUMBER", "")
RC_CALLER_ID     = os.environ.get("RC_CALLER_ID", "")

_rc_token_cache  = {"access_token": "", "expires_at": 0.0}
_rc_token_lock   = threading.Lock()

def _rc_configured():
    return bool(RC_CLIENT_ID and RC_CLIENT_SECRET and RC_JWT)

def _rc_get_access_token():
    """Exchange the JWT credential for an OAuth access_token. Cached for ~50 min."""
    import time, base64 as _b64
    with _rc_token_lock:
        now = time.time()
        if _rc_token_cache["access_token"] and _rc_token_cache["expires_at"] > now + 60:
            return _rc_token_cache["access_token"]
        basic = _b64.b64encode(f"{RC_CLIENT_ID}:{RC_CLIENT_SECRET}".encode()).decode()
        r = requests.post(
            f"{RC_SERVER_URL}/restapi/oauth/token",
            headers={"Authorization": f"Basic {basic}",
                     "Content-Type": "application/x-www-form-urlencoded",
                     "Accept": "application/json"},
            data={"grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
                  "assertion": RC_JWT},
            timeout=30,
        )
        r.raise_for_status()
        d = r.json()
        _rc_token_cache["access_token"] = d.get("access_token", "")
        # Cache for up to 50 min (RC tokens last 3600s; refresh well before expiry)
        _rc_token_cache["expires_at"] = now + min(int(d.get("expires_in", 3600)) - 120, 3000)
        return _rc_token_cache["access_token"]

def _rc_provision_sip():
    """Provision a SIP device → returns sipInfo dict the WebPhone SDK consumes."""
    tok = _rc_get_access_token()
    r = requests.post(
        f"{RC_SERVER_URL}/restapi/v1.0/client-info/sip-provision",
        headers={"Authorization": f"Bearer {tok}",
                 "Content-Type": "application/json",
                 "Accept": "application/json"},
        json={"sipInfo": [{"transport": "WSS"}]},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()

@app.route("/api/rc/status")
def api_rc_status():
    """Diagnostic — returns config + auth status, never secrets."""
    if not _rc_configured():
        missing = [k for k,v in {"RC_CLIENT_ID":RC_CLIENT_ID,
                                  "RC_CLIENT_SECRET":RC_CLIENT_SECRET,
                                  "RC_JWT":RC_JWT}.items() if not v]
        return jsonify({"configured": False, "missing_env_vars": missing})
    try:
        tok = _rc_get_access_token()
        return jsonify({
            "configured": True,
            "auth_ok": bool(tok),
            "server": RC_SERVER_URL,
            "from_ext": RC_FROM_EXT,
            "from_number": RC_FROM_NUMBER,
            "caller_id": RC_CALLER_ID,
        })
    except requests.HTTPError as e:
        body = ""
        try: body = e.response.text[:300]
        except: pass
        return jsonify({"configured": True, "auth_ok": False,
                        "error": f"HTTP {e.response.status_code}", "detail": body}), 200
    except Exception as e:
        return jsonify({"configured": True, "auth_ok": False, "error": str(e)[:300]}), 200

@app.route("/api/rc/sip")
def api_rc_sip():
    """Return sipInfo so the browser's WebPhone SDK can register over WSS."""
    if not _rc_configured():
        return jsonify({"ok": False, "error": "RingCentral not configured on the server"}), 400
    # Minimal identity guard — caller must have identified themselves in the dashboard.
    caller = (request.args.get("caller","") or "").strip()
    if not caller:
        return jsonify({"ok": False,
                        "error": "Select your name in the 'Calling as:' dropdown first."}), 400
    try:
        d = _rc_provision_sip()
        sip_info = (d.get("sipInfo") or [{}])[0]
        return jsonify({
            "ok": True,
            "sipInfo": sip_info,
            "device_id": (d.get("device") or {}).get("id"),
            "caller_id": RC_CALLER_ID,
            "from_number": RC_FROM_NUMBER,
            "from_ext": RC_FROM_EXT,
        })
    except requests.HTTPError as e:
        body = ""
        try: body = e.response.text[:400]
        except: pass
        return jsonify({"ok": False,
                        "error": f"RC API error: HTTP {e.response.status_code}",
                        "detail": body}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:400]}), 500


@app.route("/api/rc/ringout", methods=["POST"])
def api_rc_ringout():
    """Place a call via RingOut REST API — bypasses WebRTC entirely.
    Flow: RC rings Muhanna's phone first → she picks up → RC dials prospect."""
    if not _rc_configured():
        return jsonify({"ok": False, "error": "RingCentral not configured on server"}), 400
    data   = request.json or {}
    to_num = str(data.get("to","")).strip()
    caller = str(data.get("caller","")).strip()
    from_override = str(data.get("from","")).strip()
    from_num = from_override or RC_FROM_NUMBER
    if not to_num: return jsonify({"ok": False, "error": "Missing destination number"}), 400
    if not caller: return jsonify({"ok": False, "error": "Select 'Calling as:' first"}), 400
    if not from_num:
        return jsonify({"ok": False, "error": "No 'from' number — set RC_FROM_NUMBER or pass one"}), 400

    try:
        tok = _rc_get_access_token()
        payload = {
            "from": {"phoneNumber": from_num},
            "to":   {"phoneNumber": to_num},
            "callerId": {"phoneNumber": RC_CALLER_ID or RC_FROM_NUMBER},
            "playPrompt": False,
        }
        r = requests.post(
            f"{RC_SERVER_URL}/restapi/v1.0/account/~/extension/~/ring-out",
            headers={"Authorization": f"Bearer {tok}",
                     "Content-Type": "application/json",
                     "Accept": "application/json"},
            json=payload,
            timeout=30,
        )
        if r.status_code >= 400:
            body = ""
            try: body = r.text[:600]
            except: pass
            return jsonify({"ok": False,
                            "error": f"RC RingOut error HTTP {r.status_code}",
                            "detail": body}), 500
        d = r.json()
        return jsonify({
            "ok": True,
            "call_id": d.get("id"),
            "status":  (d.get("status") or {}).get("callStatus"),
            "to":      to_num,
            "from":    from_num,
            "caller_id_shown": RC_CALLER_ID or from_num,
            "uri":     d.get("uri"),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:400]}), 500


@app.route("/api/rc/ringout/<call_id>")
def api_rc_ringout_status(call_id):
    """Poll RingOut call status — useful for showing live status in the UI."""
    if not _rc_configured():
        return jsonify({"ok": False, "error": "Not configured"}), 400
    try:
        tok = _rc_get_access_token()
        r = requests.get(
            f"{RC_SERVER_URL}/restapi/v1.0/account/~/extension/~/ring-out/{call_id}",
            headers={"Authorization": f"Bearer {tok}", "Accept": "application/json"},
            timeout=15,
        )
        if r.status_code >= 400:
            return jsonify({"ok": False, "error": f"HTTP {r.status_code}", "detail": r.text[:300]}), 500
        d = r.json()
        st = d.get("status")
        if not isinstance(st, dict): st = {}
        def _leg(v):
            # callerStatus / calleeStatus can be a plain string OR an object
            if isinstance(v, dict):
                return v.get("status") or v.get("reason") or v.get("description")
            return v  # already a string (e.g. "Success", "InProgress", "NoAnswer")
        return jsonify({
            "ok": True,
            "status":       st.get("callStatus"),
            "callerStatus": _leg(st.get("callerStatus")),
            "calleeStatus": _leg(st.get("calleeStatus")),
            "raw":          st,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:300]}), 500


def _rc_fetch_call_duration(to_number, max_wait=12):
    """Fetch the actual call duration (seconds) from RC call log for the most recent
    call to to_number. Polls up to max_wait seconds for the record to appear."""
    import time as _time
    if not _rc_configured() or not to_number:
        return None, "RC not configured or no number"
    norm = lambda s: "".join(ch for ch in str(s or "") if ch.isdigit())[-9:]
    want = norm(to_number)
    if not want:
        return None, "invalid number"
    deadline = _time.time() + max_wait
    while _time.time() < deadline:
        try:
            tok = _rc_get_access_token()
            from datetime import datetime as _dt, timedelta as _td
            date_from = (_dt.utcnow() - _td(hours=1)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
            r = requests.get(
                f"{RC_SERVER_URL}/restapi/v1.0/account/~/extension/~/call-log",
                headers={"Authorization": f"Bearer {tok}", "Accept": "application/json"},
                params={"type": "Voice", "direction": "Outbound",
                        "dateFrom": date_from, "perPage": 10, "view": "Detailed"},
                timeout=15)
            if r.status_code >= 400:
                return None, f"call-log HTTP {r.status_code}"
            records = r.json().get("records", []) or []
            for rec in records:
                to_ph = (rec.get("to") or {}).get("phoneNumber") or ""
                if norm(to_ph) == want:
                    dur = rec.get("duration")
                    if dur is not None and int(dur) > 0:
                        return int(dur), ""
                    # Record exists but duration is 0 — call may still be finishing
            _time.sleep(2)
        except Exception as e:
            return None, str(e)[:100]
    return None, "RC call log not ready after wait"

def _rc_find_recording(session_id, to_number):
    """Find a recorded call matching our RingOut session / prospect number.
    Returns (content_uri, rec_id, err)."""
    tok = _rc_get_access_token()
    from datetime import datetime as _dt, timedelta as _td
    date_from = (_dt.utcnow() - _td(hours=3)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    r = requests.get(
        f"{RC_SERVER_URL}/restapi/v1.0/account/~/extension/~/call-log",
        headers={"Authorization": f"Bearer {tok}", "Accept": "application/json"},
        params={"type": "Voice", "withRecording": "true", "dateFrom": date_from,
                "perPage": 100, "view": "Detailed"},
        timeout=30)
    if r.status_code >= 400:
        return None, None, f"call-log HTTP {r.status_code}: {r.text[:200]}"
    records = r.json().get("records", []) or []
    norm = lambda s: "".join(ch for ch in str(s or "") if ch.isdigit())[-9:]
    want = norm(to_number)
    best = None
    for rec in records:
        if not rec.get("recording"):
            continue
        sid = rec.get("telephonySessionId") or rec.get("sessionId") or ""
        if session_id and sid and (session_id == sid or session_id.lstrip("s-") == str(sid).lstrip("s-")):
            best = rec; break
        to = rec.get("to", {}) or {}
        if want and norm(to.get("phoneNumber")) == want:
            best = rec; break   # records are newest-first
    if not best:
        return None, None, "no recording available yet"
    recd = best.get("recording", {}) or {}
    return (recd.get("contentUri") or recd.get("uri")), recd.get("id"), ""

def _rc_download_recording(content_uri):
    tok = _rc_get_access_token()
    url = content_uri if str(content_uri).startswith("http") else f"{RC_SERVER_URL}{content_uri}"
    r = requests.get(url, headers={"Authorization": f"Bearer {tok}"}, timeout=90)
    r.raise_for_status()
    return r.content

def _openai_transcribe(audio_bytes, filename="call.mp3"):
    if not OPENAI_API_KEY:
        raise Exception("OPENAI_API_KEY not set")
    r = requests.post(
        "https://api.openai.com/v1/audio/transcriptions",
        headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
        files={"file": (filename, audio_bytes, "audio/mpeg")},
        data={"model": "whisper-1"},
        timeout=180)
    if r.status_code >= 400:
        raise Exception(f"Whisper HTTP {r.status_code}: {r.text[:200]}")
    return (r.json().get("text") or "").strip()

def _save_transcript(reg, page, caller, transcript):
    """Save a transcript to dashboard comments + Maximizer Note."""
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    text = "📝 Call transcript:\n" + transcript
    synced = 0
    try:
        found = mx_find_by_org_number(reg)
        if found and found.get("key"):
            resp = mx_create_note(found["key"], text, caller, timestamp=ts)
            if resp and resp.get("Code") == 0:
                synced = 1
    except Exception as e:
        print(f"  transcript→mx error: {e}")
    db_exec("INSERT INTO comment_history(reg_number,page,text,caller,timestamp,synced_to_maximizer) VALUES(?,?,?,?,?,?)",
            (reg, page, text, caller, ts, synced))
    print(f"  transcript saved for {reg} (mx_synced={synced})")

def _transcribe_call_async(call_id, reg, to_number, caller, page):
    """Background worker: wait for the recording to be ready, transcribe, save."""
    import time
    for attempt in range(8):           # ~ up to 7 minutes (recordings process after the call)
        time.sleep(45 if attempt else 20)
        try:
            uri, rec_id, err = _rc_find_recording(call_id, to_number)
            if not uri:
                print(f"  transcript wait {reg}: {err} (attempt {attempt+1})")
                continue
            audio = _rc_download_recording(uri)
            transcript = _openai_transcribe(audio)
            if transcript:
                _save_transcript(reg, page, caller, transcript)
            else:
                print(f"  transcript empty for {reg}")
            return
        except Exception as e:
            print(f"  transcribe attempt {attempt+1} error for {reg}: {e}")
    print(f"  transcript gave up for {reg} (no recording after retries)")

@app.route("/api/rc/transcribe", methods=["POST"])
def api_rc_transcribe():
    """Kick off (or retry) transcription for a completed RingOut call."""
    if not _rc_configured():
        return jsonify({"ok": False, "error": "RingCentral not configured"}), 400
    if not OPENAI_API_KEY:
        return jsonify({"ok": False, "error": "OPENAI_API_KEY not set"}), 400
    data = request.json or {}
    call_id = str(data.get("call_id", "")).strip()
    reg     = str(data.get("reg", "")).strip()
    to_num  = str(data.get("to", "")).strip()
    caller  = str(data.get("caller", "")).strip() or "Unknown"
    page    = str(data.get("page", "")).strip()
    if not reg:
        return jsonify({"ok": False, "error": "Missing reg"}), 400
    t = threading.Thread(target=_transcribe_call_async,
                         args=(call_id, reg, to_num, caller, page), daemon=True)
    t.start()
    return jsonify({"ok": True, "status": "started"})
RC_TEST_PAGE_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>9M · RingCentral WebPhone Test</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  *{box-sizing:border-box}
  body{font-family:-apple-system,Segoe UI,Inter,sans-serif;background:#f3f5f8;color:#1e3a5f;margin:0;padding:30px;line-height:1.5}
  .wrap{max-width:760px;margin:0 auto;background:#fff;border-radius:14px;padding:32px;box-shadow:0 4px 20px rgba(30,58,95,.10)}
  h1{margin:0 0 4px;font-size:24px}
  .sub{color:#7b8896;font-size:13px;margin-bottom:24px}
  .status-row{display:flex;align-items:center;gap:10px;margin-bottom:18px;padding:12px 16px;background:#f7f9fc;border-radius:8px;border-left:4px solid #ccc;font-size:14px}
  .status-row.ok{border-color:#2d7a3a;background:#eef9f1}
  .status-row.err{border-color:#c0392b;background:#fdecec}
  .status-row.busy{border-color:#e8b339;background:#fef8e8}
  .pill{font-size:11px;font-weight:700;padding:2px 8px;border-radius:99px;background:#e0e7ee;color:#1e3a5f;text-transform:uppercase;letter-spacing:.04em}
  .ok .pill{background:#2d7a3a;color:#fff}
  .err .pill{background:#c0392b;color:#fff}
  .busy .pill{background:#e8b339;color:#fff}
  .row{display:flex;gap:10px;margin-bottom:14px}
  input[type=tel]{flex:1;padding:12px 14px;border:1.5px solid #cfd6df;border-radius:8px;font-size:15px;font-family:inherit;color:#1e3a5f}
  input[type=tel]:focus{outline:none;border-color:#1e3a5f;box-shadow:0 0 0 3px rgba(232,179,57,.2)}
  button{padding:12px 22px;border:none;border-radius:8px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit;transition:all .12s}
  .btn-call{background:#1e3a5f;color:#fff}
  .btn-call:hover:not(:disabled){background:#152a44;transform:translateY(-1px);box-shadow:0 4px 10px rgba(30,58,95,.25)}
  .btn-hang{background:#c0392b;color:#fff}
  .btn-hang:hover:not(:disabled){background:#a02f23}
  button:disabled{opacity:.35;cursor:not-allowed;transform:none;box-shadow:none}
  .timer{font-family:'JetBrains Mono',Consolas,monospace;font-size:32px;text-align:center;color:#1e3a5f;font-weight:700;margin:20px 0;letter-spacing:.04em}
  .log{margin-top:24px;background:#0f1729;color:#a8c5e2;border-radius:8px;padding:14px 18px;height:260px;overflow-y:auto;font-family:'JetBrains Mono',Consolas,monospace;font-size:12px;line-height:1.7}
  .log .l-info{color:#a8c5e2}
  .log .l-warn{color:#e8b339}
  .log .l-err{color:#ff8a80}
  .log .l-ok{color:#69d68b}
  .log .l-evt{color:#7fb3ff}
  .log .ts{color:#5a7090;margin-right:8px}
  .footer-note{margin-top:18px;padding:10px 14px;background:#fef8e8;border:1px solid #f0d68a;border-radius:6px;font-size:12px;color:#8a6d1f}
</style>
</head>
<body>
<div class="wrap">
  <h1>📞 RingCentral WebPhone — Connection Test</h1>
  <div class="sub">Proves WebRTC + SIP works end-to-end before integration into the main dashboard.</div>

  <div id="status" class="status-row"><span class="pill" id="pill">Loading</span><span id="status-text">Initialising…</span></div>

  <div class="row" style="margin-bottom:8px">
    <label for="cidSelect" style="display:flex;align-items:center;font-size:12px;color:#7b8896;font-weight:600;padding-right:10px">Outbound caller ID:</label>
    <select id="cidSelect" style="flex:1;padding:10px;border:1.5px solid #cfd6df;border-radius:8px;font-size:13px;font-family:inherit;color:#1e3a5f">
      <option value="default">[Default — let RingCentral pick]</option>
      <option value="freephone">+44 800 208 4749 (company freephone)</option>
      <option value="direct">+44 121 818 4992 (her direct line)</option>
    </select>
  </div>
  <div class="row" style="margin-bottom:8px">
    <button id="micBtn" onclick="checkMic()" style="background:#7b8896;color:#fff;flex:1">🎤 Check microphone permission</button>
    <button id="inspectBtn" onclick="inspectWP()" style="background:#7b8896;color:#fff;flex:1">🔍 Inspect WebPhone state</button>
  </div>
  <div class="row">
    <input id="dest" type="tel" placeholder="Number to dial — e.g. +447700900123" value="">
    <button id="callBtn" class="btn-call" disabled onclick="placeCall()">📞 Call</button>
    <button id="hangBtn" class="btn-hang" disabled onclick="hangUp()">✕ Hang Up</button>
  </div>

  <div class="timer" id="timer">00:00</div>

  <div class="log" id="log"></div>

  <div class="footer-note">
    <strong>Tip:</strong> First time you click Call, the browser will ask for microphone permission — click Allow.
    Try calling your own mobile first to verify audio works both ways.
  </div>
</div>

<audio id="remoteAudio" autoplay></audio>

<script>
// ── Robust SDK loader — tries multiple CDN paths, logs what works ───────────
function loadSDK(){
  return new Promise((resolve, reject) => {
    if(typeof WebPhone !== 'undefined'){ log('WebPhone already present','l-ok'); return resolve(); }
    const urls = [
      'https://cdn.jsdelivr.net/npm/ringcentral-web-phone/dist/index.iife.js',
      'https://unpkg.com/ringcentral-web-phone/dist/index.iife.js',
      'https://cdn.jsdelivr.net/npm/ringcentral-web-phone@latest/dist/index.iife.js',
      'https://cdn.jsdelivr.net/npm/ringcentral-web-phone/dist/index.umd.js',
      'https://cdn.jsdelivr.net/npm/ringcentral-web-phone/dist/esm/index.umd.js'
    ];
    let i = 0;
    function tryNext(){
      if(i >= urls.length){ reject(new Error('all CDN URLs exhausted')); return; }
      const url = urls[i++];
      log('Trying: '+url,'l-info');
      const tag = document.createElement('script');
      tag.src = url;
      tag.onload = () => {
        // Locate the class — could be at any of several globals depending on bundle format
        const candidates = [
          window.WebPhone,
          window.RingCentralWebPhone,
          window.ringcentralWebPhone,
          (window['ringcentral-web-phone'] && window['ringcentral-web-phone'].default),
          (window.WebPhone && window.WebPhone.default),
          (window.WebPhone && window.WebPhone.WebPhone)
        ].filter(Boolean);
        if(candidates.length){
          window.WebPhone = candidates[0];
          log('SDK loaded · class found at global · '+url,'l-ok');
          resolve();
        }else{
          log('Script loaded but no WebPhone class found in window globals','l-warn');
          tryNext();
        }
      };
      tag.onerror = () => { log('Network failure for: '+url,'l-warn'); tryNext(); };
      document.head.appendChild(tag);
    }
    tryNext();
  });
}

let webPhone = null, callSession = null, callerName = '', timerInterval = null, callStart = 0;

function $(id){return document.getElementById(id);}
function ts(){return new Date().toTimeString().slice(0,8);}
function log(msg, cls){
  const line = document.createElement('div');
  line.innerHTML = `<span class="ts">${ts()}</span><span class="${cls||'l-info'}">${msg}</span>`;
  $('log').appendChild(line);
  $('log').scrollTop = $('log').scrollHeight;
}
function setStatus(txt, cls){
  $('status').className = 'status-row ' + (cls||'');
  $('pill').textContent = cls==='ok'?'Ready':cls==='busy'?'Busy':cls==='err'?'Error':'Loading';
  $('status-text').textContent = txt;
}
function tick(){
  if(!callStart) return;
  const s = Math.floor((Date.now()-callStart)/1000);
  $('timer').textContent = String(Math.floor(s/60)).padStart(2,'0')+':'+String(s%60).padStart(2,'0');
}
function startTimer(){callStart = Date.now(); timerInterval = setInterval(tick, 500);}
function stopTimer(){clearInterval(timerInterval); timerInterval=null; callStart=0; $('timer').textContent='00:00';}

async function init(){
  // 1) Read caller name from URL (?caller=Muhanna) or prompt
  const params = new URLSearchParams(window.location.search);
  callerName = (params.get('caller')||'').trim();
  if(!callerName) callerName = prompt('Your name (matches the "Calling as:" dropdown in the dashboard):','Muhanna') || '';
  if(!callerName){
    setStatus('No caller name — refresh and provide one.','err');
    log('Aborted: no caller name provided','l-err'); return;
  }
  log(`Caller identified as: ${callerName}`,'l-info');

  // 2) Verify server config
  setStatus('Checking server configuration…','busy');
  log('Fetching /api/rc/status …','l-info');
  try{
    const s = await fetch('/api/rc/status').then(r=>r.json());
    if(!s.configured){
      setStatus('Server missing RC env vars: '+(s.missing_env_vars||[]).join(', '),'err');
      log('Server not configured: '+JSON.stringify(s),'l-err'); return;
    }
    if(!s.auth_ok){
      setStatus('Server can\\'t authenticate to RingCentral.','err');
      log('Auth failure: '+(s.error||'unknown')+' — '+(s.detail||''),'l-err'); return;
    }
    log(`Server auth OK · from_ext=${s.from_ext} · caller_id=${s.caller_id}`,'l-ok');
  }catch(e){
    setStatus('Failed to reach /api/rc/status','err');
    log('Network error: '+e.message,'l-err'); return;
  }

  // 3) Fetch sipInfo
  setStatus('Provisioning SIP credentials…','busy');
  log('Fetching sipInfo from /api/rc/sip …','l-info');
  let sipResp;
  try{
    sipResp = await fetch('/api/rc/sip?caller='+encodeURIComponent(callerName)).then(r=>r.json());
  }catch(e){
    setStatus('Failed to fetch sipInfo','err'); log('Network: '+e.message,'l-err'); return;
  }
  if(!sipResp.ok){
    setStatus('SIP provisioning failed','err');
    log('Provisioning error: '+(sipResp.error||'')+' '+(sipResp.detail||''),'l-err'); return;
  }
  log('sipInfo received: domain='+(sipResp.sipInfo.domain||'?')+', transport='+(sipResp.sipInfo.transport||'?'),'l-ok');

  // 4) Load the WebPhone SDK from a CDN (try several known-good URLs)
  setStatus('Loading WebPhone SDK from CDN…','busy');
  try{
    await loadSDK();
  }catch(e){
    setStatus('Could not load WebPhone SDK from any CDN','err');
    log('All CDN sources failed: '+e.message,'l-err'); return;
  }

  setStatus('Registering with RingCentral SIP server…','busy');
  log('Initialising WebPhone instance …','l-info');

  // ── PATCH: inject STUN + TURN servers so WebRTC media can traverse NAT ───────
  // The v2 SDK only configures STUN, doesn't pick up RC's TURN credentials from
  // p-rc-ice-servers header. We wrap RTCPeerConnection so every PC the SDK creates
  // gets a full ICE config with public-TURN fallback (openrelay.metered.ca, free).
  (function(){
    const OriginalPC = window.RTCPeerConnection;
    if(!OriginalPC){ log('Browser has no RTCPeerConnection — cannot proceed','l-err'); return; }
    if(OriginalPC._9mPatched){ log('RTCPeerConnection already patched (skipping)','l-info'); return; }
    const InjectedServers = [
      {urls: 'stun:stun.l.google.com:19302'},
      {urls: 'stun:stun.cloudflare.com:3478'},
      {urls: ['turn:openrelay.metered.ca:80'],            username: 'openrelayproject', credential: 'openrelayproject'},
      {urls: ['turn:openrelay.metered.ca:443'],           username: 'openrelayproject', credential: 'openrelayproject'},
      {urls: ['turn:openrelay.metered.ca:443?transport=tcp'], username: 'openrelayproject', credential: 'openrelayproject'}
    ];
    function PatchedPC(config){
      config = config || {};
      const existing = Array.isArray(config.iceServers) ? config.iceServers : [];
      // Merge existing (SDK-provided) with our injected servers
      config.iceServers = existing.concat(InjectedServers);
      config.iceTransportPolicy = 'all';
      log('PC created · '+config.iceServers.length+' ICE servers · policy=all','l-info');
      return new OriginalPC(config);
    }
    PatchedPC.prototype = OriginalPC.prototype;
    Object.setPrototypeOf(PatchedPC, OriginalPC);
    PatchedPC._9mPatched = true;
    window.RTCPeerConnection = PatchedPC;
    log('RTCPeerConnection monkey-patched · TURN relay injected','l-ok');
  })();

  try{
    webPhone = new WebPhone({ sipInfo: sipResp.sipInfo });
    await webPhone.start();
    log('WebPhone registered successfully','l-ok');

    // Monkey-patch webPhone.emit to surface ALL events fired on the parent
    try{
      const _origEmit = webPhone.emit.bind(webPhone);
      webPhone.emit = function(ev, ...args){
        log('[webPhone] event: '+ev+(args.length?' · '+_briefArg(args[0]):''),'l-evt');
        return _origEmit(ev, ...args);
      };
      log('webPhone.emit hooked — all parent events will be logged','l-info');
    }catch(e){ log('Could not hook webPhone.emit: '+e.message,'l-warn'); }

    // Also listen for outboundCall (RC docs say this fires before await resolves)
    try{
      webPhone.on('outboundCall', (cs) => log('[webPhone] outboundCall fired · session keys: '+Object.keys(cs||{}).slice(0,12).join(','), 'l-ok'));
    }catch(e){}

    // Monkey-patch sipClient.emit (one level below webPhone) — this is where raw SIP signaling events live
    try{
      if(webPhone.sipClient && webPhone.sipClient.emit){
        const _origSip = webPhone.sipClient.emit.bind(webPhone.sipClient);
        webPhone.sipClient.emit = function(ev){
          const args = Array.prototype.slice.call(arguments, 1);
          log('[sipClient] '+ev+(args.length?' · '+_briefArg(args[0]):''),'l-evt');
          return _origSip.apply(webPhone.sipClient, arguments);
        };
        log('sipClient.emit hooked — raw SIP events will be logged','l-info');
        if(webPhone.sipClient.eventNames) log('sipClient event names registered: '+webPhone.sipClient.eventNames().join(',')||'(none)','l-info');
      }else{
        log('webPhone has no sipClient.emit to hook','l-warn');
      }
    }catch(e){ log('Could not hook sipClient: '+e.message,'l-warn'); }
    setStatus('Ready to make calls — enter a number above','ok');
    $('callBtn').disabled = false;
  }catch(e){
    setStatus('SIP registration failed','err');
    log('WebPhone error: '+(e.message||e),'l-err'); return;
  }

  // Save caller_id for outbound calls
  window._callerId = sipResp.caller_id || sipResp.from_number || '';
  log('Outbound caller_id will be: '+window._callerId,'l-info');
}

function _sipBriefLine(msg){
  // Extract first useful line of a SIP message for logging
  try{
    const txt = (typeof msg === 'string') ? msg : (msg && msg.toString ? msg.toString() : JSON.stringify(msg));
    const lines = txt.split(/\\r?\\n/).filter(Boolean);
    for(const ln of lines){
      if(/^SIP\/2\.0\s+\d{3}/.test(ln) || /^(INVITE|ACK|BYE|CANCEL|REGISTER|SUBSCRIBE|NOTIFY|OPTIONS|REFER|UPDATE|INFO)\s/.test(ln)) return ln.substring(0,160);
    }
    return (lines[0]||txt).substring(0,160);
  }catch(_){ return '?'; }
}

async function placeCall(){
  const dest = $('dest').value.trim();
  if(!dest){alert('Enter a phone number first'); return;}
  if(!webPhone){alert('WebPhone not ready'); return;}

  // Pre-flight: confirm mic permission BEFORE attempting call (silent mic = silent failure)
  log('Pre-flight: requesting microphone access…','l-info');
  try{
    const ms = await navigator.mediaDevices.getUserMedia({audio:true, video:false});
    log('Mic stream OK · tracks='+ms.getAudioTracks().length+' · device='+(ms.getAudioTracks()[0]||{}).label,'l-ok');
    ms.getTracks().forEach(t=>t.stop()); // release immediately; SDK will request again
  }catch(e){
    log('MIC ERROR: '+e.name+' — '+e.message,'l-err');
    setStatus('Microphone unavailable','err');
    return;
  }

  const cidChoice = $('cidSelect').value;
  let callerId = '';
  if(cidChoice === 'freephone') callerId = '+448002084749';
  else if(cidChoice === 'direct') callerId = '+441218184992';

  setStatus('Calling '+dest+' …','busy');
  log('Placing call to '+dest+' · callerId='+(callerId||'[default]'),'l-evt');
  $('callBtn').disabled = true;

  try{
    callSession = await webPhone.call(dest, callerId);
    log('webPhone.call() returned · session.id='+(callSession.callId||'?'),'l-info');
    log('Session keys: '+Object.keys(callSession).slice(0,20).join(','),'l-info');

    // Monkey-patch session.emit so we see EVERY event the session fires
    try{
      const _origEmit = callSession.emit.bind(callSession);
      callSession.emit = function(ev, ...args){
        log('[session] event: '+ev+(args.length?' · '+_briefArg(args[0]):''),'l-evt');
        return _origEmit(ev, ...args);
      };
      log('session.emit hooked — all events will be logged','l-info');
    }catch(e){ log('Could not hook session.emit: '+e.message,'l-warn'); }

    // Monitor the WebRTC peer connection — if ICE fails, calls die silently like this
    try{
      const pc = callSession.rtcPeerConnection;
      if(pc){
        log('RTC initial · conn='+pc.connectionState+' ice='+pc.iceConnectionState+' gather='+pc.iceGatheringState+' sig='+pc.signalingState,'l-info');
        pc.addEventListener('connectionstatechange', function(){ log('RTC connection → '+pc.connectionState, pc.connectionState==='failed'?'l-err':'l-evt'); });
        pc.addEventListener('iceconnectionstatechange', function(){ log('RTC ICE → '+pc.iceConnectionState, pc.iceConnectionState==='failed'?'l-err':'l-evt'); });
        pc.addEventListener('icegatheringstatechange', function(){ log('RTC gather → '+pc.iceGatheringState,'l-evt'); });
        pc.addEventListener('signalingstatechange', function(){ log('RTC signaling → '+pc.signalingState,'l-evt'); });
        let cands=0;
        pc.addEventListener('icecandidate', function(e){
          if(e.candidate){ cands++;
            const c = e.candidate.candidate || '';
            // Log only first few full candidates; then summarize
            if(cands<=4) log('ICE candidate #'+cands+': '+c.substring(0,140),'l-evt');
            else if(cands===5) log('ICE candidate #5+ (further candidates suppressed in log)','l-evt');
          }else log('ICE gathering DONE — total '+cands+' candidates','l-ok');
        });
        pc.addEventListener('icecandidateerror', function(e){
          log('ICE candidate ERROR: url='+(e.url||'?')+' code='+(e.errorCode||'?')+' text='+(e.errorText||'?'),'l-err');
        });

        // Log the configured ICE servers (TURN/STUN) — if empty, that's the bug
        try{
          const cfg = pc.getConfiguration();
          const servers = cfg.iceServers||[];
          log('RTCPeerConnection has '+servers.length+' ICE servers configured','l-info');
          servers.forEach(function(srv, idx){
            const urls = Array.isArray(srv.urls)?srv.urls:[srv.urls];
            urls.forEach(function(u){ log('  ICE server #'+idx+': '+u,'l-info'); });
          });
          if(servers.length===0) log('WARNING: No ICE servers — WebRTC cannot traverse NAT','l-err');
        }catch(e){ log('Could not read RTC config: '+e.message,'l-warn'); }

        // Examine the local SDP — count ICE candidates by type (host/srflx/relay)
        setTimeout(function(){
          try{
            const sdp = pc.localDescription && pc.localDescription.sdp;
            if(sdp){
              const lines = sdp.split(String.fromCharCode(10));
              const candLines = lines.filter(function(l){return l.indexOf('a=candidate:')===0;});
              log('Local SDP has '+candLines.length+' ICE candidates','l-info');
              const counts = {host:0, srflx:0, prflx:0, relay:0, other:0};
              candLines.forEach(function(l){
                const m = l.match(/typ\s+(\w+)/);
                if(m) counts[m[1]] = (counts[m[1]]||0) + 1; else counts.other++;
              });
              log('  Candidate types: host='+counts.host+'  srflx='+counts.srflx+'  prflx='+counts.prflx+'  relay='+counts.relay+'  other='+counts.other,'l-info');
              if(counts.relay===0 && counts.srflx===0) log('  → No public-reachable candidates (only local host). Media WILL fail through NAT.','l-err');
            }
          }catch(e){ log('Local SDP inspect failed: '+e.message,'l-warn'); }
        }, 500);

        // Poll RTP stats every 5s for 30s — detect whether any audio is flowing
        let statsPolls = 0;
        const statsInterval = setInterval(function(){
          statsPolls++;
          if(statsPolls > 6 || (callSession && callSession.state === 'disposed')){ clearInterval(statsInterval); return; }
          try{
            pc.getStats().then(function(report){
              let sent = 0, recv = 0, selectedPair = null;
              report.forEach(function(s){
                if(s.type==='outbound-rtp' && s.kind==='audio') sent = s.packetsSent || 0;
                if(s.type==='inbound-rtp'  && s.kind==='audio') recv = s.packetsReceived || 0;
                if(s.type==='candidate-pair' && s.nominated && s.state==='succeeded') selectedPair = s;
              });
              const pairInfo = selectedPair ? (' · pair=succeeded') : ' · NO SUCCEEDED CANDIDATE PAIR';
              log('Stats T+'+(statsPolls*5)+'s · RTP sent='+sent+' recv='+recv+pairInfo, sent>0&&recv>0?'l-ok':'l-warn');
            });
          }catch(_){}
        }, 5000);
      }else{
        log('No rtcPeerConnection on session — WebRTC not engaged','l-warn');
      }
    }catch(e){ log('RTC hook error: '+e.message,'l-err'); }

    // Print the outgoing SIP INVITE (the message the SDK built for this call)
    try{
      if(callSession.sipMessage){
        let txt = '';
        try{ txt = String(callSession.sipMessage); }catch(_){}
        if(txt){
          log('Outgoing SIP INVITE — first 14 lines:','l-info');
          const NL = String.fromCharCode(10);
          txt.split(NL).slice(0,14).forEach(function(ln){
            log('  ' + ln.substring(0, 200), 'l-info');
          });
        }else log('sipMessage present but could not stringify','l-warn');
      }else log('No sipMessage on session — INVITE may not have been constructed','l-warn');
    }catch(e){ log('SIP dump error: '+e.message,'l-warn'); }

    // Normal listeners (in case hook missed anything)
    ['answered','ringing','disposed','inboundMessage','outboundMessage','accepted','rejected','failed','progress','bye','canceled','terminated'].forEach(ev=>{
      try{ callSession.on(ev, (p) => log('  → on('+ev+') fired'+(p?' · '+_briefArg(p):''),'l-ok')); }catch(_){}
    });

    callSession.on('disposed', () => {
      setStatus('Ready to make calls','ok');
      $('callBtn').disabled = false; $('hangBtn').disabled = true;
      // keep callSession reference for inspection
      stopTimer();
    });
    callSession.on('answered', () => {
      setStatus('In call with '+dest,'ok');
      $('hangBtn').disabled = false;
      startTimer();
    });
  }catch(e){
    log('webPhone.call() threw: '+(e.message||e),'l-err');
    setStatus('Call failed — see log','err');
    $('callBtn').disabled = false; $('hangBtn').disabled = true;
  }
}

function _briefArg(a){
  try{
    if(a==null) return 'null';
    if(typeof a==='string') return a.length>180 ? a.substring(0,180)+'…' : a;
    if(typeof a==='object'){
      const keys = Object.keys(a).slice(0,8);
      // If it looks like a SIP message
      if(a.toString && a.toString.toString().indexOf('[native code]')<0){
        const t = String(a);
        if(t.startsWith('SIP') || /^[A-Z]+\s.*SIP/.test(t)) return t.split(/\\r?\\n/)[0].substring(0,160);
      }
      return '{' + keys.join(',') + '}';
    }
    return String(a).substring(0,180);
  }catch(_){ return '?'; }
}

async function checkMic(){
  log('Manual mic check…','l-info');
  try{
    const ms = await navigator.mediaDevices.getUserMedia({audio:true});
    const t = ms.getAudioTracks()[0];
    log('Mic OK · device="'+(t&&t.label)+'" · enabled='+(t&&t.enabled),'l-ok');
    ms.getTracks().forEach(t=>t.stop());
  }catch(e){
    log('Mic FAIL · '+e.name+' · '+e.message,'l-err');
  }
}

function inspectWP(){
  if(!webPhone){log('webPhone not initialised','l-warn'); return;}
  try{
    log('webPhone keys: '+Object.keys(webPhone).slice(0,30).join(','),'l-info');
    if(webPhone.eventNames) log('webPhone event names: '+webPhone.eventNames().join(','),'l-info');
    if(webPhone.userAgent) log('userAgent state: '+(webPhone.userAgent.state||'?'),'l-info');
  }catch(e){ log('inspect error: '+e.message,'l-err'); }
  if(callSession){
    log('callSession keys: '+Object.keys(callSession).slice(0,30).join(','),'l-info');
    if(callSession.state !== undefined) log('callSession.state: '+callSession.state,'l-info');
    if(callSession.status !== undefined) log('callSession.status: '+callSession.status,'l-info');
  }
}

async function hangUp(){
  if(!callSession) return;
  log('Hanging up…','l-evt');
  try{ await callSession.dispose(); }catch(e){ log('Hangup error: '+e.message,'l-err'); }
}

window.addEventListener('load', init);
</script>
</body>
</html>
"""

@app.route("/rc-test")
def rc_test_page():
    """Standalone test page — verifies WebRTC + SIP end-to-end before main integration."""
    from flask import render_template_string
    return render_template_string(RC_TEST_PAGE_HTML)


# ── RingOut (PSTN-only) test page — no WebRTC, works through any firewall ─────
RC_RINGOUT_PAGE_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>9M · RingCentral RingOut Test</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  *{box-sizing:border-box}
  body{font-family:-apple-system,Segoe UI,Inter,sans-serif;background:#f3f5f8;color:#1e3a5f;margin:0;padding:30px;line-height:1.5}
  .wrap{max-width:680px;margin:0 auto;background:#fff;border-radius:14px;padding:32px;box-shadow:0 4px 20px rgba(30,58,95,.10)}
  h1{margin:0 0 4px;font-size:24px}
  .sub{color:#7b8896;font-size:13px;margin-bottom:20px}
  .info{background:#eef9f1;border:1px solid #9fd1ad;border-radius:8px;padding:12px 14px;margin-bottom:18px;font-size:13px;line-height:1.55}
  .info strong{color:#1d7a3a}
  .row{display:flex;gap:10px;margin-bottom:14px;align-items:center}
  label{font-size:12px;font-weight:700;color:#7b8896;text-transform:uppercase;letter-spacing:.04em;min-width:130px}
  input[type=tel]{flex:1;padding:12px 14px;border:1.5px solid #cfd6df;border-radius:8px;font-size:15px;font-family:inherit;color:#1e3a5f}
  input[type=tel]:focus{outline:none;border-color:#1e3a5f;box-shadow:0 0 0 3px rgba(232,179,57,.2)}
  button{padding:14px 28px;border:none;border-radius:8px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit;transition:all .15s}
  .btn-call{background:#1e3a5f;color:#fff;width:100%;font-size:16px;padding:16px}
  .btn-call:hover:not(:disabled){background:#152a44;box-shadow:0 4px 12px rgba(30,58,95,.30)}
  button:disabled{opacity:.4;cursor:not-allowed}
  .status{margin:18px 0;padding:14px 18px;border-radius:8px;border-left:4px solid #ccc;font-size:14px;background:#f7f9fc}
  .status.ok{border-color:#1d7a3a;background:#eef9f1}
  .status.err{border-color:#c0392b;background:#fdecec}
  .status.busy{border-color:#e8b339;background:#fef8e8}
  .pill{display:inline-block;font-size:10px;font-weight:800;padding:2px 8px;border-radius:99px;background:#e0e7ee;color:#1e3a5f;text-transform:uppercase;letter-spacing:.04em;margin-right:8px}
  .ok .pill{background:#1d7a3a;color:#fff}
  .err .pill{background:#c0392b;color:#fff}
  .busy .pill{background:#e8b339;color:#fff}
  .log{margin-top:18px;background:#0f1729;color:#a8c5e2;border-radius:8px;padding:14px;height:280px;overflow-y:auto;font-family:'JetBrains Mono',Consolas,monospace;font-size:12px;line-height:1.7}
  .log .l-ok{color:#69d68b}
  .log .l-warn{color:#e8b339}
  .log .l-err{color:#ff8a80}
  .log .l-evt{color:#7fb3ff}
  .log .ts{color:#5a7090;margin-right:8px}
  .meta{font-size:11px;color:#7b8896;margin-top:10px}
</style>
</head>
<body>
<div class="wrap">
  <h1>📞 RingCentral RingOut — PSTN Test</h1>
  <div class="sub">Bypasses WebRTC entirely. Your phone rings first, then RC dials the prospect.</div>

  <div class="info">
    <strong>How this works:</strong> Click Call → RingCentral platform rings your phone (whichever device is registered to <code>RC_FROM_NUMBER</code>) → you pick up → RC dials the prospect → you're connected. Audio flows through RC's network, not your browser. Works through any firewall.
  </div>

  <div class="row">
    <label for="caller">Calling as</label>
    <input id="caller" type="text" value="Muhanna" style="flex:1">
  </div>
  <div class="row">
    <label for="fromNum">Ring my phone</label>
    <input id="fromNum" type="tel" placeholder="Your MOBILE, e.g. +48xxxxxxxxx (NOT a RingCentral number)">
  </div>
  <div class="row">
    <label for="dest">Then dial</label>
    <input id="dest" type="tel" placeholder="+447920116516 (the prospect)">
  </div>
  <button class="btn-call" id="callBtn" onclick="ringOut()">📞 Place RingOut call</button>

  <div class="status" id="status"><span class="pill" id="pill">Ready</span><span id="status-text">Enter a number and click Call</span></div>

  <div class="log" id="log"></div>

  <div class="meta">Server endpoint: <code>POST /api/rc/ringout</code> · Poll: <code>GET /api/rc/ringout/{id}</code></div>
</div>

<script>
function $(id){return document.getElementById(id);}
function ts(){return new Date().toTimeString().slice(0,8);}
function log(msg, cls){
  const line = document.createElement('div');
  line.innerHTML = '<span class="ts">'+ts()+'</span><span class="'+(cls||'')+'">'+msg+'</span>';
  $('log').appendChild(line); $('log').scrollTop = $('log').scrollHeight;
}
function setStatus(txt, cls){
  $('status').className = 'status ' + (cls||'');
  $('pill').textContent = cls==='ok'?'Connected':cls==='busy'?'Calling':cls==='err'?'Error':'Ready';
  $('status-text').textContent = txt;
}

async function ringOut(){
  const dest = $('dest').value.trim();
  const caller = $('caller').value.trim();
  const fromNum = $('fromNum').value.trim();
  if(!dest){alert('Enter the prospect number to dial'); return;}
  if(!fromNum){alert('Enter YOUR phone number (the one that should ring first)'); return;}
  if(!caller){alert('Enter your name first'); return;}

  $('callBtn').disabled = true;
  setStatus('Sending request to RingCentral platform…','busy');
  log('POST /api/rc/ringout from='+fromNum+' to='+dest+' caller='+caller, 'l-evt');

  try{
    const r = await fetch('/api/rc/ringout', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({to: dest, from: fromNum, caller: caller})
    });
    const d = await r.json();
    if(!d.ok){
      log('Request failed: '+(d.error||'')+' '+(d.detail||''),'l-err');
      setStatus('RingCentral rejected the call','err');
      $('callBtn').disabled = false;
      return;
    }
    log('Call created · id='+d.call_id+' · status='+d.status, 'l-ok');
    log('  · Your phone ('+d.from+') will ring first', 'l-ok');
    log('  · Once you answer, RC will dial '+d.to, 'l-ok');
    log('  · Prospect sees caller ID: '+d.caller_id_shown, 'l-ok');
    setStatus('Calling your phone now — pick up to connect','busy');

    // Poll status every 4 seconds for up to 2 minutes
    let polls = 0, last = '';
    const interval = setInterval(async function(){
      polls++;
      if(polls > 30){
        log('Status polling timed out after 2 minutes','l-warn');
        setStatus('Polling stopped (call may still be active)','err');
        clearInterval(interval); $('callBtn').disabled = false; return;
      }
      try{
        const s = await fetch('/api/rc/ringout/'+d.call_id).then(r=>r.json());
        if(s.ok){
          // Build a rich status line
          let line = 'Status: '+(s.status||'?');
          if(s.callerStatus) line += '  ·  your-phone='+s.callerStatus;
          if(s.calleeStatus) line += '  ·  prospect='+s.calleeStatus;
          if(line !== last){
            last = line;
            const cls = s.status==='Success' ? 'l-ok' : s.status==='InProgress' ? 'l-evt' : 'l-err';
            log(line, cls);
            if(s.status==='Success'){
              setStatus('Connected — talk!','ok');
            } else if(s.status && s.status !== 'InProgress'){
              // Terminal state of any kind — re-enable button
              setStatus('Call ended: '+s.status,'err');
              if(s.raw){
                try{ log('Raw status: '+JSON.stringify(s.raw).substring(0,400),'l-warn'); }catch(_){}
              }
              $('callBtn').disabled = false;
              clearInterval(interval);
            }
          }
        }else{
          log('Poll error: '+(s.error||'?'),'l-err');
        }
      }catch(e){ log('Poll error: '+e.message,'l-warn'); }
    }, 4000);
  }catch(e){
    log('Network error: '+e.message,'l-err');
    setStatus('Could not reach server','err');
    $('callBtn').disabled = false;
  }
}

// On load, populate caller from query string if present
window.addEventListener('load', function(){
  const params = new URLSearchParams(window.location.search);
  const c = params.get('caller');
  if(c) $('caller').value = c;
  log('Page ready · click Call to test','l-ok');
});
</script>
</body>
</html>
"""

@app.route("/rc-ringout")
def rc_ringout_page():
    """RingOut test page — works through any firewall, no browser audio."""
    from flask import render_template_string
    return render_template_string(RC_RINGOUT_PAGE_HTML)
